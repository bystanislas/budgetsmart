"""Modules métier d'APEX Budget : Mariage, Immobilier & Terrain, Business, Objectifs, Dettes."""
from __future__ import annotations

from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import apex_data as D
from apex_style import (
    ChartZone, nav_bar,
    BG, BLUE, CARD, FMT_DATE, FMT_INT, FMT_MONEY, FMT_PCT, GOLD, GREEN, INK,
    LINE, MUTED, NAVY, NAVY_MID, NAVY_SOFT, ORANGE, PURPLE, RED, SERIES, TEAL,
    WHITE, al, banner, box_border, color_scale_ok, currency_format, databar, f,
    hide_zeros, input_style, kpi_card, note_box, paint, section, setup_sheet,
    percent_labels, solid, tune_chart,
    table_header, widths, zebra,
)
from sheets_core import REAL

# bloc « sorties par module et par mois » du moteur de calcul
CALC_MOD_HEAD = 21
CALC_MOD_FIRST = 22

OUT = 'J_Type,"<>Revenu",J_Type,"<>Transfert"'


def _color(serie, hexcolor):
    serie.graphicalProperties = GraphicalProperties(solidFill=hexcolor)
    serie.graphicalProperties.line.solidFill = hexcolor


def _field(ws, row, label, value=None, formula=None, numfmt=None, editable=True,
           lab_cols=("C", "D"), val_cols=("E", "F"), help_text="", help_cols=("G", "J")):
    ws.merge_cells(f"{lab_cols[0]}{row}:{lab_cols[1]}{row}")
    c = ws[f"{lab_cols[0]}{row}"]
    c.value = label
    c.font = f(10, True, INK)
    c.alignment = al("left", "center", indent=1)
    ws.merge_cells(f"{val_cols[0]}{row}:{val_cols[1]}{row}")
    v = ws[f"{val_cols[0]}{row}"]
    v.value = formula if formula is not None else value
    v.number_format = numfmt or "General"
    v.alignment = al("left", "center", indent=1)
    if editable:
        input_style(ws, f"{val_cols[0]}{row}:{val_cols[1]}{row}")
        v.font = f(11, True, INK)
    else:
        paint(ws, f"{val_cols[0]}{row}:{val_cols[1]}{row}", fill=solid(BG),
              border=box_border(LINE))
        v.font = f(11, True, BLUE)
    if help_text:
        ws.merge_cells(f"{help_cols[0]}{row}:{help_cols[1]}{row}")
        h = ws[f"{help_cols[0]}{row}"]
        h.value = help_text
        h.font = f(9, False, MUTED, italic=True)
        h.alignment = al("left", "center", wrap=True, indent=1)
    ws.row_dimensions[row].height = 22
    return v


def module_month_chart(wb, ws, module, zone, color):
    """Barres des sorties du module, mois par mois, lues dans le moteur de calcul."""
    calc = wb["Calculs"]
    col = 3 + D.MODULES.index(module)
    ch = BarChart()
    ch.type = "col"
    ch.title = f"Sorties {module} — mois par mois"
    ch.add_data(Reference(calc, min_col=col, min_row=CALC_MOD_HEAD,
                          max_row=CALC_MOD_FIRST + 11), titles_from_data=True)
    ch.set_categories(Reference(calc, min_col=2, min_row=CALC_MOD_FIRST,
                                max_row=CALC_MOD_FIRST + 11))
    _color(ch.series[0], color)
    ch.height, ch.width = 8.0, 17.0
    ch.gapWidth = 60
    tune_chart(ch, axis_size=750, no_legend=True)
    zone.add(ch)
    return ch


# ==========================================================================
# MARIAGE
# ==========================================================================
MAR_FIRST = 20
MAR_LAST = 49


def build_mariage(wb):
    ws = wb.create_sheet("Mariage")
    setup_sheet(ws, tab_color="9C3A5F", freeze="A19", zoom=85)
    nav_bar(ws, "P", row=3)
    banner(ws, "P", "BUDGET MARIAGE",
           "Estimations, devis comparés, paiements réels et plan de financement")
    widths(ws, {"A": 2, "B": 5, "C": 30, "D": 22, "E": 15, "F": 14, "G": 14, "H": 14,
                "I": 13, "J": 15, "K": 15, "L": 15, "M": 11, "N": 13, "O": 14, "P": 26})

    section(ws, 4, "C", "P", "①  LE MARIAGE")
    _field(ws, 6, "Les mariés", value="Prénom & Prénom", help_cols=("G", "P"))
    _field(ws, 7, "Date du mariage", value=None, numfmt=FMT_DATE,
           help_text="Alimente le compte à rebours ci-dessous.", help_cols=("G", "P"))
    _field(ws, 8, "Lieu de la réception", value="", help_cols=("G", "P"))
    _field(ws, 9, "Nombre d'invités", value=100, numfmt=FMT_INT, help_cols=("G", "P"))
    _field(ws, 10, "Budget cible", value=0, numfmt=FMT_MONEY,
           help_text="Le montant que vous ne souhaitez pas dépasser.", help_cols=("G", "P"))
    _field(ws, 11, "Jours restants",
           formula='=IF($E$7="","—",MAX(0,$E$7-TODAY()))', numfmt=FMT_INT,
           editable=False, help_cols=("G", "P"))
    currency_format(ws, "E10:F10")

    T = f"$J${MAR_FIRST}:$J${MAR_LAST}"
    kpi_card(ws, 13, 2, "COÛT ESTIMÉ", f"=SUM($E${MAR_FIRST}:$E${MAR_LAST})",
             FMT_MONEY, MUTED, span=3, note="Somme des estimations initiales")
    kpi_card(ws, 13, 5, "COÛT RETENU", f"=SUM({T})", FMT_MONEY, NAVY, span=3,
             note="Devis retenus ou estimations")
    kpi_card(ws, 13, 8, "DÉJÀ PAYÉ", f"=SUM($K${MAR_FIRST}:$K${MAR_LAST})",
             FMT_MONEY, GREEN, span=3, note_formula=f'="Soit "&TEXT(IFERROR(SUM($K${MAR_FIRST}'
             f':$K${MAR_LAST})/SUM({T}),0),"0%")&" du coût retenu"')
    kpi_card(ws, 13, 11, "RESTE À PAYER",
             f"=SUM({T})-SUM($K${MAR_FIRST}:$K${MAR_LAST})", FMT_MONEY, ORANGE, span=3)
    kpi_card(ws, 13, 14, "COÛT PAR INVITÉ",
             f'=IFERROR(SUM({T})/$E$9,0)', FMT_MONEY, PURPLE, span=3,
             note_formula=f'=IF($E$10=0,"Aucun budget cible saisi",'
                          f'IF(SUM({T})>$E$10,"⚠ Budget cible dépassé de "&'
                          f'TEXT(SUM({T})-$E$10,"#,##0"),"✓ Dans le budget cible"))')
    for start, end in (("B", "D"), ("E", "G"), ("H", "J"), ("K", "M"), ("N", "P")):
        currency_format(ws, f"{start}14:{end}14")

    section(ws, 17, "B", "P", "②  POSTES DE DÉPENSE — ESTIMATION, DEVIS, RÉALISÉ")
    table_header(ws, 19, "B", "P",
                 ["N°", "Poste", "Prestataire retenu", "Estimation", "Devis 1",
                  "Devis 2", "Devis 3", "Devis retenu", "Coût retenu", "Déjà payé",
                  "Reste à payer", "% payé", "Échéance", "Statut", "Notes"], height=34)
    postes = D.CAT_MARIAGE
    for i in range(MAR_LAST - MAR_FIRST + 1):
        r = MAR_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        if i < len(postes):
            ws.cell(row=r, column=3, value=postes[i])
        for col in (5, 6, 7, 8):
            ws.cell(row=r, column=col).number_format = FMT_MONEY
        ws.cell(row=r, column=10,
                value=(f'=IF($I{r}="Devis 1",$F{r},IF($I{r}="Devis 2",$G{r},'
                       f'IF($I{r}="Devis 3",$H{r},N($E{r}))))')).number_format = FMT_MONEY
        ws.cell(row=r, column=11,
                value=(f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Module,"Mariage",'
                       f'J_Cat,$C{r},{OUT},{REAL}))')).number_format = FMT_MONEY
        ws.cell(row=r, column=12, value=f"=N($J{r})-N($K{r})").number_format = FMT_MONEY
        ws.cell(row=r, column=13, value=f"=IFERROR($K{r}/$J{r},0)").number_format = FMT_PCT
        ws.cell(row=r, column=14).number_format = FMT_DATE
        ws.cell(row=r, column=15,
                value=(f'=IF($C{r}="","",IF($J{r}=0,"À chiffrer",IF($M{r}>=1,"✓ Soldé",'
                       f'IF($M{r}>0,"Acompte versé",IF(AND($N{r}<>"",$N{r}<TODAY()),'
                       f'"⚠ Échéance passée","À payer")))))'))
        ws.cell(row=r, column=15).alignment = al("center")
        ws.cell(row=r, column=15).font = f(9, True)
    r = MAR_LAST + 1
    ws.cell(row=r, column=3, value="TOTAL").font = f(11, True, WHITE)
    for col, letter in ((5, "E"), (6, "F"), (7, "G"), (8, "H"), (10, "J"),
                        (11, "K"), (12, "L")):
        ws.cell(row=r, column=col,
                value=f"=SUM({letter}{MAR_FIRST}:{letter}{MAR_LAST})"
                ).number_format = FMT_MONEY
    ws.cell(row=r, column=13, value=f"=IFERROR($K{r}/$J{r},0)").number_format = FMT_PCT
    paint(ws, f"B{MAR_FIRST}:P{r}", border=box_border(LINE), font=f(10))
    paint(ws, f"B{r}:P{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    input_style(ws, f"C{MAR_FIRST}:I{MAR_LAST}")
    input_style(ws, f"N{MAR_FIRST}:N{MAR_LAST}")
    input_style(ws, f"P{MAR_FIRST}:P{MAR_LAST}")
    paint(ws, f"J{MAR_FIRST}:M{MAR_LAST}", fill=solid(BG))
    zebra(ws, f"B{MAR_FIRST}:P{MAR_LAST}", MAR_FIRST)
    databar(ws, f"M{MAR_FIRST}:M{MAR_LAST}", "9C3A5F")
    hide_zeros(ws, f"E{MAR_FIRST}:L{MAR_LAST}")
    currency_format(ws, f"E{MAR_FIRST}:L{MAR_LAST}", hide_zero=True)
    currency_format(ws, f"E{r}:L{r}")

    dv = DataValidation(type="list", formula1='"Estimation,Devis 1,Devis 2,Devis 3"',
                        allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv)
    dv.add(f"I{MAR_FIRST}:I{MAR_LAST}")
    dvc = DataValidation(type="list", formula1="Cat_2", allow_blank=True,
                         showErrorMessage=False)
    ws.add_data_validation(dvc)
    dvc.add(f"C{MAR_FIRST}:C{MAR_LAST}")

    # --- Plan de financement
    section(ws, 53, "B", "H", "③  PLAN DE FINANCEMENT")
    table_header(ws, 54, "B", "E", ["", "Source de financement", "Montant", "Part"])
    sources = ["Épargne dédiée du couple", "Apport du marié", "Apport de la mariée",
               "Famille du marié", "Famille de la mariée", "Cotisations & dons",
               "Crédit / prêt", "Autre source"]
    for i, srcname in enumerate(sources):
        r = 55 + i
        ws[f"C{r}"] = srcname
        ws[f"D{r}"] = 0
        ws[f"D{r}"].number_format = FMT_MONEY
        ws[f"E{r}"] = f"=IFERROR($D{r}/SUM($D$55:$D$62),0)"
        ws[f"E{r}"].number_format = FMT_PCT
    ws["C63"] = "TOTAL FINANCEMENT"
    ws["D63"] = "=SUM(D55:D62)"
    ws["D63"].number_format = FMT_MONEY
    ws["C64"] = "COUVERTURE DU COÛT RETENU"
    ws["D64"] = f"=IFERROR($D$63/SUM({T}),0)"
    ws["D64"].number_format = FMT_PCT
    ws["C65"] = "RESTE À FINANCER"
    ws["D65"] = f"=MAX(0,SUM({T})-$D$63)"
    ws["D65"].number_format = FMT_MONEY
    ws["C66"] = "EFFORT D'ÉPARGNE MENSUEL NÉCESSAIRE"
    ws["D66"] = ('=IF(OR($E$7="",$E$7<=TODAY()),$D$65,'
                 'IFERROR($D$65/MAX(1,ROUND(($E$7-TODAY())/30.4,0)),0))')
    ws["D66"].number_format = FMT_MONEY
    paint(ws, "B55:E66", border=box_border(LINE))
    input_style(ws, "C55:D62")
    paint(ws, "C63:E66", fill=solid(BG), font=f(10, True, NAVY))
    zebra(ws, "B55:E62", 55)
    currency_format(ws, "D55:D63")
    currency_format(ws, "D65:D66")

    # --- familles de postes : six parts lisibles plutôt que trente
    FAM = 70
    section(ws, 68, "B", "H", "④  RÉPARTITION PAR FAMILLE DE POSTES")
    table_header(ws, 69, "C", "F",
                 ["Famille de postes", "Coût retenu", "Déjà payé", "Part"])
    for i, (nom, idx) in enumerate(D.FAMILLES_MARIAGE):
        r = FAM + i
        ws[f"C{r}"] = nom
        ws[f"D{r}"] = "=" + "+".join(f"N($J${MAR_FIRST + k})" for k in idx)
        ws[f"E{r}"] = "=" + "+".join(f"N($K${MAR_FIRST + k})" for k in idx)
        ws[f"F{r}"] = f"=IFERROR($D{r}/SUM($D${FAM}:$D${FAM + 5}),0)"
        ws[f"D{r}"].number_format = FMT_MONEY
        ws[f"E{r}"].number_format = FMT_MONEY
        ws[f"F{r}"].number_format = FMT_PCT
    rt = FAM + 6
    ws[f"C{rt}"] = "TOTAL"
    ws[f"D{rt}"] = f"=SUM($D${FAM}:$D${FAM + 5})"
    ws[f"E{rt}"] = f"=SUM($E${FAM}:$E${FAM + 5})"
    ws[f"D{rt}"].number_format = FMT_MONEY
    ws[f"E{rt}"].number_format = FMT_MONEY
    paint(ws, f"C{FAM}:F{rt}", border=box_border(LINE), font=f(10))
    paint(ws, f"C{rt}:F{rt}", fill=solid(NAVY), font=f(10, True, WHITE))
    zebra(ws, f"C{FAM}:F{FAM + 5}", FAM)
    databar(ws, f"F{FAM}:F{FAM + 5}", "9C3A5F")
    currency_format(ws, f"D{FAM}:E{rt}", hide_zero=True)

    ch1 = DoughnutChart()
    ch1.title = "Répartition du coût retenu par famille"
    ch1.add_data(Reference(ws, min_col=4, min_row=FAM, max_row=FAM + 5),
                 titles_from_data=False)
    ch1.set_categories(Reference(ws, min_col=3, min_row=FAM, max_row=FAM + 5))
    ch1.dataLabels = percent_labels()
    ch1.holeSize = 55
    ch1.height, ch1.width = 10.0, 15.0
    tune_chart(ch1, axis_size=800)
    zm = ChartZone(ws, "B", 80)
    zm2 = ChartZone(ws, "J", 80)
    zm.add(ch1)

    ch2 = BarChart()
    ch2.type = "bar"
    ch2.title = "Coût retenu vs déjà payé (par poste)"
    d2 = Reference(ws, min_col=10, max_col=11, min_row=19, max_row=MAR_LAST)
    c2 = Reference(ws, min_col=3, min_row=MAR_FIRST, max_row=MAR_LAST)
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(c2)
    _color(ch2.series[0], NAVY_SOFT)
    _color(ch2.series[1], "9C3A5F")
    ch2.height, ch2.width = 22.0, 16.0
    tune_chart(ch2, axis_size=700, legend="t")
    zm2.add(ch2)

    module_month_chart(wb, ws, "Mariage", zm, "9C3A5F")
    note_box(ws, "B", "P", 18,
             "Les colonnes « Déjà payé » se remplissent toutes seules : enregistrez vos "
             "paiements dans le Journal avec le module « Mariage » et la même catégorie que "
             "le poste. Comparez jusqu'à trois devis et choisissez celui retenu.", height=28)
    return ws


# ==========================================================================
# IMMOBILIER & TERRAIN
# ==========================================================================
IMM_FIRST = 10
IMM_LAST = 19
LOT_FIRST = 25
ECH_FIRST = 44
ECH_LAST = 63


def build_immobilier(wb):
    ws = wb.create_sheet("Immobilier")
    setup_sheet(ws, tab_color="1E6B3C", freeze="A10", zoom=80)
    nav_bar(ws, "S", row=3)
    banner(ws, "S", "IMMOBILIER & TERRAIN",
           "Acquisition, construction, échéanciers de paiement, location et rendement")
    widths(ws, {"A": 2, "B": 4, "C": 26, "D": 16, "E": 20, "F": 11, "G": 12, "H": 16,
                "I": 13, "J": 16, "K": 16, "L": 16, "M": 10, "N": 14, "O": 14,
                "P": 14, "Q": 11, "R": 15, "S": 18})

    kpi_card(ws, 4, 2, "BIENS SUIVIS", f"=COUNTA($C${IMM_FIRST}:$C${IMM_LAST})",
             FMT_INT, NAVY, span=2)
    kpi_card(ws, 4, 4, "BUDGET TOTAL", f"=SUM($J${IMM_FIRST}:$J${IMM_LAST})",
             FMT_MONEY, NAVY_SOFT, span=3)
    kpi_card(ws, 4, 7, "INVESTI À CE JOUR", f"=SUM($K${IMM_FIRST}:$K${IMM_LAST})",
             FMT_MONEY, TEAL, span=3,
             note_formula=f'="Soit "&TEXT(IFERROR(SUM($K${IMM_FIRST}:$K${IMM_LAST})/'
                          f'SUM($J${IMM_FIRST}:$J${IMM_LAST}),0),"0%")&" du budget"')
    kpi_card(ws, 4, 10, "RESTE À FINANCER", f"=SUM($L${IMM_FIRST}:$L${IMM_LAST})",
             FMT_MONEY, ORANGE, span=3)
    kpi_card(ws, 4, 13, "LOYERS MENSUELS", f"=SUM($N${IMM_FIRST}:$N${IMM_LAST})",
             FMT_MONEY, GREEN, span=3)
    kpi_card(ws, 4, 16, "CASH-FLOW MENSUEL", f"=SUM($P${IMM_FIRST}:$P${IMM_LAST})",
             FMT_MONEY, BLUE, span=3,
             note_formula=f'="Rendement brut moyen : "&TEXT(IFERROR(SUM($N${IMM_FIRST}:'
                          f'$N${IMM_LAST})*12/SUM($J${IMM_FIRST}:$J${IMM_LAST}),0),"0.0%")')
    for start, end in (("D", "F"), ("G", "I"), ("J", "L"), ("M", "O"), ("P", "R")):
        currency_format(ws, f"{start}5:{end}5")

    section(ws, 8, "B", "S", "①  PORTEFEUILLE — BIENS BÂTIS & TERRAINS")
    table_header(ws, 9, "B", "S",
                 ["N°", "Désignation", "Type de bien", "Localisation", "Surface (m²)",
                  "Prix / m²", "Coût d'acquisition", "Frais annexes", "Budget total",
                  "Investi à ce jour", "Reste à financer", "% financé",
                  "Loyer mensuel", "Charges mensuelles", "Cash-flow", "Rend. brut",
                  "Statut", "Titre foncier"], height=36)
    for i in range(IMM_LAST - IMM_FIRST + 1):
        r = IMM_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=6).number_format = FMT_INT
        ws.cell(row=r, column=7,
                value=f'=IF(OR($F{r}="",$H{r}=""),"",IFERROR($H{r}/$F{r},""))'
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=8).number_format = FMT_MONEY
        ws.cell(row=r, column=9).number_format = FMT_MONEY
        ws.cell(row=r, column=10, value=f"=N($H{r})+N($I{r})").number_format = FMT_MONEY
        ws.cell(row=r, column=11,
                value=(f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Module,"Immobilier",'
                       f'J_Projet,$C{r},{OUT},{REAL}))')).number_format = FMT_MONEY
        ws.cell(row=r, column=12, value=f"=MAX(0,N($J{r})-N($K{r}))").number_format = FMT_MONEY
        ws.cell(row=r, column=13, value=f"=IFERROR($K{r}/$J{r},0)").number_format = FMT_PCT
        ws.cell(row=r, column=14).number_format = FMT_MONEY
        ws.cell(row=r, column=15).number_format = FMT_MONEY
        ws.cell(row=r, column=16, value=f"=N($N{r})-N($O{r})").number_format = FMT_MONEY
        ws.cell(row=r, column=17, value=f"=IFERROR($N{r}*12/$J{r},0)").number_format = FMT_PCT
    r = IMM_LAST + 1
    ws.cell(row=r, column=3, value="TOTAL").font = f(11, True, WHITE)
    for col in (9, 10, 11, 12, 14, 15, 16):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}{IMM_FIRST}:{L}{IMM_LAST})"
                ).number_format = FMT_MONEY
    paint(ws, f"B{IMM_FIRST}:S{r}", border=box_border(LINE), font=f(10))
    paint(ws, f"B{r}:S{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    input_style(ws, f"C{IMM_FIRST}:F{IMM_LAST}")
    input_style(ws, f"H{IMM_FIRST}:I{IMM_LAST}")
    input_style(ws, f"N{IMM_FIRST}:O{IMM_LAST}")
    input_style(ws, f"R{IMM_FIRST}:S{IMM_LAST}")
    paint(ws, f"J{IMM_FIRST}:M{IMM_LAST}", fill=solid(BG))
    paint(ws, f"P{IMM_FIRST}:Q{IMM_LAST}", fill=solid(BG))
    zebra(ws, f"B{IMM_FIRST}:S{IMM_LAST}", IMM_FIRST)
    databar(ws, f"M{IMM_FIRST}:M{IMM_LAST}", TEAL)
    color_scale_ok(ws, f"P{IMM_FIRST}:P{IMM_LAST}")
    hide_zeros(ws, f"G{IMM_FIRST}:L{IMM_LAST}")
    hide_zeros(ws, f"N{IMM_FIRST}:P{IMM_LAST}")
    currency_format(ws, f"G{IMM_FIRST}:L{IMM_LAST}", hide_zero=True)
    currency_format(ws, f"N{IMM_FIRST}:P{IMM_LAST}", hide_zero=True)
    currency_format(ws, f"G{r}:P{r}")

    for col, name in (("D", "Liste_TypesBien"), ("R", "Liste_StatutsBien"),
                      ("S", "Liste_Titres")):
        dv = DataValidation(type="list", formula1=name, allow_blank=True,
                            showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}{IMM_FIRST}:{col}{IMM_LAST}")

    # --- Lots de construction
    section(ws, 23, "B", "I", "②  BUDGET DE CONSTRUCTION PAR LOT")
    table_header(ws, 24, "B", "I",
                 ["N°", "Lot de travaux", "Bien concerné", "Budget prévu",
                  "Dépensé (Journal)", "Reste", "% consommé", "Statut"])
    for i, lot in enumerate(D.LOTS_CONSTRUCTION):
        r = LOT_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=3, value=lot)
        ws.cell(row=r, column=5).number_format = FMT_MONEY
        ws.cell(row=r, column=6,
                value=(f'=IF($D{r}="",'
                       f'SUMIFS(J_Montant,J_Module,"Immobilier",'
                       f'J_Libelle,"*"&$C{r}&"*",{OUT},{REAL}),'
                       f'SUMIFS(J_Montant,J_Module,"Immobilier",'
                       f'J_Libelle,"*"&$C{r}&"*",J_Projet,$D{r},{OUT},{REAL}))')
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=7, value=f"=MAX(0,N($E{r})-N($F{r}))").number_format = FMT_MONEY
        ws.cell(row=r, column=8, value=f"=IFERROR($F{r}/$E{r},0)").number_format = FMT_PCT
        ws.cell(row=r, column=9,
                value=(f'=IF($E{r}=0,"Non budgété",IF($H{r}>1,"⚠ Dépassement",'
                       f'IF($H{r}>=1,"✓ Terminé",IF($H{r}>0,"En cours","Non démarré"))))'))
        ws.cell(row=r, column=9).alignment = al("center")
    r = LOT_FIRST + len(D.LOTS_CONSTRUCTION)
    ws.cell(row=r, column=3, value="TOTAL TRAVAUX").font = f(11, True, WHITE)
    for col in (5, 6, 7):
        L = get_column_letter(col)
        ws.cell(row=r, column=col,
                value=f"=SUM({L}{LOT_FIRST}:{L}{r - 1})").number_format = FMT_MONEY
    ws.cell(row=r, column=8, value=f"=IFERROR($F{r}/$E{r},0)").number_format = FMT_PCT
    paint(ws, f"B{LOT_FIRST}:I{r}", border=box_border(LINE), font=f(10))
    paint(ws, f"B{r}:I{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    input_style(ws, f"C{LOT_FIRST}:E{r - 1}")
    paint(ws, f"F{LOT_FIRST}:H{r - 1}", fill=solid(BG))
    zebra(ws, f"B{LOT_FIRST}:I{r - 1}", LOT_FIRST)
    databar(ws, f"H{LOT_FIRST}:H{r - 1}", "1E6B3C")
    hide_zeros(ws, f"E{LOT_FIRST}:G{r - 1}")
    currency_format(ws, f"E{LOT_FIRST}:G{r - 1}", hide_zero=True)
    currency_format(ws, f"E{r}:G{r}")
    dvb = DataValidation(type="list", formula1="Liste_Projets", allow_blank=True,
                         showErrorMessage=False)
    ws.add_data_validation(dvb)
    dvb.add(f"D{LOT_FIRST}:D{r - 1}")

    # --- Échéancier
    section(ws, 42, "B", "I", "③  ÉCHÉANCIER DE PAIEMENT (terrain, acquisition, travaux)")
    table_header(ws, 43, "B", "I",
                 ["N°", "Échéance", "Bien concerné", "Objet", "Montant prévu",
                  "Payé", "Reste", "Alerte"])
    for i in range(ECH_LAST - ECH_FIRST + 1):
        r = ECH_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=3).number_format = FMT_DATE
        ws.cell(row=r, column=6).number_format = FMT_MONEY
        ws.cell(row=r, column=7).number_format = FMT_MONEY
        ws.cell(row=r, column=8, value=f"=N($F{r})-N($G{r})").number_format = FMT_MONEY
        ws.cell(row=r, column=9,
                value=(f'=IF($C{r}="","",IF($H{r}<=0,"✓ Soldé",'
                       f'IF($C{r}<TODAY(),"⚠ En retard",IF($C{r}<=TODAY()+30,'
                       f'"⏳ Sous 30 jours","À venir"))))'))
        ws.cell(row=r, column=9).alignment = al("center")
    paint(ws, f"B{ECH_FIRST}:I{ECH_LAST}", border=box_border(LINE), font=f(10))
    input_style(ws, f"C{ECH_FIRST}:G{ECH_LAST}")
    paint(ws, f"H{ECH_FIRST}:I{ECH_LAST}", fill=solid(BG))
    zebra(ws, f"B{ECH_FIRST}:I{ECH_LAST}", ECH_FIRST)
    hide_zeros(ws, f"F{ECH_FIRST}:H{ECH_LAST}")
    currency_format(ws, f"F{ECH_FIRST}:H{ECH_LAST}", hide_zero=True)
    dve = DataValidation(type="list", formula1="Liste_Projets", allow_blank=True,
                         showErrorMessage=False)
    ws.add_data_validation(dve)
    dve.add(f"D{ECH_FIRST}:D{ECH_LAST}")
    ws[f"C{ECH_LAST + 1}"] = "TOTAL ÉCHÉANCIER"
    for col in (6, 7, 8):
        L = get_column_letter(col)
        ws.cell(row=ECH_LAST + 1, column=col,
                value=f"=SUM({L}{ECH_FIRST}:{L}{ECH_LAST})").number_format = FMT_MONEY
    paint(ws, f"B{ECH_LAST + 1}:I{ECH_LAST + 1}", fill=solid(NAVY),
          font=f(10, True, WHITE))

    # --- Simulateur de crédit
    section(ws, 67, "C", "J", "④  SIMULATEUR DE CRÉDIT IMMOBILIER")
    _field(ws, 69, "Montant emprunté", value=0, numfmt=FMT_MONEY,
           help_text="Capital sollicité auprès de la banque.")
    _field(ws, 70, "Taux annuel", value=0.075, numfmt="0.00%")
    _field(ws, 71, "Durée (mois)", value=180, numfmt=FMT_INT,
           help_text="180 mois = 15 ans.")
    _field(ws, 72, "Assurance mensuelle", value=0, numfmt=FMT_MONEY)
    _field(ws, 73, "Mensualité totale",
           formula='=IFERROR(-PMT($E$70/12,$E$71,$E$69),0)+N($E$72)',
           numfmt=FMT_MONEY, editable=False,
           help_text="Capital + intérêts + assurance.")
    _field(ws, 74, "Coût total du crédit",
           formula='=$E$73*$E$71', numfmt=FMT_MONEY, editable=False)
    _field(ws, 75, "Dont intérêts et assurance",
           formula='=$E$74-$E$69', numfmt=FMT_MONEY, editable=False)
    _field(ws, 76, "Taux d'endettement estimé",
           formula='=IFERROR($E$73/(SUMIFS(J_Montant,J_Annee,Annee_Travail,'
                   'J_Type,"Revenu",' + REAL + ')/12),0)',
           numfmt=FMT_PCT, editable=False,
           help_text="Au-delà de 33 %, les banques refusent généralement.")
    currency_format(ws, "E69:F69")
    currency_format(ws, "E72:F75")

    ch1 = BarChart()
    ch1.type = "col"
    ch1.title = "Budget, investi et reste à financer par bien"
    d1 = Reference(ws, min_col=10, max_col=12, min_row=9, max_row=IMM_LAST)
    c1 = Reference(ws, min_col=3, min_row=IMM_FIRST, max_row=IMM_LAST)
    ch1.add_data(d1, titles_from_data=True)
    ch1.set_categories(c1)
    for i, serie in enumerate(ch1.series):
        _color(serie, [NAVY_SOFT, TEAL, ORANGE][i])
    ch1.height, ch1.width = 10.0, 17.0
    tune_chart(ch1, axis_size=700, legend="t")
    zi = ChartZone(ws, "B", 82)
    zi2 = ChartZone(ws, "K", 82)
    zi.add(ch1)

    ch2 = BarChart()
    ch2.type = "bar"
    ch2.title = "Consommation du budget travaux par lot"
    d2 = Reference(ws, min_col=5, max_col=6, min_row=24, max_row=LOT_FIRST + 13)
    c2 = Reference(ws, min_col=3, min_row=LOT_FIRST, max_row=LOT_FIRST + 13)
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(c2)
    _color(ch2.series[0], NAVY_SOFT)
    _color(ch2.series[1], "1E6B3C")
    ch2.height, ch2.width = 12.0, 16.0
    tune_chart(ch2, axis_size=700, legend="t")
    zi2.add(ch2)

    module_month_chart(wb, ws, "Immobilier", zi, "1E6B3C")
    note_box(ws, "B", "S", 7,
             "Reliez chaque dépense du Journal à un bien via la colonne « Projet / Bien » : "
             "l'investi, le reste à financer et le rendement se calculent seuls. Le budget "
             "par lot se rapproche du Journal via le libellé de l'opération.", height=28)
    return ws


# ==========================================================================
# BUSINESS & PROJETS
# ==========================================================================
BUS_FIRST = 10
BUS_LAST = 21
BUS_PL_FIRST = 27


def build_business(wb):
    ws = wb.create_sheet("Business")
    setup_sheet(ws, tab_color="E07B22", freeze="A10", zoom=80)
    nav_bar(ws, "R", row=3)
    banner(ws, "R", "BUSINESS & PROJETS",
           "Budget, chiffre d'affaires, marge et rentabilité de chaque activité")
    widths(ws, {"A": 2, "B": 4, "C": 26, "D": 16, "E": 14, "F": 12, "G": 12, "H": 15,
                "I": 15, "J": 15, "K": 11, "L": 15, "M": 15, "N": 15, "O": 11,
                "P": 11, "Q": 18, "R": 24})

    kpi_card(ws, 4, 2, "PROJETS ACTIFS",
             f'=COUNTIFS($E${BUS_FIRST}:$E${BUS_LAST},"En cours")', FMT_INT, NAVY, span=2)
    kpi_card(ws, 4, 4, "CA RÉALISÉ", f"=SUM($M${BUS_FIRST}:$M${BUS_LAST})",
             FMT_MONEY, GREEN, span=3,
             note_formula=f'="Prévu : "&TEXT(SUM($L${BUS_FIRST}:$L${BUS_LAST}),"#,##0")')
    kpi_card(ws, 4, 7, "CHARGES ENGAGÉES", f"=SUM($I${BUS_FIRST}:$I${BUS_LAST})",
             FMT_MONEY, RED, span=3,
             note_formula=f'="Budget : "&TEXT(SUM($H${BUS_FIRST}:$H${BUS_LAST}),"#,##0")')
    kpi_card(ws, 4, 10, "MARGE", f"=SUM($N${BUS_FIRST}:$N${BUS_LAST})",
             FMT_MONEY, BLUE, span=3)
    kpi_card(ws, 4, 13, "TAUX DE MARGE",
             f"=IFERROR(SUM($N${BUS_FIRST}:$N${BUS_LAST})/SUM($M${BUS_FIRST}:$M${BUS_LAST}),0)",
             FMT_PCT, PURPLE, span=3)
    kpi_card(ws, 4, 16, "BUDGET CONSOMMÉ",
             f"=IFERROR(SUM($I${BUS_FIRST}:$I${BUS_LAST})/SUM($H${BUS_FIRST}:$H${BUS_LAST}),0)",
             FMT_PCT, ORANGE, span=3)
    for start, end in (("D", "F"), ("G", "I"), ("J", "L")):
        currency_format(ws, f"{start}5:{end}5")

    section(ws, 8, "B", "R", "①  PORTEFEUILLE DE PROJETS & ACTIVITÉS")
    table_header(ws, 9, "B", "R",
                 ["N°", "Projet / Activité", "Type", "Statut", "Début", "Fin",
                  "Budget prévu", "Dépensé", "Reste", "% consommé", "CA prévu",
                  "CA réalisé", "Marge", "Marge %", "ROI", "Responsable", "Notes"],
                 height=34)
    for i in range(BUS_LAST - BUS_FIRST + 1):
        r = BUS_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=6).number_format = FMT_DATE
        ws.cell(row=r, column=7).number_format = FMT_DATE
        ws.cell(row=r, column=8).number_format = FMT_MONEY
        ws.cell(row=r, column=9,
                value=(f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Module,"Business",'
                       f'J_Projet,$C{r},{OUT},{REAL}))')).number_format = FMT_MONEY
        ws.cell(row=r, column=10, value=f"=N($H{r})-N($I{r})").number_format = FMT_MONEY
        ws.cell(row=r, column=11, value=f"=IFERROR($I{r}/$H{r},0)").number_format = FMT_PCT
        ws.cell(row=r, column=12).number_format = FMT_MONEY
        ws.cell(row=r, column=13,
                value=(f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Module,"Business",'
                       f'J_Projet,$C{r},J_Type,"Revenu",{REAL}))')).number_format = FMT_MONEY
        ws.cell(row=r, column=14, value=f"=N($M{r})-N($I{r})").number_format = FMT_MONEY
        ws.cell(row=r, column=15, value=f"=IFERROR($N{r}/$M{r},0)").number_format = FMT_PCT
        ws.cell(row=r, column=16, value=f"=IFERROR($N{r}/$I{r},0)").number_format = FMT_PCT
    r = BUS_LAST + 1
    ws.cell(row=r, column=3, value="TOTAL").font = f(11, True, WHITE)
    for col in (8, 9, 10, 12, 13, 14):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}{BUS_FIRST}:{L}{BUS_LAST})"
                ).number_format = FMT_MONEY
    ws.cell(row=r, column=15, value=f"=IFERROR($N{r}/$M{r},0)").number_format = FMT_PCT
    ws.cell(row=r, column=16, value=f"=IFERROR($N{r}/$I{r},0)").number_format = FMT_PCT
    paint(ws, f"B{BUS_FIRST}:R{r}", border=box_border(LINE), font=f(10))
    paint(ws, f"B{r}:R{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    input_style(ws, f"C{BUS_FIRST}:H{BUS_LAST}")
    input_style(ws, f"L{BUS_FIRST}:L{BUS_LAST}")
    input_style(ws, f"Q{BUS_FIRST}:R{BUS_LAST}")
    paint(ws, f"I{BUS_FIRST}:K{BUS_LAST}", fill=solid(BG))
    paint(ws, f"M{BUS_FIRST}:P{BUS_LAST}", fill=solid(BG))
    zebra(ws, f"B{BUS_FIRST}:R{BUS_LAST}", BUS_FIRST)
    databar(ws, f"K{BUS_FIRST}:K{BUS_LAST}", PURPLE)
    color_scale_ok(ws, f"N{BUS_FIRST}:N{BUS_LAST}")
    hide_zeros(ws, f"H{BUS_FIRST}:J{BUS_LAST}")
    hide_zeros(ws, f"L{BUS_FIRST}:N{BUS_LAST}")
    currency_format(ws, f"H{BUS_FIRST}:J{BUS_LAST}", hide_zero=True)
    currency_format(ws, f"L{BUS_FIRST}:N{BUS_LAST}", hide_zero=True)
    currency_format(ws, f"H{r}:N{r}")
    for col, name in (("D", "Liste_TypesProjet"), ("E", "Liste_StatutsProjet")):
        dv = DataValidation(type="list", formula1=name, allow_blank=True,
                            showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}{BUS_FIRST}:{col}{BUS_LAST}")

    # --- Compte de résultat mensuel
    section(ws, 25, "B", "I", "②  COMPTE DE RÉSULTAT MENSUEL — MODULE BUSINESS")
    table_header(ws, 26, "B", "I",
                 ["", "Mois", "Chiffre d'affaires", "Charges", "Marge",
                  "Taux de marge", "Marge cumulée", "Tendance"])
    for i in range(12):
        r = BUS_PL_FIRST + i
        ws.cell(row=r, column=3, value=f"=INDEX(Liste_Mois,{i + 1})")
        ws.cell(row=r, column=4,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{i + 1},'
                       f'J_Module,"Business",J_Type,"Revenu",{REAL})')
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=5,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{i + 1},'
                       f'J_Module,"Business",{OUT},{REAL})')).number_format = FMT_MONEY
        ws.cell(row=r, column=6, value=f"=$D{r}-$E{r}").number_format = FMT_MONEY
        ws.cell(row=r, column=7, value=f"=IFERROR($F{r}/$D{r},0)").number_format = FMT_PCT
        prev = "0" if i == 0 else f"$H{r - 1}"
        ws.cell(row=r, column=8, value=f"={prev}+$F{r}").number_format = FMT_MONEY
        ws.cell(row=r, column=9,
                value=(f'=IF($D{r}=0,"—",IF($F{r}>0,"▲ Bénéfice",'
                       f'IF($F{r}=0,"= Équilibre","▼ Perte")))'))
        ws.cell(row=r, column=9).alignment = al("center")
    r = BUS_PL_FIRST + 12
    ws.cell(row=r, column=3, value="TOTAL").font = f(11, True, WHITE)
    for col in (4, 5, 6):
        L = get_column_letter(col)
        ws.cell(row=r, column=col,
                value=f"=SUM({L}{BUS_PL_FIRST}:{L}{r - 1})").number_format = FMT_MONEY
    ws.cell(row=r, column=7, value=f"=IFERROR($F{r}/$D{r},0)").number_format = FMT_PCT
    paint(ws, f"B{BUS_PL_FIRST}:I{r}", border=box_border(LINE), font=f(10))
    paint(ws, f"B{r}:I{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    zebra(ws, f"B{BUS_PL_FIRST}:I{r - 1}", BUS_PL_FIRST)
    color_scale_ok(ws, f"F{BUS_PL_FIRST}:F{r - 1}")
    currency_format(ws, f"D{BUS_PL_FIRST}:F{r}")
    currency_format(ws, f"H{BUS_PL_FIRST}:H{r - 1}")

    # --- Seuil de rentabilité
    section(ws, 42, "C", "J", "③  SEUIL DE RENTABILITÉ")
    _field(ws, 44, "Charges fixes mensuelles", value=0, numfmt=FMT_MONEY,
           help_text="Loyer, salaires, abonnements — indépendantes du volume.")
    _field(ws, 45, "Taux de charges variables", value=0.40, numfmt="0.0%",
           help_text="Part du chiffre d'affaires absorbée par les coûts variables.")
    _field(ws, 46, "Taux de marge sur coûts variables",
           formula="=1-$E$45", numfmt=FMT_PCT, editable=False)
    _field(ws, 47, "Chiffre d'affaires d'équilibre (mois)",
           formula="=IFERROR($E$44/$E$46,0)", numfmt=FMT_MONEY, editable=False,
           help_text="CA minimum pour couvrir toutes vos charges.")
    _field(ws, 48, "Chiffre d'affaires d'équilibre (an)",
           formula="=$E$47*12", numfmt=FMT_MONEY, editable=False)
    _field(ws, 49, "CA réalisé cette année",
           formula=f"=$D{BUS_PL_FIRST + 12}", numfmt=FMT_MONEY, editable=False)
    _field(ws, 50, "Situation",
           formula='=IF($E$48=0,"Renseignez vos charges fixes",'
                   'IF($E$49>=$E$48,"✓ Seuil de rentabilité atteint",'
                   '"Reste "&TEXT($E$48-$E$49,"#,##0")&" "&Symbole_Base&" à réaliser"))',
           editable=False, help_cols=("G", "J"))
    currency_format(ws, "E44:F44")
    currency_format(ws, "E47:F49")

    ch1 = BarChart()
    ch1.type = "col"
    ch1.title = "Chiffre d'affaires et charges par mois"
    d1 = Reference(ws, min_col=4, max_col=5, min_row=26, max_row=BUS_PL_FIRST + 11)
    c1 = Reference(ws, min_col=3, min_row=BUS_PL_FIRST, max_row=BUS_PL_FIRST + 11)
    ch1.add_data(d1, titles_from_data=True)
    ch1.set_categories(c1)
    _color(ch1.series[0], GREEN)
    _color(ch1.series[1], RED)
    ch1.height, ch1.width = 9.0, 16.0
    tune_chart(ch1, axis_size=750, legend="t")
    zbu = ChartZone(ws, "B", 56)
    zbu2 = ChartZone(ws, "K", 56)
    zbu.add(ch1)

    ch2 = LineChart()
    ch2.title = "Marge cumulée"
    d2 = Reference(ws, min_col=8, min_row=26, max_row=BUS_PL_FIRST + 11)
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(c1)
    _color(ch2.series[0], PURPLE)
    ch2.height, ch2.width = 8.0, 16.0
    tune_chart(ch2, axis_size=750, legend="t")
    zbu2.add(ch2)

    ch3 = BarChart()
    ch3.type = "bar"
    ch3.title = "CA réalisé vs charges par projet"
    d3 = Reference(ws, min_col=13, min_row=9, max_row=BUS_LAST)
    d3b = Reference(ws, min_col=9, min_row=9, max_row=BUS_LAST)
    ch3.add_data(d3, titles_from_data=True)
    ch3.add_data(d3b, titles_from_data=True)
    ch3.set_categories(Reference(ws, min_col=3, min_row=BUS_FIRST, max_row=BUS_LAST))
    _color(ch3.series[0], GREEN)
    _color(ch3.series[1], RED)
    ch3.height, ch3.width = 12.0, 16.0
    tune_chart(ch3, axis_size=700, legend="t")
    zbu.add(ch3)

    module_month_chart(wb, ws, "Business", zbu2, "E07B22")
    note_box(ws, "B", "R", 7,
             "Chaque écriture du Journal portant le module « Business » et le nom du projet "
             "alimente automatiquement le dépensé, le chiffre d'affaires et la marge.",
             height=26)
    return ws


# ==========================================================================
# OBJECTIFS & ÉPARGNE
# ==========================================================================
OBJ_FIRST = 10
OBJ_LAST = 21
SIM_FIRST = 33


def build_objectifs(wb):
    ws = wb.create_sheet("Objectifs")
    setup_sheet(ws, tab_color="10706B", freeze="A10", zoom=85)
    nav_bar(ws, "N", row=3)
    banner(ws, "N", "OBJECTIFS D'ÉPARGNE & PROJETS DE VIE",
           "Chaque objectif, son échéance, l'effort mensuel requis et la progression réelle")
    widths(ws, {"A": 2, "B": 4, "C": 32, "D": 14, "E": 16, "F": 13, "G": 16, "H": 16,
                "I": 11, "J": 16, "K": 11, "L": 18, "M": 22, "N": 24})

    kpi_card(ws, 4, 2, "OBJECTIFS SUIVIS", f"=COUNTA($C${OBJ_FIRST}:$C${OBJ_LAST})",
             FMT_INT, NAVY, span=2)
    kpi_card(ws, 4, 4, "TOTAL À CONSTITUER", f"=SUM($E${OBJ_FIRST}:$E${OBJ_LAST})",
             FMT_MONEY, NAVY_SOFT, span=3)
    kpi_card(ws, 4, 7, "DÉJÀ CONSTITUÉ", f"=SUM($G${OBJ_FIRST}:$G${OBJ_LAST})",
             FMT_MONEY, GREEN, span=3,
             note_formula=f'="Soit "&TEXT(IFERROR(SUM($G${OBJ_FIRST}:$G${OBJ_LAST})/'
                          f'SUM($E${OBJ_FIRST}:$E${OBJ_LAST}),0),"0%")&" des objectifs"')
    kpi_card(ws, 4, 10, "RESTE À FINANCER", f"=SUM($H${OBJ_FIRST}:$H${OBJ_LAST})",
             FMT_MONEY, ORANGE, span=3)
    kpi_card(ws, 4, 13, "EFFORT MENSUEL REQUIS", f"=SUM($J${OBJ_FIRST}:$J${OBJ_LAST})",
             FMT_MONEY, TEAL, span=3, note="Toutes échéances confondues")
    for start, end in (("D", "F"), ("G", "I"), ("J", "L"), ("M", "O")):
        currency_format(ws, f"{start}5:{end}5")

    section(ws, 8, "B", "N", "①  MES OBJECTIFS")
    table_header(ws, 9, "B", "N",
                 ["N°", "Objectif", "Module", "Montant cible", "Échéance",
                  "Déjà constitué", "Reste", "Mois restants", "Effort mensuel",
                  "% atteint", "Statut", "Compte dédié", "Notes"], height=32)
    for i in range(OBJ_LAST - OBJ_FIRST + 1):
        r = OBJ_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        if i < len(D.OBJECTIFS_DEFAUT):
            name, mod = D.OBJECTIFS_DEFAUT[i]
            ws.cell(row=r, column=3, value=name)
            ws.cell(row=r, column=4, value=mod)
        ws.cell(row=r, column=5).number_format = FMT_MONEY
        ws.cell(row=r, column=6).number_format = FMT_DATE
        ws.cell(row=r, column=7,
                value=(f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Projet,$C{r},'
                       f'J_Type,"Épargne",{REAL})+SUMIFS(J_Montant,J_Projet,$C{r},'
                       f'J_Type,"Investissement",{REAL}))')).number_format = FMT_MONEY
        ws.cell(row=r, column=8, value=f"=MAX(0,N($E{r})-N($G{r}))").number_format = FMT_MONEY
        ws.cell(row=r, column=9,
                value=f'=IF($F{r}="","",MAX(0,ROUND(($F{r}-TODAY())/30.4,0)))'
                ).number_format = FMT_INT
        ws.cell(row=r, column=10,
                value=f'=IF($C{r}="",0,IFERROR($H{r}/MAX(1,$I{r}),$H{r}))'
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=11, value=f"=IFERROR($G{r}/$E{r},0)").number_format = FMT_PCT
        ws.cell(row=r, column=12,
                value=(f'=IF($C{r}="","",IF($K{r}>=1,"✓ Atteint",'
                       f'IF(AND($F{r}<>"",$F{r}<TODAY()),"⚠ Échéance dépassée",'
                       f'IF($K{r}>=0.5,"Bien engagé",IF($K{r}>0,"Démarré","À lancer")))))'))
        ws.cell(row=r, column=12).alignment = al("center")
        ws.cell(row=r, column=12).font = f(9, True)
    r = OBJ_LAST + 1
    ws.cell(row=r, column=3, value="TOTAL").font = f(11, True, WHITE)
    for col in (5, 7, 8, 10):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}{OBJ_FIRST}:{L}{OBJ_LAST})"
                ).number_format = FMT_MONEY
    ws.cell(row=r, column=11, value=f"=IFERROR($G{r}/$E{r},0)").number_format = FMT_PCT
    paint(ws, f"B{OBJ_FIRST}:N{r}", border=box_border(LINE), font=f(10))
    paint(ws, f"B{r}:N{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    input_style(ws, f"C{OBJ_FIRST}:F{OBJ_LAST}")
    input_style(ws, f"M{OBJ_FIRST}:N{OBJ_LAST}")
    paint(ws, f"G{OBJ_FIRST}:L{OBJ_LAST}", fill=solid(BG))
    zebra(ws, f"B{OBJ_FIRST}:N{OBJ_LAST}", OBJ_FIRST)
    databar(ws, f"K{OBJ_FIRST}:K{OBJ_LAST}", TEAL)
    hide_zeros(ws, f"E{OBJ_FIRST}:H{OBJ_LAST}")
    hide_zeros(ws, f"J{OBJ_FIRST}:J{OBJ_LAST}")
    currency_format(ws, f"E{OBJ_FIRST}:H{OBJ_LAST}", hide_zero=True)
    currency_format(ws, f"J{OBJ_FIRST}:J{OBJ_LAST}", hide_zero=True)
    currency_format(ws, f"E{r}:J{r}")
    dvm = DataValidation(type="list", formula1="Liste_Modules", allow_blank=True,
                         showErrorMessage=False)
    ws.add_data_validation(dvm)
    dvm.add(f"D{OBJ_FIRST}:D{OBJ_LAST}")
    dvc = DataValidation(type="list", formula1="Liste_Comptes", allow_blank=True,
                         showErrorMessage=False)
    ws.add_data_validation(dvc)
    dvc.add(f"M{OBJ_FIRST}:M{OBJ_LAST}")

    # --- Simulateur d'épargne
    section(ws, 25, "B", "H", "②  SIMULATEUR D'ÉPARGNE ET D'INTÉRÊTS COMPOSÉS")
    _field(ws, 27, "Versement mensuel", value=50000, numfmt=FMT_MONEY,
           lab_cols=("B", "C"), val_cols=("D", "E"), help_cols=("F", "H"),
           help_text="Ce que vous placez chaque mois.")
    _field(ws, 28, "Apport initial", value=0, numfmt=FMT_MONEY,
           lab_cols=("B", "C"), val_cols=("D", "E"), help_cols=("F", "H"))
    _field(ws, 29, "Rendement annuel brut", value=0.05, numfmt="0.00%",
           lab_cols=("B", "C"), val_cols=("D", "E"), help_cols=("F", "H"),
           help_text="Taux moyen espéré de votre placement.")
    _field(ws, 30, "Horizon (années)", value=20, numfmt=FMT_INT,
           lab_cols=("B", "C"), val_cols=("D", "E"), help_cols=("F", "H"))
    currency_format(ws, "D27:E28")

    table_header(ws, 32, "B", "F",
                 ["Année", "Versements cumulés", "Valeur acquise", "Gains", "Progression"])
    for i in range(30):
        r = SIM_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=3,
                value=f"=IF($B{r}>$D$30,\"\",$D$28+$D$27*12*$B{r})").number_format = FMT_MONEY
        prev = "$D$28" if i == 0 else f"$D{r - 1}"
        ws.cell(row=r, column=4,
                value=(f'=IF($B{r}>$D$30,"",({prev}+$D$27*12)*(1+$D$29))')
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=5,
                value=f'=IF($B{r}>$D$30,"",$D{r}-$C{r})').number_format = FMT_MONEY
        ws.cell(row=r, column=6,
                value=f'=IF($B{r}>$D$30,"",REPT("█",MIN(30,ROUND($D{r}/MAX(1,$D${SIM_FIRST + 29})*30,0))))')
        ws.cell(row=r, column=6).font = f(9, False, TEAL)
    paint(ws, f"B{SIM_FIRST}:F{SIM_FIRST + 29}", border=box_border(LINE), font=f(10))
    zebra(ws, f"B{SIM_FIRST}:F{SIM_FIRST + 29}", SIM_FIRST)
    currency_format(ws, f"C{SIM_FIRST}:E{SIM_FIRST + 29}")

    ch1 = BarChart()
    ch1.type = "bar"
    ch1.title = "Progression des objectifs"
    d1 = Reference(ws, min_col=7, max_col=8, min_row=9, max_row=OBJ_LAST)
    c1 = Reference(ws, min_col=3, min_row=OBJ_FIRST, max_row=OBJ_LAST)
    ch1.add_data(d1, titles_from_data=True)
    ch1.set_categories(c1)
    ch1.grouping = "stacked"
    ch1.overlap = 100
    _color(ch1.series[0], TEAL)
    _color(ch1.series[1], "D7E0EC")
    ch1.height, ch1.width = 12.0, 16.0
    tune_chart(ch1, axis_size=700, legend="t")
    zo = ChartZone(ws, "B", 66)
    zo2 = ChartZone(ws, "J", 66)
    zo.add(ch1)

    ch2 = LineChart()
    ch2.title = "Épargne : versements vs valeur acquise"
    d2 = Reference(ws, min_col=3, max_col=4, min_row=32, max_row=SIM_FIRST + 29)
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=2, min_row=SIM_FIRST, max_row=SIM_FIRST + 29))
    _color(ch2.series[0], MUTED)
    _color(ch2.series[1], TEAL)
    ch2.height, ch2.width = 10.0, 16.0
    tune_chart(ch2, axis_size=750, legend="t", skip=3)
    zo2.add(ch2)

    note_box(ws, "B", "N", 7,
             "Pour qu'un objectif se remplisse tout seul, enregistrez vos versements dans le "
             "Journal avec le type « Épargne » (ou « Investissement ») et le nom de l'objectif "
             "dans la colonne « Projet / Bien ».", height=26)
    return ws


# ==========================================================================
# DETTES & CRÉDITS
# ==========================================================================
DET_FIRST = 10
DET_LAST = 19
AMO_FIRST = 27
AMO_ROWS = 240


def build_dettes(wb):
    ws = wb.create_sheet("Dettes")
    setup_sheet(ws, tab_color="8B1A1A", freeze="A10", zoom=85)
    nav_bar(ws, "N", row=3)
    banner(ws, "N", "DETTES & CRÉDITS",
           "Mensualités, capital restant dû et tableau d'amortissement")
    widths(ws, {"A": 2, "B": 4, "C": 26, "D": 20, "E": 16, "F": 11, "G": 10, "H": 13,
                "I": 15, "J": 16, "K": 17, "L": 11, "M": 13, "N": 24})

    kpi_card(ws, 4, 2, "CRÉDITS EN COURS", f"=COUNTA($C${DET_FIRST}:$C${DET_LAST})",
             FMT_INT, NAVY, span=2)
    kpi_card(ws, 4, 4, "CAPITAL EMPRUNTÉ", f"=SUM($E${DET_FIRST}:$E${DET_LAST})",
             FMT_MONEY, NAVY_SOFT, span=3)
    kpi_card(ws, 4, 7, "MENSUALITÉS TOTALES", f"=SUM($I${DET_FIRST}:$I${DET_LAST})",
             FMT_MONEY, RED, span=3)
    kpi_card(ws, 4, 10, "CAPITAL RESTANT DÛ", f"=SUM($K${DET_FIRST}:$K${DET_LAST})",
             FMT_MONEY, ORANGE, span=3)
    kpi_card(ws, 4, 13, "TAUX D'ENDETTEMENT",
             f'=IFERROR(SUM($I${DET_FIRST}:$I${DET_LAST})/(SUMIFS(J_Montant,J_Annee,'
             f'Annee_Travail,J_Type,"Revenu",{REAL})/12),0)', FMT_PCT, PURPLE, span=2,
             note="Seuil bancaire usuel : 33 %")
    for start, end in (("D", "F"), ("G", "I"), ("J", "L")):
        currency_format(ws, f"{start}5:{end}5")

    section(ws, 8, "B", "N", "①  MES CRÉDITS ET DETTES")
    table_header(ws, 9, "B", "N",
                 ["N°", "Crédit / Dette", "Organisme", "Capital emprunté",
                  "Taux annuel", "Durée (mois)", "1re échéance", "Mensualité",
                  "Déjà remboursé", "Capital restant dû", "% remboursé",
                  "Fin prévue", "Notes"], height=32)
    for i in range(DET_LAST - DET_FIRST + 1):
        r = DET_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=5).number_format = FMT_MONEY
        ws.cell(row=r, column=6).number_format = "0.00%"
        ws.cell(row=r, column=7).number_format = FMT_INT
        ws.cell(row=r, column=8).number_format = FMT_DATE
        ws.cell(row=r, column=9,
                value=(f'=IF(OR($E{r}="",$G{r}=""),0,IF($F{r}=0,$E{r}/$G{r},'
                       f'-PMT($F{r}/12,$G{r},$E{r})))')).number_format = FMT_MONEY
        ws.cell(row=r, column=10,
                value=(f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Projet,$C{r},'
                       f'J_Type,"Remboursement crédit",{REAL}))')).number_format = FMT_MONEY
        ws.cell(row=r, column=11,
                value=(f'=IF($C{r}="",0,IF($F{r}=0,MAX(0,N($E{r})-N($J{r})),'
                       f'MAX(0,$E{r}*(1+$F{r}/12)^ROUND(IFERROR($J{r}/$I{r},0),0)'
                       f'-$I{r}*((1+$F{r}/12)^ROUND(IFERROR($J{r}/$I{r},0),0)-1)/($F{r}/12))))')
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=12,
                value=f"=IFERROR(1-$K{r}/$E{r},0)").number_format = FMT_PCT
        ws.cell(row=r, column=13,
                value=f'=IF(OR($H{r}="",$G{r}=""),"",EDATE($H{r},$G{r}-1))'
                ).number_format = FMT_DATE
    r = DET_LAST + 1
    ws.cell(row=r, column=3, value="TOTAL").font = f(11, True, WHITE)
    for col in (5, 9, 10, 11):
        L = get_column_letter(col)
        ws.cell(row=r, column=col, value=f"=SUM({L}{DET_FIRST}:{L}{DET_LAST})"
                ).number_format = FMT_MONEY
    ws.cell(row=r, column=12, value=f"=IFERROR(1-$K{r}/$E{r},0)").number_format = FMT_PCT
    paint(ws, f"B{DET_FIRST}:N{r}", border=box_border(LINE), font=f(10))
    paint(ws, f"B{r}:N{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    input_style(ws, f"C{DET_FIRST}:H{DET_LAST}")
    input_style(ws, f"N{DET_FIRST}:N{DET_LAST}")
    paint(ws, f"I{DET_FIRST}:M{DET_LAST}", fill=solid(BG))
    zebra(ws, f"B{DET_FIRST}:N{DET_LAST}", DET_FIRST)
    databar(ws, f"L{DET_FIRST}:L{DET_LAST}", "8B1A1A")
    hide_zeros(ws, f"E{DET_FIRST}:E{DET_LAST}")
    hide_zeros(ws, f"I{DET_FIRST}:K{DET_LAST}")
    currency_format(ws, f"E{DET_FIRST}:E{DET_LAST}", hide_zero=True)
    currency_format(ws, f"I{DET_FIRST}:K{DET_LAST}", hide_zero=True)
    currency_format(ws, f"E{r}:K{r}")

    # --- Tableau d'amortissement
    section(ws, 23, "B", "H", "②  TABLEAU D'AMORTISSEMENT")
    ws["C25"] = "Crédit à détailler :"
    ws["C25"].font = f(10, True, INK)
    ws["E25"] = "=IF($C10=\"\",\"\",$C10)"
    input_style(ws, "E25:F25")
    ws.merge_cells("E25:F25")
    ws["E25"].font = f(11, True, INK)
    ws["E25"].alignment = al("center")
    ws["G25"] = f'=IFERROR(MATCH($E$25,$C${DET_FIRST}:$C${DET_LAST},0),1)'
    ws["G25"].font = f(8, False, MUTED)
    dvd = DataValidation(
        type="list", formula1=f"$C${DET_FIRST}:$C${DET_LAST}", allow_blank=True,
        showErrorMessage=False)
    ws.add_data_validation(dvd)
    dvd.add("E25")

    table_header(ws, 26, "B", "G",
                 ["Éch.", "Date", "Mensualité", "Intérêts", "Capital amorti",
                  "Capital restant dû"])
    idx = f"INDEX($E${DET_FIRST}:$E${DET_LAST},$G$25)"
    rate = f"INDEX($F${DET_FIRST}:$F${DET_LAST},$G$25)/12"
    nper = f"INDEX($G${DET_FIRST}:$G${DET_LAST},$G$25)"
    start = f"INDEX($H${DET_FIRST}:$H${DET_LAST},$G$25)"
    pay = f"INDEX($I${DET_FIRST}:$I${DET_LAST},$G$25)"
    for i in range(AMO_ROWS):
        r = AMO_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=3,
                value=f'=IF($B{r}>{nper},"",IFERROR(EDATE({start},$B{r}-1),""))'
                ).number_format = FMT_DATE
        ws.cell(row=r, column=4,
                value=f'=IF($B{r}>{nper},"",{pay})').number_format = FMT_MONEY
        prev = idx if i == 0 else f"$G{r - 1}"
        ws.cell(row=r, column=5,
                value=f'=IF($B{r}>{nper},"",{prev}*{rate})').number_format = FMT_MONEY
        ws.cell(row=r, column=6,
                value=f'=IF($B{r}>{nper},"",$D{r}-$E{r})').number_format = FMT_MONEY
        ws.cell(row=r, column=7,
                value=f'=IF($B{r}>{nper},"",MAX(0,{prev}-$F{r}))').number_format = FMT_MONEY
    paint(ws, f"B{AMO_FIRST}:G{AMO_FIRST + AMO_ROWS - 1}", border=box_border(LINE),
          font=f(9))
    zebra(ws, f"B{AMO_FIRST}:G{AMO_FIRST + AMO_ROWS - 1}", AMO_FIRST)
    hide_zeros(ws, f"D{AMO_FIRST}:G{AMO_FIRST + AMO_ROWS - 1}")
    currency_format(ws, f"D{AMO_FIRST}:G{AMO_FIRST + 119}", hide_zero=True)

    ch1 = LineChart()
    ch1.title = "Capital restant dû dans le temps"
    d1 = Reference(ws, min_col=7, min_row=26, max_row=AMO_FIRST + 179)
    ch1.add_data(d1, titles_from_data=True)
    ch1.set_categories(Reference(ws, min_col=2, min_row=AMO_FIRST,
                                 max_row=AMO_FIRST + 179))
    _color(ch1.series[0], "8B1A1A")
    ch1.height, ch1.width = 9.0, 16.0
    tune_chart(ch1, axis_size=700, no_legend=True, skip=12)
    zde = ChartZone(ws, "P", 8)
    zde.add(ch1)

    ch2 = BarChart()
    ch2.type = "bar"
    ch2.title = "Capital emprunté vs restant dû"
    d2 = Reference(ws, min_col=5, min_row=9, max_row=DET_LAST)
    d2b = Reference(ws, min_col=11, min_row=9, max_row=DET_LAST)
    ch2.add_data(d2, titles_from_data=True)
    ch2.add_data(d2b, titles_from_data=True)
    ch2.set_categories(Reference(ws, min_col=3, min_row=DET_FIRST, max_row=DET_LAST))
    _color(ch2.series[0], NAVY_SOFT)
    _color(ch2.series[1], "8B1A1A")
    ch2.height, ch2.width = 10.0, 16.0
    tune_chart(ch2, axis_size=700, legend="t")
    zde.add(ch2)

    note_box(ws, "B", "N", 7,
             "Saisissez vos remboursements dans le Journal avec le type « Remboursement "
             "crédit » et le nom du crédit dans « Projet / Bien » : le capital restant dû "
             "se recalcule automatiquement.", height=26)
    return ws

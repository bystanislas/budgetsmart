"""Feuille « Budget 12 Mois » : la vue maîtresse façon Budget Gold.

Une seule grille : les postes en lignes, les 12 mois en colonnes, et pour chaque
mois deux colonnes — Prévu (issu du Estimation annuelle) et Réel (issu du Journal).
Aucune saisie ici : tout est calculé.
"""
from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter

import apex_data as D
from apex_style import (
    ChartZone, nav_bar, tune_chart,
    BG, BLUE, FMT_MONEY, FMT_MONEY_BLANK, FMT_PCT, GOLD, GOLD_SOFT, GREEN,
    INK, LINE, MUTED,
    NAVY, NAVY_MID, NAVY_SOFT, ORANGE, PURPLE, RED, TEAL, WHITE, al, banner,
    box_border, color_scale_ok, currency_format, f, hide_zeros, note_box,
    paint, section, setup_sheet, solid, widths,
)
from sheets_core import REAL

# Colonnes
COL_FIRST_MONTH = 4                    # D
COL_TOT_PREV = 28                      # AB
COL_TOT_REAL = 29                      # AC
COL_ECART = 30                         # AD
COL_PCT = 31                           # AE
LAST_COL = "AE"

HEAD_MONTH = 8
HEAD_SUB = 9

REV_CATS = D.CAT_GENERAL[:11] + ["Chiffre d'affaires", "Prestation de service",
                                 "Vente de produits"]
SKIP = {"Épargne (versement)", "Investissement", "Remboursement de crédit"}
EXP_CATS = [c for c in D.CAT_GENERAL[11:] if c not in SKIP]
MODULES_OUT = ["Mariage", "Immobilier", "Business"]


def _pc(m):
    return COL_FIRST_MONTH + 2 * (m - 1)


def _rc(m):
    return COL_FIRST_MONTH + 2 * (m - 1) + 1


def build_budget12(wb):
    ws = wb.create_sheet("Budget 12 Mois")
    setup_sheet(ws, tab_color=GOLD, freeze="D10", zoom=70)
    nav_bar(ws, LAST_COL, row=3, current="Budget 12 Mois")
    banner(ws, LAST_COL, "BUDGET 12 MOIS — ENTRÉES & SORTIES",
           '="Année "&Annee_Travail&"  •  "&Nom_Foyer&"  •  Prévu issu du Estimation annuelle, '
           'Réel issu du Journal  •  montants en "&Devise_Base')

    widths(ws, {"A": 2, "B": 4, "C": 40})
    for m in range(1, 13):
        widths(ws, {get_column_letter(_pc(m)): 12, get_column_letter(_rc(m)): 12})
    widths(ws, {"AB": 15, "AC": 15, "AD": 15, "AE": 12})

    note_box(ws, "B", LAST_COL, 5,
             "Cette feuille ne se remplit pas à la main. Colonne « Prévu » = ce que vous "
             "avez inscrit dans le Estimation annuelle. Colonne « Réel » = ce que vous avez "
             "enregistré dans le Journal. Changez l'année en haut du Tableau de bord.",
             height=30)

    # ---- En-têtes de mois -------------------------------------------------
    for m in range(1, 13):
        p, r = get_column_letter(_pc(m)), get_column_letter(_rc(m))
        ws.merge_cells(f"{p}{HEAD_MONTH}:{r}{HEAD_MONTH}")
        c = ws[f"{p}{HEAD_MONTH}"]
        c.value = D.MONTHS[m - 1]
        c.font = f(11, True, WHITE)
        c.alignment = al("center", "center")
        paint(ws, f"{p}{HEAD_MONTH}:{r}{HEAD_MONTH}",
              fill=solid(NAVY if m % 2 else NAVY_MID))
        for col, label in ((p, "Prévu"), (r, "Réel")):
            s = ws[f"{col}{HEAD_SUB}"]
            s.value = label
            s.font = f(9, True, WHITE)
            s.alignment = al("center", "center")
            s.fill = solid(NAVY_SOFT)
            s.border = box_border(NAVY_SOFT)
    ws.merge_cells(f"AB{HEAD_MONTH}:AE{HEAD_MONTH}")
    t = ws[f"AB{HEAD_MONTH}"]
    t.value = "TOTAL DE L'ANNÉE"
    t.font = f(11, True, NAVY)
    t.alignment = al("center", "center")
    paint(ws, f"AB{HEAD_MONTH}:AE{HEAD_MONTH}", fill=solid(GOLD_SOFT))
    for col, label in (("AB", "Prévu"), ("AC", "Réel"), ("AD", "Écart"),
                       ("AE", "% consommé")):
        s = ws[f"{col}{HEAD_SUB}"]
        s.value = label
        s.font = f(9, True, WHITE)
        s.alignment = al("center", "center", wrap=True)
        s.fill = solid(GOLD)
        s.border = box_border(GOLD)
    ws.row_dimensions[HEAD_MONTH].height = 22
    ws.row_dimensions[HEAD_SUB].height = 22
    ws.merge_cells(f"B{HEAD_MONTH}:C{HEAD_SUB}")
    h = ws[f"B{HEAD_MONTH}"]
    h.value = "POSTE"
    h.font = f(11, True, WHITE)
    h.alignment = al("left", "center", indent=1)
    paint(ws, f"B{HEAD_MONTH}:C{HEAD_SUB}", fill=solid(NAVY))

    row = 10
    blocks = {}

    def sec_header(text, color=NAVY_MID):
        nonlocal row
        section(ws, row, "B", LAST_COL, text, color)
        row += 1

    def data_row(label, prev_tpl, real_tpl, num=None, link=None):
        """Écrit une ligne de poste. `*_tpl` sont des lambdas(m) -> formule."""
        nonlocal row
        r = row
        if num is not None:
            ws.cell(row=r, column=2, value=num).alignment = al("center")
        cc = ws.cell(row=r, column=3, value=link if link else label)
        cc.font = f(10)
        cc.alignment = al("left", "center", indent=1)
        for m in range(1, 13):
            ws.cell(row=r, column=_pc(m), value=prev_tpl(m, r)).number_format = FMT_MONEY
            ws.cell(row=r, column=_rc(m), value=real_tpl(m, r)).number_format = FMT_MONEY
        row += 1
        return r

    def totals_cols(first_r, last_r, favourable_real=False):
        """Colonnes AB..AE pour un bloc de lignes déjà écrit."""
        for r in range(first_r, last_r + 1):
            prev_cells = "+".join(f"{get_column_letter(_pc(m))}{r}" for m in range(1, 13))
            real_cells = "+".join(f"{get_column_letter(_rc(m))}{r}" for m in range(1, 13))
            ws.cell(row=r, column=COL_TOT_PREV, value=f"={prev_cells}").number_format = FMT_MONEY
            ws.cell(row=r, column=COL_TOT_REAL, value=f"={real_cells}").number_format = FMT_MONEY
            sign = (f"$AC{r}-$AB{r}" if favourable_real else f"$AB{r}-$AC{r}")
            ws.cell(row=r, column=COL_ECART, value=f"={sign}").number_format = FMT_MONEY
            ws.cell(row=r, column=COL_PCT,
                    value=f"=IFERROR($AC{r}/$AB{r},0)").number_format = FMT_PCT

    def subtotal_row(label, first_r, last_r, color=GOLD_SOFT, fg=INK,
                     favourable_real=False):
        nonlocal row
        r = row
        ws.cell(row=r, column=3, value=label).font = f(10, True, fg)
        for m in range(1, 13):
            for col in (_pc(m), _rc(m)):
                L = get_column_letter(col)
                ws.cell(row=r, column=col,
                        value=f"=SUM({L}{first_r}:{L}{last_r})").number_format = FMT_MONEY
        paint(ws, f"B{r}:{LAST_COL}{r}", fill=solid(color), font=f(10, True, fg))
        totals_cols(r, r, favourable_real)
        row += 1
        return r

    # ======================================================================
    # A. ENTRÉES
    # ======================================================================
    sec_header("A ▸ ENTRÉES — REVENUS", NAVY)
    a_first = row
    for i, cat in enumerate(REV_CATS + [""] * 4):
        data_row(
            cat,
            lambda m, r: (f'=IF($C{r}="",0,SUMIFS(P_M{m},P_Annee,Annee_Travail,'
                          f'P_Cat,$C{r},P_Type,"Revenu"))'),
            lambda m, r: (f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Annee,Annee_Travail,'
                          f'J_Mois,{m},J_Cat,$C{r},J_Type,"Revenu",{REAL}))'),
            num=i + 1)
    a_last = row - 1
    totals_cols(a_first, a_last, favourable_real=True)
    blocks["entrees"] = subtotal_row("TOTAL DES ENTRÉES", a_first, a_last,
                                     color=GREEN, fg=WHITE, favourable_real=True)
    row += 1

    # ======================================================================
    # B. ÉPARGNE, PLACEMENTS & PROVISIONS
    # ======================================================================
    sec_header("B ▸ ÉPARGNE, PLACEMENTS & PROVISIONS  (vos objectifs)", TEAL)
    b_first = row
    for i in range(12):
        src = 10 + i
        data_row(
            None,
            lambda m, r, s=src: f"=IF($C{r}=\"\",0,N(Objectifs!$J${s}))",
            lambda m, r: (f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Annee,Annee_Travail,'
                          f'J_Mois,{m},J_Projet,$C{r},J_Type,"Épargne",{REAL})'
                          f'+SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
                          f'J_Projet,$C{r},J_Type,"Investissement",{REAL}))'),
            num=i + 1,
            link=f"=IF(Objectifs!$C${src}=\"\",\"\",Objectifs!$C${src})")
    b_objs_last = row - 1
    data_row(
        "Épargne / investissement non affecté à un objectif",
        lambda m, r: "=0",
        lambda m, r, bf=b_first, bl=b_objs_last: (
            f'=MAX(0,SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
            f'J_Type,"Épargne",{REAL})'
            f'+SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
            f'J_Type,"Investissement",{REAL})'
            f'-SUM({get_column_letter(_rc(m))}{bf}:{get_column_letter(_rc(m))}{bl}))'),
        num=13)
    b_last = row - 1
    totals_cols(b_first, b_last)
    blocks["epargne"] = subtotal_row("TOTAL ÉPARGNE & PROVISIONS", b_first, b_last,
                                     color=TEAL, fg=WHITE)
    row += 1

    # ======================================================================
    # C. CRÉDITS & DETTES
    # ======================================================================
    sec_header("C ▸ CRÉDITS & DETTES  (remboursements)", "8B1A1A")
    c_first = row
    for i in range(10):
        src = 10 + i
        data_row(
            None,
            lambda m, r, s=src: f"=IF($C{r}=\"\",0,N(Dettes!$I${s}))",
            lambda m, r: (f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Annee,Annee_Travail,'
                          f'J_Mois,{m},J_Projet,$C{r},J_Type,"Remboursement crédit",'
                          f'{REAL}))'),
            num=i + 1,
            link=f"=IF(Dettes!$C${src}=\"\",\"\",Dettes!$C${src})")
    c_credits_last = row - 1
    data_row(
        "Autres remboursements non rattachés à un crédit",
        lambda m, r: "=0",
        lambda m, r, cf=c_first, cl=c_credits_last: (
            f'=MAX(0,SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
            f'J_Type,"Remboursement crédit",{REAL})'
            f'-SUM({get_column_letter(_rc(m))}{cf}:{get_column_letter(_rc(m))}{cl}))'),
        num=11)
    c_last = row - 1
    totals_cols(c_first, c_last)
    blocks["credits"] = subtotal_row("TOTAL CRÉDITS & DETTES", c_first, c_last,
                                     color="8B1A1A", fg=WHITE)
    row += 1

    # ======================================================================
    # D. DÉPENSES COURANTES
    # ======================================================================
    sec_header("D ▸ DÉPENSES COURANTES  (module Général)", NAVY_MID)
    d_first = row
    for i, cat in enumerate(EXP_CATS):
        if cat == D.CAT_DIME:
            # la dîme prévue est calculée, pas saisie : taux × revenus soumis
            prev_tpl = (lambda m, r: (
                f'=IF(Dime_Active="Oui",INDEX(Calculs!$D$308:$D$319,{m}),'
                f'IF($C{r}="",0,SUMIFS(P_M{m},P_Annee,Annee_Travail,P_Cat,$C{r},'
                f'P_Type,"Dépense",P_Module,"Général")))'))
        else:
            prev_tpl = (lambda m, r: (
                f'=IF($C{r}="",0,SUMIFS(P_M{m},P_Annee,Annee_Travail,'
                f'P_Cat,$C{r},P_Type,"Dépense",P_Module,"Général"))'))
        data_row(
            cat,
            prev_tpl,
            lambda m, r: (f'=IF($C{r}="",0,SUMIFS(J_Montant,J_Annee,Annee_Travail,'
                          f'J_Mois,{m},J_Module,"Général",J_Cat,$C{r},J_Type,"Dépense",'
                          f'{REAL}))'),
            num=i + 1)
    d_cats_last = row - 1
    data_row(
        "Autres dépenses (catégorie non renseignée)",
        lambda m, r, df=d_first, dl=d_cats_last: (
            f'=MAX(0,SUMIFS(P_M{m},P_Annee,Annee_Travail,P_Type,"Dépense",'
            f'P_Module,"Général")'
            f'-SUM({get_column_letter(_pc(m))}{df}:{get_column_letter(_pc(m))}{dl}))'),
        lambda m, r, df=d_first, dl=d_cats_last: (
            f'=MAX(0,SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
            f'J_Module,"Général",J_Type,"Dépense",{REAL})'
            f'-SUM({get_column_letter(_rc(m))}{df}:{get_column_letter(_rc(m))}{dl}))'),
        num=len(EXP_CATS) + 1)
    d_last = row - 1
    totals_cols(d_first, d_last)
    blocks["courantes"] = subtotal_row("TOTAL DÉPENSES COURANTES", d_first, d_last,
                                       color=NAVY_MID, fg=WHITE)
    row += 1

    # ======================================================================
    # E. DÉPENSES DES MODULES
    # ======================================================================
    sec_header("E ▸ DÉPENSES DES MODULES  (mariage, immobilier & terrain, business)",
               PURPLE)
    e_first = row
    for i, mod in enumerate(MODULES_OUT):
        data_row(
            f"Module {mod}",
            lambda m, r, mo=mod: (f'=SUMPRODUCT((P_Annee=Annee_Travail)*'
                                  f'(P_Module="{mo}")*(P_Type="Dépense")*'
                                  f'INDEX(P_Mois,0,{m}))'),
            lambda m, r, mo=mod: (f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
                                  f'J_Module,"{mo}",J_Type,"Dépense",{REAL})'),
            num=i + 1)
    e_last = row - 1
    totals_cols(e_first, e_last)
    blocks["modules"] = subtotal_row("TOTAL DÉPENSES DES MODULES", e_first, e_last,
                                     color=PURPLE, fg=WHITE)
    row += 1

    # ======================================================================
    # F. SYNTHÈSE
    # ======================================================================
    sec_header("F ▸ SYNTHÈSE DU MOIS", NAVY)
    r_in = blocks["entrees"]
    r_ep, r_cr, r_dc, r_dm = (blocks["epargne"], blocks["credits"],
                              blocks["courantes"], blocks["modules"])

    def summary_row(label, tpl, numfmt=FMT_MONEY, color=NAVY, fg=WHITE, size=10):
        nonlocal row
        r = row
        ws.cell(row=r, column=3, value=label).font = f(size, True, fg)
        for m in range(1, 13):
            for col in (_pc(m), _rc(m)):
                ws.cell(row=r, column=col,
                        value=tpl(get_column_letter(col), r, m)).number_format = numfmt
        for col, L in ((COL_TOT_PREV, "AB"), (COL_TOT_REAL, "AC")):
            ws.cell(row=r, column=col, value=tpl(L, r, None)).number_format = numfmt
        paint(ws, f"B{r}:{LAST_COL}{r}", fill=solid(color), font=f(size, True, fg))
        ws.row_dimensions[r].height = 20
        row += 1
        return r

    r_tot_in = summary_row("TOTAL DES ENTRÉES",
                           lambda L, r, m: f"={L}{r_in}", color=GREEN)
    r_tot_out = summary_row(
        "TOTAL DES SORTIES",
        lambda L, r, m: f"={L}{r_ep}+{L}{r_cr}+{L}{r_dc}+{L}{r_dm}", color=RED)
    r_bal = summary_row("BALANCE DU MOIS",
                        lambda L, r, m: f"={L}{r_tot_in}-{L}{r_tot_out}", color=NAVY)

    # trésorerie cumulée : colonne par colonne
    r_cum = row
    ws.cell(row=r_cum, column=3, value="TRÉSORERIE CUMULÉE").font = f(10, True, WHITE)
    for m in range(1, 13):
        for col in (_pc(m), _rc(m)):
            L = get_column_letter(col)
            if m == 1:
                prev_ref = "Solde_Initial"
            else:
                prev_col = _pc(m - 1) if col == _pc(m) else _rc(m - 1)
                prev_ref = f"{get_column_letter(prev_col)}{r_cum}"
            ws.cell(row=r_cum, column=col,
                    value=f"={prev_ref}+{L}{r_bal}").number_format = FMT_MONEY
    for col, L in ((COL_TOT_PREV, "AB"), (COL_TOT_REAL, "AC")):
        last_col = get_column_letter(_pc(12) if col == COL_TOT_PREV else _rc(12))
        ws.cell(row=r_cum, column=col,
                value=f"={last_col}{r_cum}").number_format = FMT_MONEY
    paint(ws, f"B{r_cum}:{LAST_COL}{r_cum}", fill=solid(NAVY_SOFT), font=f(10, True, WHITE))
    ws.row_dimensions[r_cum].height = 20
    row += 1

    r_taux = summary_row(
        "TAUX D'ÉPARGNE",
        lambda L, r, m: f"=IFERROR({L}{r_ep}/{L}{r_tot_in},0)",
        numfmt=FMT_PCT, color=TEAL)
    r_cons = summary_row(
        "BUDGET CONSOMMÉ (sorties réel / prévu)",
        lambda L, r, m: (
            f"=IFERROR({get_column_letter(_rc(m))}{r_tot_out}/"
            f"{get_column_letter(_pc(m))}{r_tot_out},0)" if m else
            f"=IFERROR($AC{r_tot_out}/$AB{r_tot_out},0)"),
        numfmt=FMT_PCT, color=ORANGE)

    # ---- Mise en forme globale -------------------------------------------
    last_row = row - 1
    paint(ws, f"B10:{LAST_COL}{last_row}", border=box_border(LINE))
    # colonnes « Prévu » légèrement grisées pour distinguer du réel
    for m in range(1, 13):
        L = get_column_letter(_pc(m))
        for r in range(10, last_row + 1):
            c = ws.cell(row=r, column=_pc(m))
            if c.fill is None or c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = solid(BG)
    paint(ws, f"AB10:AE{last_row}", fill=solid("FDFBF3"))
    for r in (r_tot_in, r_tot_out, r_bal, r_cum, r_taux, r_cons):
        paint(ws, f"B{r}:{LAST_COL}{r}",
              fill=solid({r_tot_in: GREEN, r_tot_out: RED, r_bal: NAVY,
                          r_cum: NAVY_SOFT, r_taux: TEAL, r_cons: ORANGE}[r]),
              font=f(10, True, WHITE))
    color_scale_ok(ws, f"AD10:AD{last_row}")
    # les zéros disparaissent dans le corps de la grille, pas dans la synthèse
    hide_zeros(ws, f"D10:AC{r_tot_in - 1}")
    for m in range(1, 13):
        currency_format(ws, f"{get_column_letter(_pc(m))}10:"
                            f"{get_column_letter(_rc(m))}{r_tot_in - 1}", hide_zero=True)
        currency_format(ws, f"{get_column_letter(_pc(m))}{r_tot_in}:"
                            f"{get_column_letter(_rc(m))}{r_bal}")
    currency_format(ws, f"AB10:AC{r_tot_in - 1}", hide_zero=True)
    currency_format(ws, f"AB{r_tot_in}:AC{r_bal}")

    # ---- Graphique de synthèse -------------------------------------------
    ch = BarChart()
    ch.type = "col"
    ch.title = "Entrées, sorties et balance — réalisé par mois"
    calc = wb["Calculs"]
    data = Reference(calc, min_col=4, max_col=4, min_row=5, max_row=17)
    data2 = Reference(calc, min_col=9, max_col=10, min_row=5, max_row=17)
    ch.add_data(data, titles_from_data=True)
    ch.add_data(data2, titles_from_data=True)
    ch.set_categories(Reference(calc, min_col=3, min_row=6, max_row=17))
    for i, serie in enumerate(ch.series):
        gp = GraphicalProperties(solidFill=[GREEN, RED, NAVY][i])
        gp.line.solidFill = [GREEN, RED, NAVY][i]
        serie.graphicalProperties = gp
    ch.height, ch.width = 9.0, 24.0
    ch.gapWidth = 60
    tune_chart(ch, axis_size=750, legend="t")
    ChartZone(ws, "B", last_row + 3).add(ch)

    note_box(ws, "B", LAST_COL, last_row + 25,
             "Pour modifier un montant PRÉVU : allez dans « Estimation annuelle ». Pour corriger "
             "un montant RÉEL : allez dans « Journal ». Cette feuille est le miroir des "
             "deux, mois par mois.", height=26)
    return ws

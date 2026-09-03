#!/usr/bin/env python3
"""Génère le classeur APEX Budget (.xlsx).

Usage :
    python3 build_apex_budget.py [chemin/de/sortie.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Protection
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import apex_data as D
from apex_style import (
    BG, BLUE, BLUE_SOFT, CARD, FMT_INT, FMT_MONEY, FMT_PCT, GOLD, GOLD_SOFT,
    GREEN, INK, LINE, MUTED, NAVY, NAVY_MID, NAVY_SOFT, ORANGE, PURPLE, RED,
    TEAL, WHITE, al, banner, box_border, currency_format, f, kpi_card, link,
    big_button, nav_bar, note_box, paint, section, setup_sheet, solid,
    table_header,
    widths,
)
import sheets_budget12 as SB
import sheets_core as SC
import sheets_dash as SD
import sheets_modules as SM

ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DEFAULT = ROOT / "Budget-Smart-by-Apex-Africa.xlsx"
BRAND = ROOT / "assets" / "brand"


def place_logo(ws, filename, anchor, width, height):
    """Pose un visuel de marque sur une feuille (silencieux si le fichier manque)."""
    path = BRAND / filename
    if not path.exists():
        return None
    img = XLImage(str(path))
    img.width, img.height = width, height
    ws.add_image(img, anchor)
    return img

ETAPES = [
    ("①   REMPLIR MES INFORMATIONS", "Nom, coordonnées, devise, comptes bancaires",
     "Paramètres", NAVY),
    ("②   ESTIMER MON ANNÉE", "Charges fixes, variables et imprévus, mois par mois",
     "Estimation Annuelle", NAVY_SOFT),
    ("③   BUDGET SMART", "Mon journal de tous les jours — tout se synchronise",
     "Journal", GOLD),
]

MODULES_BTN = [
    ("MARIAGE", "Postes, devis, financement", "Mariage", "9C3A5F"),
    ("PROJET", "Business, chiffre d'affaires, marge", "Business", "E07B22"),
    ("IMMO", "Terrains, biens, travaux, loyers", "Immobilier", "1E6B3C"),
]

CONSULTER = [
    ("Tableau de bord", "Tableau de Bord"), ("Budget 12 mois", "Budget 12 Mois"),
    ("Synthèse annuelle", "Synthèse Annuelle"), ("Objectifs & épargne", "Objectifs"),
    ("Dettes & crédits", "Dettes"), ("Charges récurrentes", "Récurrents"),
    ("Guide d'utilisation", "Guide"), ("Référentiels & listes", "Listes"),
]


def build_accueil(wb):
    ws = wb.create_sheet("Accueil")
    setup_sheet(ws, tab_color=GOLD, zoom=90)
    widths(ws, {"A": 3, "B": 26, "C": 26, "D": 26, "E": 26, "F": 26, "G": 26, "H": 3})

    # ---------- en-tête de marque -------------------------------------
    for r in range(1, 6):
        ws.row_dimensions[r].height = 30
    place_logo(ws, "budget-smart-logo-2400.png", "B1", 476, 137)
    ws.merge_cells("B6:G6")
    paint(ws, "B6:G6", fill=solid(GOLD))
    ws.row_dimensions[6].height = 5
    ws.merge_cells("B7:G7")
    sub = ws["B7"]
    sub.value = D.APP_BASELINE
    sub.font = f(11, True, NAVY)
    sub.alignment = al("left", "center", indent=1)
    paint(ws, "B7:G7", fill=solid(GOLD_SOFT))
    ws.row_dimensions[7].height = 26

    ws.merge_cells("B9:G9")
    w = ws["B9"]
    w.value = ('="Bonjour "&Nom_Foyer&"   •   période suivie : "&Mois_Suivi&" "&'
               'Annee_Travail&"   •   devise : "&Devise_Base&" ("&Symbole_Base&")"')
    w.font = f(12, True, NAVY)
    w.alignment = al("left", "center", indent=1)
    ws.row_dimensions[9].height = 24
    ws.merge_cells("B10:G10")
    co = ws["B10"]
    co.value = ('=TRIM(Adresse_Entite&IF(Adresse_Entite="","",", ")&Ville_Entite'
                '&IF(Pays_Entite="","",", "&Pays_Entite)'
                '&IF(Tel_Entite="",""," · "&Tel_Entite)'
                '&IF(Email_Entite="",""," · "&Email_Entite)'
                '&IF(Site_Entite="",""," · "&Site_Entite))')
    co.font = f(9, False, MUTED)
    co.alignment = al("left", "center", indent=1)
    ws.row_dimensions[10].height = 18

    # ---------- le parcours -------------------------------------------
    section(ws, 12, "B", "G", "VOTRE PARCOURS — DANS CET ORDRE")
    for i, (label, sub_, sheet, color) in enumerate(ETAPES):
        big_button(ws, 13, 2 + i * 2, 2, 4, label, sub_, sheet, color)
    ws.merge_cells("B18:G18")
    fl = ws["B18"]
    fl.value = ("Une fois vos informations et votre estimation posées, il ne reste plus "
                "qu'à tenir le journal : tableau de bord, budget 12 mois, synthèse et "
                "modules se remplissent tout seuls.")
    fl.font = f(10, False, MUTED, italic=True)
    fl.alignment = al("left", "center", indent=1)
    ws.row_dimensions[18].height = 22

    # ---------- les modules -------------------------------------------
    section(ws, 20, "B", "G", "MES MODULES — chacun sa feuille, aucun mélange")
    for i, (label, sub_, sheet, color) in enumerate(MODULES_BTN):
        big_button(ws, 21, 2 + i * 2, 2, 3, label, sub_, sheet, color)

    # ---------- où j'en suis -------------------------------------------
    section(ws, 26, "B", "G", "OÙ J'EN SUIS AUJOURD'HUI")
    kpi_card(ws, 27, 2, "SOLDE DU MOIS",
             f"=INDEX(Calculs!$J${SD.C_MONTH_FIRST}:$J${SD.C_MONTH_FIRST + 11},Mois_Num)",
             FMT_MONEY, BLUE, span=2)
    kpi_card(ws, 27, 4, "DÉPENSES DU MOIS",
             f"=INDEX(Calculs!$E${SD.C_MONTH_FIRST}:$E${SD.C_MONTH_FIRST + 11},Mois_Num)",
             FMT_MONEY, RED, span=2)
    kpi_card(ws, 27, 6, "TRÉSORERIE DISPONIBLE",
             f"=Paramètres!$H${SC.PARAM_TOTAL_ROW}", FMT_MONEY, GREEN, span=2)
    for rng in ("B28:C28", "D28:E28", "F28:G28"):
        currency_format(ws, rng)

    # ---------- consulter ----------------------------------------------
    section(ws, 32, "B", "G", "CONSULTER")
    for i, (label, sheet) in enumerate(CONSULTER):
        r = 33 + (i // 3) * 2
        c = 2 + (i % 3) * 2
        c1 = get_column_letter(c)
        c2 = get_column_letter(c + 1)
        ws.merge_cells(f"{c1}{r}:{c2}{r}")
        link(ws, f"{c1}{r}", "➜  " + label, sheet, "A1", fill_color=CARD)
        paint(ws, f"{c1}{r}:{c2}{r}", fill=solid(CARD), border=box_border(LINE))
        ws.row_dimensions[r].height = 22

    last = 39
    note_box(ws, "B", "G", last,
             "Les cases sur fond jaune pâle sont les seules à remplir. Tout le reste est "
             "calculé et verrouillé — mot de passe APEX-2026 si vous devez y toucher.",
             height=30)
    ws.merge_cells(f"B{last + 2}:G{last + 2}")
    paint(ws, f"B{last + 2}:G{last + 2}", fill=solid(GOLD))
    ws.row_dimensions[last + 2].height = 4
    for i, line in enumerate((
            D.APP_SIGNATURE,
            "contact@apxafrica.com · www.apxafrica.com",
            D.APP_COPYRIGHT,
            f"{D.APP_NAME} {D.APP_BRAND} v{D.APP_VERSION} — feuilles protégées, "
            f"aucune macro.")):
        r = last + 3 + i
        ws.merge_cells(f"B{r}:G{r}")
        c = ws[f"B{r}"]
        c.value = line
        c.font = f(9, i == 0, NAVY if i == 0 else MUTED, italic=(i >= 2))
        c.alignment = al("left", "center", indent=1)
    return ws


GUIDE_SECTIONS = [
    ("Protection du classeur et droits", [
        "Les feuilles sont protégées : les cellules de calcul sont verrouillées et "
        "leurs formules masquées. Seules les zones de saisie — repérables à leur fond "
        "jaune pâle — restent modifiables.",
        "Mot de passe de déverrouillage : APEX-2026. Excel : Révision → Ôter la "
        "protection de la feuille. LibreOffice : Outils → Protéger la feuille. "
        "Changez ce mot de passe avant toute diffusion.",
        "La structure du classeur est également protégée : on ne peut ni ajouter, ni "
        "supprimer, ni renommer un onglet sans le mot de passe. La feuille technique "
        "« Calculs » est masquée pour la même raison.",
        "À savoir, honnêtement : la protection Excel empêche les modifications "
        "accidentelles et la recopie facile des formules. Ce n'est pas du chiffrement — "
        "un outil spécialisé lève ce verrou. Elle protège l'intégrité du classeur, "
        "pas le secret de son contenu.",
        "« Budget Smart » et son logo sont la propriété d'APEX AFRICA. Le droit "
        "d'auteur s'applique dès la création ; pour protéger le NOM comme marque, il "
        "faut un dépôt à l'OAPI (valable dans les 17 États membres, dont la Côte "
        "d'Ivoire) — le droit d'auteur seul ne réserve pas un nom commercial.",
    ]),
    ("Où est « le budget sur 12 mois » ?", [
        "Feuille BUDGET 12 MOIS : c'est la grande grille équivalente à la feuille "
        "« Budget Gold ». Les postes en lignes, les 12 mois en colonnes, et pour chaque "
        "mois deux colonnes côte à côte : Prévu et Réel.",
        "Elle est organisée en cinq blocs : A. Entrées (revenus) — B. Épargne, placements "
        "et provisions — C. Crédits et dettes — D. Dépenses courantes — E. Dépenses des "
        "modules (mariage, immobilier, business). Puis la synthèse : total des entrées, "
        "total des sorties, balance du mois, trésorerie cumulée, taux d'épargne.",
        "On n'y saisit RIEN. Le Prévu vient de la feuille Estimation annuelle, le Réel vient de la "
        "feuille Journal. C'est le miroir des deux.",
        "Les trois feuilles se répondent : Estimation annuelle = ce que je décide de dépenser ; "
        "Journal = ce que je dépense vraiment ; Budget 12 mois = la comparaison des deux, "
        "mois par mois, sur une seule page.",
    ]),
    ("Le principe en une phrase", [
        "Vous saisissez vos opérations dans une seule feuille — le Journal — et tout le "
        "reste du classeur se calcule automatiquement : tableau de bord, synthèse annuelle, "
        "budget mariage, immobilier, business, objectifs, crédits.",
    ]),
    ("1. Paramètres — à faire une seule fois", [
        "Bloc ① IDENTITÉ & COORDONNÉES : raison sociale, responsable, activité, "
        "adresse, ville, pays, téléphone, email, site web, numéro d'identification. "
        "Ces informations remontent en tête de la page d'accueil.",
        "Rappel : seules les cases sur fond JAUNE PÂLE sont modifiables. Le reste de la "
        "feuille est verrouillé pour protéger les formules.",
        "Choisissez votre DEVISE DE BASE (XOF, XAF, EUR, USD, MAD, NGN, GHS, CDF…). "
        "Tous les montants et tous les symboles du classeur s'y adaptent.",
        "Listez vos comptes, caisses et portefeuilles mobile money, avec leur solde "
        "d'ouverture. Leur solde actuel se recalcule ensuite tout seul.",
        "Fixez votre taux d'épargne cible et votre seuil d'alerte budgétaire.",
    ]),
    ("2. Devises multiples", [
        "Chaque ligne du Journal peut être saisie dans une autre devise : indiquez le code "
        "(EUR, USD…) dans la colonne « Devise ».",
        "Le taux de conversion est repris automatiquement de la feuille Listes, et le "
        "montant est converti dans votre devise de base.",
        "Les cours sont indicatifs : mettez-les à jour dans Listes, colonne « Cours "
        "indicatif ». Le classeur recalcule immédiatement tous les taux croisés.",
        "Laissez la colonne « Devise » vide pour utiliser votre devise de base.",
        "Choisissez votre devise de base AVANT de commencer à saisir. Une ligne laissée "
        "sans code devise est toujours interprétée dans la devise de base du moment : si "
        "vous changez de devise de base plus tard, ces lignes suivent. Pour figer une "
        "opération dans une monnaie précise, indiquez toujours son code.",
    ]),
    ("3. Journal — la saisie quotidienne", [
        "La feuille est découpée en quatre blocs, repérés par un bandeau de couleur.",
        "BLOC ① À REMPLIR (fond jaune) — sept colonnes, et c'est tout : Date, Type "
        "d'opération, Module, Catégorie, Libellé, Montant, Compte. Le montant est "
        "toujours POSITIF.",
        "BLOC ② CALCULÉ POUR VOUS — Sens (▲ ENTRÉE / ▼ SORTIE / ↔ NEUTRE), ENTRÉE, "
        "SORTIE et SOLDE CUMULÉ. Vous voyez immédiatement, ligne par ligne, ce qui est "
        "rentré, ce qui est sorti et où vous en êtes, comme sur un relevé de banque.",
        "BLOC ③ COMPLÉMENTS (facultatif) — Devise, Moyen de paiement, Projet / Bien, "
        "Nature, Statut, Récurrent, Notes. À remplir seulement si c'est utile.",
        "BLOC ④ TECHNIQUE — taux de change, montant converti, mois, année. Vous pouvez "
        "replier ces colonnes avec le petit « − » au-dessus.",
        "Le TYPE dit tout : Revenu = l'argent entre. Dépense, Épargne, Investissement et "
        "Remboursement crédit = l'argent sort. Transfert = neutre, d'un de vos comptes "
        "vers un autre.",
        "Épargne et Investissement sortent de votre poche mais ne sont pas perdus : le "
        "tableau de bord les compte à part, dans « MIS DE CÔTÉ ».",
        "Pour distinguer une épargne BLOQUÉE d'une épargne disponible, donnez la nature "
        "« Épargne » ou « Épargne bloquée » au compte concerné dans Paramètres.",
        "« Projet / Bien » relie l'opération à un bien immobilier, un projet business, un "
        "objectif d'épargne ou un crédit : c'est ce qui alimente ces modules.",
        "« Statut » : Payé compte dans le réalisé ; Prévu et En attente sont suivis à "
        "part comme des engagements ; Annulé est ignoré.",
    ]),
    ("La dîme", [
        "Dans Paramètres, bloc ⑤ : activez la dîme, choisissez le taux (10 % par défaut) "
        "et l'assiette — Salaire uniquement, Salaire + primes, ou Tous les revenus.",
        "Le classeur calcule alors seul, chaque mois, les revenus soumis, la dîme due, "
        "ce que vous avez déjà versé et ce qu'il reste à verser. Le bloc DÎME du tableau "
        "de bord et une alerte dédiée vous le rappellent.",
        "Dans la grille Budget 12 Mois, la ligne « Dîme / Offrandes » affiche "
        "automatiquement le montant dû en colonne Prévu.",
        "Pour enregistrer un versement : une ligne du Journal, type Dépense, catégorie "
        "« Dîme / Offrandes ». Le reste à verser se met à jour.",
        "À savoir : Excel sans macro ne peut pas créer cette ligne tout seul. Le classeur "
        "calcule et vous rappelle le montant ; la saisie du versement reste à vous.",
    ]),
    ("Modules séparés — le périmètre", [
        "Le mariage, l'immobilier et le business ne doivent pas polluer votre budget "
        "courant. Dans Paramètres, bloc ⑤, le PÉRIMÈTRE D'ANALYSE règle ce que comptent "
        "le tableau de bord et la synthèse annuelle.",
        "« Général seul » (par défaut) : seul votre budget courant est agrégé. "
        "« Tout confondu » : tout est additionné. Ou choisissez un module précis pour "
        "n'analyser que lui.",
        "Quel que soit ce réglage, chaque module garde sa feuille complète, avec ses "
        "propres indicateurs, ses propres tableaux et son graphique mois par mois. "
        "La grille Budget 12 Mois, elle, sépare toujours les modules en bloc E.",
        "Un seul Journal alimente tout : c'est ce qui garantit que rien n'est saisi deux "
        "fois et que tout reste cohérent. La séparation se fait à la lecture, pas à la "
        "saisie.",
    ]),
    ("4. Estimation annuelle — le budget prévisionnel", [
        "Une ligne = une catégorie budgétée pour une année donnée.",
        "Vous saisissez les douze mois ; le total, le réalisé, l'écart et le pourcentage "
        "consommé se calculent automatiquement.",
        "Pour préparer l'année suivante, recopiez les lignes en changeant l'année : "
        "l'historique des années précédentes reste intact et consultable.",
    ]),
    ("5. Tableau de bord", [
        "Choisissez l'année et le mois en haut de la feuille : tout se recalcule.",
        "Dix indicateurs, quatre graphiques, le top 10 des dépenses du mois, le comparatif "
        "prévu / réalisé par module et sept alertes automatiques.",
    ]),
    ("6. Modules spécialisés", [
        "MARIAGE : chaque poste accepte une estimation et jusqu'à trois devis ; vous "
        "choisissez le devis retenu. Les paiements réels remontent du Journal (module "
        "Mariage + même catégorie). Le plan de financement calcule le reste à financer "
        "et l'effort d'épargne mensuel jusqu'à la date du mariage.",
        "IMMOBILIER & TERRAIN : portefeuille de biens et de terrains (surface, prix au m², "
        "budget, investi, reste à financer, loyer, cash-flow, rendement brut), budget de "
        "construction par lot, échéancier de paiement avec alertes, simulateur de crédit.",
        "BUSINESS & PROJETS : budget et chiffre d'affaires par projet, marge, ROI, compte "
        "de résultat mensuel et seuil de rentabilité.",
        "OBJECTIFS : montant cible, échéance, effort mensuel requis, progression réelle, "
        "et un simulateur d'intérêts composés sur trente ans.",
        "DETTES : mensualité calculée, capital restant dû, taux d'endettement et tableau "
        "d'amortissement sur 240 échéances.",
        "RÉCURRENTS : abonnements et prélèvements ramenés au mois quelle que soit leur "
        "fréquence, avec alerte à J-30, J-7 et en cas de retard.",
    ]),
    ("7. Les couleurs", [
        "Fond jaune pâle : cellule à remplir.",
        "Fond gris clair : cellule calculée — ne pas écrire dedans.",
        "Vert : favorable. Orange : vigilance. Rouge : dépassement ou retard.",
    ]),
    ("8. Bonnes pratiques", [
        "Saisissez au moins une fois par semaine ; l'idéal reste le jour même.",
        "Utilisez le statut « Prévu » pour les engagements à venir : ils apparaissent dans "
        "les alertes sans fausser vos totaux réalisés.",
        "En début d'année, dupliquez vos lignes de plan avec la nouvelle année et mettez à "
        "jour la trésorerie initiale dans Paramètres.",
        "Le classeur ne contient aucune macro : il s'ouvre dans Excel, LibreOffice Calc, "
        "Google Sheets et les applications mobiles Excel.",
    ]),
]


def build_guide(wb):
    ws = wb.create_sheet("Guide")
    setup_sheet(ws, tab_color=NAVY_SOFT, zoom=100)
    widths(ws, {"A": 2, "B": 4, "C": 6, "D": 110, "E": 2})
    nav_bar(ws, "D", row=3)
    banner(ws, "D", "GUIDE D'UTILISATION",
           "Cinq minutes pour prendre en main Budget Smart")
    r = 5
    for title, items in GUIDE_SECTIONS:
        r += 1
        ws.merge_cells(f"B{r}:D{r}")
        c = ws[f"B{r}"]
        c.value = title
        c.font = f(12, True, WHITE)
        c.alignment = al("left", "center", indent=1)
        paint(ws, f"B{r}:D{r}", fill=solid(NAVY_MID))
        ws.row_dimensions[r].height = 24
        r += 1
        for item in items:
            ws[f"C{r}"] = "•"
            ws[f"C{r}"].font = f(11, True, GOLD)
            ws[f"C{r}"].alignment = al("center", "top")
            ws[f"D{r}"] = item
            ws[f"D{r}"].font = f(10, False, INK)
            ws[f"D{r}"].alignment = al("left", "top", wrap=True, indent=1)
            ws.row_dimensions[r].height = max(18, 14 * (1 + len(item) // 95))
            r += 1
        r += 1
    note_box(ws, "B", "D", r + 1,
             "Une question sur une cellule ? Les cellules calculées sont grises : cliquez "
             "dessus pour lire la formule. Rien n'est verrouillé, tout est modifiable.",
             height=30)
    return ws


def define_names(wb, listes_info, cat_last):
    ncur = len(D.CURRENCIES)
    names = {
        # Paramètres
        "Nom_Foyer": "'Paramètres'!$E$9",
        "Responsable": "'Paramètres'!$E$10",
        "Adresse_Entite": "'Paramètres'!$E$12",
        "Ville_Entite": "'Paramètres'!$E$13",
        "Pays_Entite": "'Paramètres'!$E$14",
        "Tel_Entite": "'Paramètres'!$E$15",
        "Email_Entite": "'Paramètres'!$E$16",
        "Site_Entite": "'Paramètres'!$E$17",
        "Devise_Base": "'Paramètres'!$E$24",
        "Symbole_Base": "'Paramètres'!$E$26",
        "Solde_Initial": "'Paramètres'!$E$33",
        "Taux_Cible": "'Paramètres'!$E$34",
        "Seuil_Alerte": "'Paramètres'!$E$35",
        "Plafond_Mensuel": "'Paramètres'!$E$36",
        "Liste_Comptes": f"'Paramètres'!$C${SC.PARAM_ACC_FIRST}:$C${SC.PARAM_ACC_LAST}",
        # Période (pilotée depuis le tableau de bord)
        "Annee_Travail": "'Tableau de Bord'!$D$4",
        "Mois_Suivi": "'Tableau de Bord'!$G$4",
        "Mois_Num": "'Tableau de Bord'!$H$4",
        # Devises
        "Codes_Devises": f"'Listes'!$B$6:$B${5 + ncur}",
        "Noms_Devises": f"'Listes'!$C$6:$C${5 + ncur}",
        "Symboles_Devises": f"'Listes'!$D$6:$D${5 + ncur}",
        "Taux_Devises": f"'Listes'!$G$6:$G${5 + ncur}",
        "Facteurs_Frequence": "'Listes'!$AO$6:$AO$11",
        # Journal
        "J_Date": f"'Journal'!$B${SC.JR_FIRST}:$B${SC.JR_LAST}",
        "J_Type": f"'Journal'!$C${SC.JR_FIRST}:$C${SC.JR_LAST}",
        "J_Module": f"'Journal'!$D${SC.JR_FIRST}:$D${SC.JR_LAST}",
        "J_Cat": f"'Journal'!$E${SC.JR_FIRST}:$E${SC.JR_LAST}",
        "J_Libelle": f"'Journal'!$F${SC.JR_FIRST}:$F${SC.JR_LAST}",
        "J_Compte": f"'Journal'!$H${SC.JR_FIRST}:$H${SC.JR_LAST}",
        "J_Entree": f"'Journal'!$J${SC.JR_FIRST}:$J${SC.JR_LAST}",
        "J_Sortie": f"'Journal'!$K${SC.JR_FIRST}:$K${SC.JR_LAST}",
        "J_Moyen": f"'Journal'!$N${SC.JR_FIRST}:$N${SC.JR_LAST}",
        "J_Projet": f"'Journal'!$O${SC.JR_FIRST}:$O${SC.JR_LAST}",
        "J_Nature": f"'Journal'!$P${SC.JR_FIRST}:$P${SC.JR_LAST}",
        "J_Statut": f"'Journal'!$Q${SC.JR_FIRST}:$Q${SC.JR_LAST}",
        "J_Montant": f"'Journal'!$U${SC.JR_FIRST}:$U${SC.JR_LAST}",
        "J_Mois": f"'Journal'!$V${SC.JR_FIRST}:$V${SC.JR_LAST}",
        "J_Annee": f"'Journal'!$W${SC.JR_FIRST}:$W${SC.JR_LAST}",
        # Estimation annuelle
        "P_Annee": f"'Estimation Annuelle'!$B${SC.PLAN_FIRST}:$B${SC.PLAN_LAST}",
        "P_Module": f"'Estimation Annuelle'!$C${SC.PLAN_FIRST}:$C${SC.PLAN_LAST}",
        "P_Type": f"'Estimation Annuelle'!$D${SC.PLAN_FIRST}:$D${SC.PLAN_LAST}",
        "P_Cat": f"'Estimation Annuelle'!$E${SC.PLAN_FIRST}:$E${SC.PLAN_LAST}",
        "P_Total": f"'Estimation Annuelle'!$R${SC.PLAN_FIRST}:$R${SC.PLAN_LAST}",
        "P_Mois": f"'Estimation Annuelle'!$F${SC.PLAN_FIRST}:$Q${SC.PLAN_LAST}",
        # Divers
        "Liste_Projets": f"'Calculs'!$B${SD.C_PRJ_FIRST}:$B${SD.C_PRJ_LAST}",
        "Dime_Active": f"'Paramètres'!$E${SC.PARAM_TOTAL_ROW + 7}",
        "Dime_Taux": f"'Paramètres'!$E${SC.PARAM_TOTAL_ROW + 8}",
        "Dime_Base": f"'Paramètres'!$E${SC.PARAM_TOTAL_ROW + 9}",
        "Perimetre": f"'Paramètres'!$E${SC.PARAM_TOTAL_ROW + 10}",
        "Perimetre_Filtre": f"'Paramètres'!$K${SC.PARAM_TOTAL_ROW + 10}",
    }
    for i in range(12):
        col = get_column_letter(6 + i)
        names[f"P_M{i + 1}"] = (f"'Estimation Annuelle'!${col}${SC.PLAN_FIRST}:"
                                f"${col}${SC.PLAN_LAST}")

    label_to_name = {
        "Modules": "Liste_Modules", "Types d'opération": "Liste_Types",
        "Nature": "Liste_Natures", "Statuts": "Liste_Statuts",
        "Moyens de paiement": "Liste_Moyens", "Fréquences": "Liste_Frequences",
        "Oui / Non": "Liste_OuiNon", "Types de bien": "Liste_TypesBien",
        "Statuts du bien": "Liste_StatutsBien", "Titres fonciers": "Liste_Titres",
        "Statuts de projet": "Liste_StatutsProjet",
        "Types de projet": "Liste_TypesProjet", "Mois": "Liste_Mois",
        "Situation familiale": "Liste_Situations",
        "Sens des types": "Sens_Types",
        "Bases de dîme": "Liste_Bases_Dime",
        "Périmètres": "Liste_Perimetres",
    }
    for label, (col, count) in listes_info["simple"].items():
        names[label_to_name[label]] = f"'Listes'!${col}$6:${col}${5 + count}"
    for idx, ref in listes_info["cat_ranges"].items():
        names[f"Cat_{idx}"] = ref

    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))


def _split_args(inner: str):
    args, depth, quoted, cur = [], 0, False, ""
    for ch in inner:
        if ch == '"':
            quoted = not quoted
        if not quoted:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                args.append(cur)
                cur = ""
                continue
        cur += ch
    args.append(cur)
    return args


def _detext(formula: str) -> str:
    """Remplace TEXT(x,"format") par des équivalents indépendants de la langue d'Excel.

    Les codes de format de TEXT() sont interprétés selon la locale : "#,##0" ou "0.0%"
    renvoient #VALEUR! sur un Excel français. FIXED() et ROUND() sont, eux, universels.
    """
    while True:
        i = formula.rfind("TEXT(")
        if i == -1:
            return formula
        j, depth, quoted = i + 5, 1, False
        while j < len(formula) and depth:
            ch = formula[j]
            if ch == '"':
                quoted = not quoted
            elif not quoted:
                if ch == "(":
                    depth += 1
                elif ch == ")":
                    depth -= 1
            j += 1
        inner = formula[i + 5:j - 1]
        args = _split_args(inner)
        expr = args[0]
        fmt = args[1].strip() if len(args) > 1 else '"#,##0"'
        if fmt in ('"0%"', '"0 %"'):
            repl = f'ROUND({expr}*100,0)&" %"'
        elif fmt in ('"0.0%"', '"0,0%"'):
            repl = f'ROUND({expr}*100,1)&" %"'
        elif "y" in fmt or "a" in fmt:
            repl = (f'RIGHT("0"&DAY({expr}),2)&"/"&RIGHT("0"&MONTH({expr}),2)'
                    f'&"/"&YEAR({expr})')
        else:
            decimals = 0
            if "." in fmt:
                decimals = len(fmt.split(".")[1].rstrip('"%'))
            repl = f"FIXED({expr},{decimals})"
        formula = formula[:i] + repl + formula[j:]


def _chart_boxes(ws):
    """Boîte en pixels de chaque graphique de la feuille."""
    from openpyxl.utils import get_column_letter
    CM = 37.8

    def colpx(i):
        w = ws.column_dimensions[get_column_letter(i)].width
        return (w if w else 8.43) * 7 + 5

    def rowpx(r):
        h = ws.row_dimensions[r].height
        return (h if h else 15) * 4 / 3

    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
    out = []
    for ch in ws._charts:
        anchor = ch.anchor
        if isinstance(anchor, str):          # ancre encore sous forme "B15"
            letter, row = coordinate_from_string(anchor)
            col0, row0 = column_index_from_string(letter) - 1, row - 1
        else:
            col0, row0 = anchor._from.col, anchor._from.row
        x = sum(colpx(i) for i in range(1, col0 + 1))
        y = sum(rowpx(r) for r in range(1, row0 + 1))
        out.append((x, y, x + ch.width * CM, y + ch.height * CM, ch))
    return out, colpx, rowpx


def check_layout(wb):
    """Refuse un graphique qui en recouvre un autre ou qui recouvre du contenu."""
    faults = []
    for ws in wb.worksheets:
        boxes, colpx, rowpx = _chart_boxes(ws)
        if not boxes:
            continue
        for i in range(len(boxes)):
            a1, b1, a2, b2, _ = boxes[i]
            for j in range(i + 1, len(boxes)):
                c1, d1, c2, d2, _ = boxes[j]
                if not (a2 <= c1 or c2 <= a1 or b2 <= d1 or d2 <= b1):
                    faults.append(f"{ws.title} : deux graphiques se superposent")
        covered = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in (None, ""):
                    continue
                x = sum(colpx(i) for i in range(1, cell.column))
                y = sum(rowpx(r) for r in range(1, cell.row))
                x2, y2 = x + colpx(cell.column), y + rowpx(cell.row)
                for a1, b1, a2, b2, _ in boxes:
                    if not (x2 <= a1 or a2 <= x or y2 <= b1 or b2 <= y):
                        covered += 1
                        break
        if covered:
            faults.append(f"{ws.title} : {covered} cellules recouvertes par un graphique")
    if faults:
        raise SystemExit("Mise en page incorrecte :\n  - " + "\n  - ".join(faults))


def check_circular(wb):
    """Refuse une formule qui se trouve dans une plage qu'elle référence.

    C'est la référence circulaire la plus courante : une carte d'indicateur
    posée par erreur à l'intérieur du tableau qu'elle additionne. Excel refuse
    alors de calculer tout le classeur.
    """
    import re
    from openpyxl.utils import column_index_from_string

    names = {}
    for n in wb.defined_names:
        try:
            names[n] = wb.defined_names[n].attr_text
        except Exception:
            pass

    RANGE = re.compile(
        r"(?:(?:'([^']+)'|([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ0-9_ ]*))!)?"
        r"\$?([A-Z]{1,3})\$?(\d+)(?:\s*:\s*\$?([A-Z]{1,3})\$?(\d+))?")

    def no_strings(text):
        return re.sub(r'"[^"]*"', '""', text)

    faults = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                body = no_strings(v)
                for nm, tgt in names.items():
                    if nm in body:
                        body = body.replace(nm, " " + tgt + " ")
                body = no_strings(body)
                for m in RANGE.finditer(body):
                    sheet = m.group(1) or m.group(2) or ws.title
                    if sheet != ws.title:
                        continue
                    c1 = column_index_from_string(m.group(3))
                    r1 = int(m.group(4))
                    c2 = column_index_from_string(m.group(5)) if m.group(5) else c1
                    r2 = int(m.group(6)) if m.group(6) else r1
                    if (min(c1, c2) <= cell.column <= max(c1, c2)
                            and min(r1, r2) <= cell.row <= max(r1, r2)):
                        faults.append(f"{ws.title}!{cell.coordinate} référence "
                                      f"{m.group(0)}, qui la contient")
                        break
    if faults:
        raise SystemExit("Références circulaires :\n  - " + "\n  - ".join(faults))


def check_merges(wb):
    """Refuse toute fusion qui en chevauche une autre.

    Excel rejette le fichier entier et propose de le « réparer » dès que deux
    plages fusionnées se recoupent. Ce contrôle fait échouer la génération avant
    que le fichier ne parte.
    """
    from openpyxl.utils.cell import range_boundaries
    faults = []
    for ws in wb.worksheets:
        rngs = [str(r) for r in ws.merged_cells.ranges]
        boxes = [range_boundaries(r) for r in rngs]
        for i in range(len(boxes)):
            c1, r1, c2, r2 = boxes[i]
            for j in range(i + 1, len(boxes)):
                d1, s1, d2, s2 = boxes[j]
                if not (c2 < d1 or d2 < c1 or r2 < s1 or s2 < r1):
                    faults.append(f"{ws.title} : {rngs[i]} chevauche {rngs[j]}")
    if faults:
        raise SystemExit("Fusions incompatibles :\n  - " + "\n  - ".join(faults))


def normalize_formulas(wb):
    """Passe finale : rend les formules indépendantes de la locale."""
    count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("=") and "TEXT(" in v:
                    c.value = _detext(v)
                    count += 1
    return count


# ==========================================================================
# Protection des feuilles
# ==========================================================================
# Mot de passe par défaut. À changer ici avant diffusion.
# Rappel honnête : la protection de feuille Excel empêche les modifications
# accidentelles et la recopie facile des formules. Ce n'est pas du chiffrement :
# un outil spécialisé lève ce verrou. Elle protège l'intégrité du classeur,
# pas le secret du contenu.
PASSWORD = "APEX-2026"

# Plages laissées saisissables, feuille par feuille. Tout le reste est verrouillé
# et les formules sont masquées dans la barre de formule.
ZONES_SAISIE = {
    "Paramètres": ["E9:F20", "E24:F24", "E29:F29", "E33:F36",
                   f"E{SC.PARAM_TOTAL_ROW + 7}:F{SC.PARAM_TOTAL_ROW + 10}",
                   f"C{SC.PARAM_ACC_FIRST}:D{SC.PARAM_ACC_LAST}",
                   f"F{SC.PARAM_ACC_FIRST}:F{SC.PARAM_ACC_LAST}"],
    "Tableau de Bord": ["D4:D4", "G4:G4"],
    "Journal": [f"B{SC.JR_FIRST}:H{SC.JR_LAST}",
                f"M{SC.JR_FIRST}:S{SC.JR_LAST}"],
    "Estimation Annuelle": [f"C{SC.PLAN_FIRST}:Q{SC.PLAN_LAST}",
                    f"V{SC.PLAN_FIRST}:V{SC.PLAN_LAST}"],
    "Récurrents": [f"B{SC.REC_FIRST}:J{SC.REC_LAST}",
                   f"O{SC.REC_FIRST}:O{SC.REC_LAST}"],
    "Mariage": ["E6:F10", f"C{SM.MAR_FIRST}:I{SM.MAR_LAST}",
                f"N{SM.MAR_FIRST}:N{SM.MAR_LAST}",
                f"P{SM.MAR_FIRST}:P{SM.MAR_LAST}", "C55:D62"],
    "Immobilier": [f"C{SM.IMM_FIRST}:F{SM.IMM_LAST}",
                   f"H{SM.IMM_FIRST}:I{SM.IMM_LAST}",
                   f"N{SM.IMM_FIRST}:O{SM.IMM_LAST}",
                   f"R{SM.IMM_FIRST}:S{SM.IMM_LAST}",
                   f"C{SM.LOT_FIRST}:E{SM.LOT_FIRST + 13}",
                   f"C{SM.ECH_FIRST}:G{SM.ECH_LAST}", "E69:F72"],
    "Business": [f"C{SM.BUS_FIRST}:H{SM.BUS_LAST}",
                 f"L{SM.BUS_FIRST}:L{SM.BUS_LAST}",
                 f"Q{SM.BUS_FIRST}:R{SM.BUS_LAST}", "E44:F45"],
    "Objectifs": [f"C{SM.OBJ_FIRST}:F{SM.OBJ_LAST}",
                  f"M{SM.OBJ_FIRST}:N{SM.OBJ_LAST}", "D27:E30"],
    "Dettes": [f"C{SM.DET_FIRST}:H{SM.DET_LAST}",
               f"N{SM.DET_FIRST}:N{SM.DET_LAST}", "E25:F25"],
    "Listes": ["E6:E40", "I6:V40", "X6:AA60"],
}


def protect_workbook(wb, password=PASSWORD):
    """Verrouille les feuilles, masque les formules, laisse les zones de saisie ouvertes."""
    verrou = Protection(locked=True, hidden=True)
    ouvert = Protection(locked=False, hidden=False)

    for ws in wb.worksheets:
        # 1. toute cellule de formule est verrouillée et sa formule masquée
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    c.protection = verrou
        # 2. les zones de saisie redeviennent modifiables
        for ref in ZONES_SAISIE.get(ws.title, []):
            for row in ws[ref]:
                for c in row:
                    c.protection = ouvert
        # 3. protection de la feuille : on autorise la sélection, le tri,
        #    les filtres et le redimensionnement des colonnes
        ws.protection = SheetProtection(
            sheet=True, objects=True, scenarios=True,
            formatCells=False, formatColumns=False, formatRows=False,
            insertColumns=True, insertRows=True, insertHyperlinks=True,
            deleteColumns=True, deleteRows=True, pivotTables=True,
            selectLockedCells=False, selectUnlockedCells=False,
            sort=False, autoFilter=False)
        ws.protection.password = password

    # 4. le moteur de calcul disparaît de la liste des onglets
    wb["Calculs"].sheet_state = "veryHidden"

    # 5. structure du classeur : ni ajout, ni suppression, ni renommage d'onglet
    wb.security = WorkbookProtection(lockStructure=True)
    wb.security.workbookPassword = password


def main(out_path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True

    listes_info = SC.build_listes(wb)
    SC.build_parametres(wb)
    SC.build_journal(wb)
    SC.build_plan(wb)
    SC.build_recurrents(wb)
    wsc, cat_last = SD.build_calculs(wb)
    SD.build_dashboard(wb, wsc)
    SD.build_annuel(wb, wsc)
    SB.build_budget12(wb)
    SM.build_mariage(wb)
    SM.build_immobilier(wb)
    SM.build_business(wb)
    SM.build_objectifs(wb)
    SM.build_dettes(wb)
    build_accueil(wb)
    build_guide(wb)

    place_logo(wb["Tableau de Bord"], "budget-smart-monogramme-512.png", "P1", 40, 40)
    place_logo(wb["Budget 12 Mois"], "budget-smart-monogramme-512.png", "AE1", 40, 40)
    place_logo(wb["Guide"], "budget-smart-monogramme-512.png", "D1", 40, 40)

    define_names(wb, listes_info, cat_last)
    normalize_formulas(wb)
    check_merges(wb)
    check_layout(wb)
    check_circular(wb)

    order = ["Accueil", "Paramètres", "Estimation Annuelle", "Journal",
             "Tableau de Bord", "Budget 12 Mois", "Synthèse Annuelle",
             "Mariage", "Business", "Immobilier",
             "Objectifs", "Dettes", "Récurrents", "Listes", "Calculs", "Guide"]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0
    protect_workbook(wb)

    wb.properties.title = f"{D.APP_NAME} {D.APP_BRAND} — {D.APP_BASELINE}"
    wb.properties.creator = "APEX AFRICA"
    wb.properties.company = "APEX AFRICA — African Premium Experience"
    wb.properties.subject = D.APP_COPYRIGHT
    wb.properties.keywords = "budget, APEX AFRICA, Budget Smart, finances"
    wb.properties.category = "Outil financier"
    wb.properties.description = (
        "Classeur de planification et de suivi budgétaire multi-devises : budget général, "
        "mariage, immobilier & terrain, business et projets.")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    target = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else OUTPUT_DEFAULT
    path = main(target)
    print(f"✓ {path}")

"""Feuilles socles d'APEX Budget : Listes, Paramètres, Journal, Plan annuel, Récurrents."""
from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import apex_data as D
from apex_style import (
    BG, BLUE, BLUE_SOFT, CARD, FMT_DATE, FMT_INT, FMT_MONEY, FMT_PCT, FMT_RATE,
    GREEN_SOFT, ORANGE_SOFT, RED_SOFT,
    GOLD, GOLD_SOFT, GREEN, INK, LINE, MUTED, NAVY, NAVY_MID, NAVY_SOFT, ORANGE,
    PURPLE, RED, TEAL, WHITE, al, banner, box_border, cell, color_scale_ok,
    currency_format, databar, f, heights, hide_zeros, input_style, kpi_card,
    link, note_box,
    nav_bar, paint, section, setup_sheet, solid, table_header, widths, zebra,
)

# --------------------------------------------------------------------------
# Géométrie du Journal
# --------------------------------------------------------------------------
JR_BLOCS = 11
JR_HEADER = 12
JR_FIRST = 13
JR_LAST = 3012

# Géométrie de la feuille Paramètres
PARAM_ACC_FIRST = 41
PARAM_ACC_LAST = 60
PARAM_TOTAL_ROW = 61

# Critères « écriture réellement encaissée / décaissée »
REAL = 'J_Statut,"<>Annulé",J_Statut,"<>Prévu"'


def build_listes(wb):
    ws = wb.create_sheet("Listes")
    setup_sheet(ws, tab_color=MUTED, freeze="B5")
    nav_bar(ws, "AA", row=3, current=None)
    banner(ws, "AA", "RÉFÉRENTIELS & LISTES DE CHOIX",
           "Cours des devises, catégories et listes déroulantes utilisées par tout le classeur")

    widths(ws, {"A": 2, "B": 9, "C": 30, "D": 10, "E": 16, "F": 9, "G": 14, "H": 2})
    for col in "IJKLMNOPQRSTUV":
        widths(ws, {col: 20})
    widths(ws, {"W": 2, "X": 32, "Y": 30, "Z": 32, "AA": 30, "AB": 2,
                "AC": 16, "AD": 20, "AE": 16})

    # ---- Devises -------------------------------------------------------
    section(ws, 4, "B", "G", "DEVISES — cours indicatifs, à actualiser librement", GOLD, INK)
    table_header(ws, 5, "B", "G",
                 ["Code", "Devise", "Symbole", "Cours indicatif\n(1 unité = X XOF)",
                  "Déci-\nmales", "Taux vers la\ndevise de base"])
    r = 6
    for code, name, sym, rate, dec in D.CURRENCIES:
        ws.cell(row=r, column=2, value=code).font = f(10, True, NAVY)
        ws.cell(row=r, column=3, value=name).font = f(10)
        ws.cell(row=r, column=4, value=sym).alignment = al("center")
        c = ws.cell(row=r, column=5, value=rate)
        c.number_format = FMT_RATE
        ws.cell(row=r, column=6, value=dec).alignment = al("center")
        g = ws.cell(
            row=r, column=7,
            value=f"=IFERROR($E{r}/INDEX($E$6:$E${5 + len(D.CURRENCIES)},"
                  f"MATCH(Devise_Base,$B$6:$B${5 + len(D.CURRENCIES)},0)),1)")
        g.number_format = FMT_RATE
        g.font = f(10, True, BLUE)
        r += 1
    last_cur = r - 1
    paint(ws, f"B6:G{last_cur}", border=box_border(LINE))
    input_style(ws, f"E6:E{last_cur}")
    zebra(ws, f"B6:G{last_cur}", 6)
    note_box(ws, "B", "G", last_cur + 2,
             "Les cours sont donnés à titre indicatif. Modifiez la colonne « Cours indicatif » "
             "pour coller au marché : le taux vers votre devise de base se recalcule seul.",
             height=30)

    # ---- Listes simples ------------------------------------------------
    simple = [
        ("I", "Modules", D.MODULES),
        ("J", "Types d'opération", D.TYPES),
        ("K", "Nature", D.NATURES),
        ("L", "Statuts", D.STATUTS),
        ("M", "Moyens de paiement", D.MOYENS),
        ("N", "Fréquences", D.FREQUENCES),
        ("O", "Oui / Non", D.OUI_NON),
        ("P", "Types de bien", D.TYPES_BIENS),
        ("Q", "Statuts du bien", D.STATUTS_BIEN),
        ("R", "Titres fonciers", D.TITRES_FONCIERS),
        ("S", "Statuts de projet", D.STATUTS_PROJET),
        ("T", "Types de projet", D.TYPES_PROJET),
        ("U", "Mois", D.MONTHS),
        ("V", "Situation familiale", ["Célibataire", "En couple", "Marié(e)",
                                      "Parent seul", "Étudiant(e)", "Autre"]),
        ("AC", "Sens des types", D.SENS_TYPES),
        ("AD", "Bases de dîme", D.DIME_BASES),
        ("AE", "Périmètres", D.PERIMETRES),
    ]
    for col, title, values in simple:
        c = ws[f"{col}5"]
        c.value = title
        c.fill = solid(NAVY_SOFT)
        c.font = f(9, True, WHITE)
        c.alignment = al("center", "center", wrap=True)
        c.border = box_border(NAVY_SOFT)
        for i, v in enumerate(values):
            cc = ws.cell(row=6 + i, column=c.column, value=v)
            cc.font = f(9)
            cc.border = box_border(LINE)

    # facteur de mensualisation des récurrents (à côté des fréquences)
    factors = {"Hebdomadaire": 4.3333, "Mensuelle": 1, "Bimestrielle": 0.5,
               "Trimestrielle": 1 / 3, "Semestrielle": 1 / 6, "Annuelle": 1 / 12}
    fc = ws["N4"]
    fc.value = "Facteur mensuel →"
    fc.font = f(8, False, MUTED)
    fc.alignment = al("right")
    for i, fr in enumerate(D.FREQUENCES):
        c = ws.cell(row=6 + i, column=41, value=factors[fr])  # colonne AO
        c.number_format = "0.0000"
    ws["AO5"] = "Facteur mensuel"
    ws["AO5"].font = f(8, True, MUTED)

    # ---- Catégories ----------------------------------------------------
    section(ws, 4, "X", "AA", "CATÉGORIES PAR MODULE", GOLD, INK)
    cat_cols = [("X", "Général", D.CAT_GENERAL), ("Y", "Mariage", D.CAT_MARIAGE),
                ("Z", "Immobilier", D.CAT_IMMOBILIER), ("AA", "Business", D.CAT_BUSINESS)]
    cat_ranges = {}
    for idx, (col, title, values) in enumerate(cat_cols, start=1):
        h = ws[f"{col}5"]
        h.value = title
        h.fill = solid(NAVY_SOFT)
        h.font = f(9, True, WHITE)
        h.alignment = al("center", "center")
        h.border = box_border(NAVY_SOFT)
        for i, v in enumerate(values):
            cc = ws.cell(row=6 + i, column=h.column, value=v)
            cc.font = f(9)
            cc.border = box_border(LINE)
        cat_ranges[idx] = f"'Listes'!${col}$6:${col}${5 + len(values)}"

    ws.sheet_state = "visible"
    return {
        "last_cur": last_cur,
        "cat_ranges": cat_ranges,
        "simple": {t: (c, len(v)) for c, t, v in simple},
    }


def build_parametres(wb):
    ws = wb.create_sheet("Paramètres")
    setup_sheet(ws, tab_color=GOLD, freeze="A5")
    banner(ws, "J", "MES INFORMATIONS",
           "Étape ① — renseignez ceci une seule fois : tout le classeur s'y adapte")
    nav_bar(ws, "J", row=3, current="Paramètres")
    widths(ws, {"A": 2, "B": 3, "C": 30, "D": 12, "E": 20, "F": 14, "G": 20,
                "H": 26, "I": 22, "J": 26})

    def field(row, label, value=None, formula=None, numfmt=None, help_text="",
              dv=None, editable=True):
        ws.merge_cells(f"C{row}:D{row}")
        lab = ws[f"C{row}"]
        lab.value = label
        lab.font = f(10, True, INK)
        lab.alignment = al("left", "center", indent=1)
        ws.merge_cells(f"E{row}:F{row}")
        v = ws[f"E{row}"]
        v.value = formula if formula is not None else value
        v.alignment = al("left", "center", indent=1)
        v.number_format = numfmt or "General"
        if editable:
            input_style(ws, f"E{row}:F{row}")
            v.font = f(11, True, INK)
        else:
            paint(ws, f"E{row}:F{row}", fill=solid(BG), border=box_border(LINE))
            v.font = f(11, True, BLUE)
        ws.merge_cells(f"G{row}:J{row}")
        h = ws[f"G{row}"]
        h.value = help_text
        h.font = f(9, False, MUTED, italic=True)
        h.alignment = al("left", "center", wrap=True, indent=1)
        ws.row_dimensions[row].height = 22
        if dv is not None:
            dv.add(v)
        return v

    note_box(ws, "C", "J", 5,
             "Seules les cases sur fond JAUNE PÂLE sont modifiables. Le reste de la "
             "feuille est verrouillé pour protéger les formules — mot de passe "
             "APEX-2026 (Révision ▸ Ôter la protection de la feuille).", height=30)

    # ① Identité & coordonnées
    section(ws, 7, "C", "J", "①  IDENTITÉ & COORDONNÉES")
    dv_sit = DataValidation(type="list", formula1="Liste_Situations", allow_blank=True)
    ws.add_data_validation(dv_sit)
    field(9, "Raison sociale / Nom", value="APEX AFRICA",
          help_text="Apparaît en tête du tableau de bord et des synthèses.")
    field(10, "Responsable / Titulaire", value="",
          help_text="La personne qui tient le budget.")
    field(11, "Activité / Fonction", value="")
    field(12, "Adresse", value="")
    field(13, "Ville", value="Abidjan")
    field(14, "Pays", value="Côte d'Ivoire")
    field(15, "Téléphone", value="", numfmt="@",
          help_text="Saisissez-le tel quel, avec l'indicatif : +225 …")
    field(16, "Email", value="contact@apxafrica.com")
    field(17, "Site web", value="www.apxafrica.com")
    field(18, "RCCM / IFU / N° d'identification", value="",
          help_text="Facultatif — utile si le classeur sert à une structure.")
    field(19, "Situation familiale", value="", dv=dv_sit,
          help_text="Pour un budget personnel. Laissez vide sinon.")
    field(20, "Personnes à charge", value=0, numfmt=FMT_INT,
          help_text="Sert au calcul de la dépense moyenne par personne.")

    # ② Devise & période
    section(ws, 22, "C", "J", "②  DEVISE & PÉRIODE")
    dv_dev = DataValidation(type="list", formula1="Codes_Devises", allow_blank=False)
    ws.add_data_validation(dv_dev)
    dv_mois = DataValidation(type="list", formula1="Liste_Mois", allow_blank=False)
    ws.add_data_validation(dv_mois)
    field(24, "Devise de base", value="XOF", dv=dv_dev,
          help_text="Toutes les synthèses sont exprimées dans cette devise. Changez-la : "
                    "symboles, taux et totaux suivent. À fixer avant de commencer à saisir.")
    field(25, "Nom de la devise",
          formula="=IFERROR(INDEX(Noms_Devises,MATCH(Devise_Base,Codes_Devises,0)),\"\")",
          editable=False)
    field(26, "Symbole",
          formula="=IFERROR(INDEX(Symboles_Devises,MATCH(Devise_Base,Codes_Devises,0)),\"\")",
          editable=False)
    field(27, "Année de travail", formula="=Annee_Travail", numfmt="0", editable=False,
          help_text="Se règle en haut du Tableau de bord — toutes les feuilles suivent.")
    field(28, "Mois de suivi", formula="=Mois_Suivi", editable=False,
          help_text="Se règle également en haut du Tableau de bord.")
    field(29, "Mois de démarrage du budget", value="Janvier", dv=dv_mois,
          help_text="Utile si votre exercice ne démarre pas en janvier.")

    # ③ Soldes & objectifs
    section(ws, 31, "C", "J", "③  SOLDES & OBJECTIFS")
    field(33, "Trésorerie initiale", value=0, numfmt=FMT_MONEY,
          help_text="Total de vos disponibilités au 1er janvier de l'année de travail.")
    field(34, "Taux d'épargne cible", value=0.20, numfmt=FMT_PCT,
          help_text="Part des revenus que vous visez à épargner chaque mois.")
    field(35, "Seuil d'alerte budget", value=0.90, numfmt=FMT_PCT,
          help_text="Au-delà de ce pourcentage du budget consommé, l'alerte passe au rouge.")
    field(36, "Plafond de dépenses mensuel", value=0, numfmt=FMT_MONEY,
          help_text="Laissez 0 si vous n'en fixez pas.")
    currency_format(ws, "E33:F33")
    currency_format(ws, "E36:F36")

    # ④ Comptes
    section(ws, 38, "C", "J", "④  COMPTES, CAISSES & PORTEFEUILLES")
    table_header(ws, 40, "C", "H",
                 ["Compte / Portefeuille", "Nature", "Devise", "Solde d'ouverture",
                  "Mouvements enregistrés", "Solde actuel calculé"])
    dv_nature_cpt = DataValidation(
        type="list",
        formula1=('"Compte courant,Épargne,Épargne bloquée,Mobile Money,Espèces,'
                  'Business,Investissement,Autre"'),
        allow_blank=True)
    ws.add_data_validation(dv_nature_cpt)
    for i in range(20):
        r = PARAM_ACC_FIRST + i
        ws.cell(row=r, column=3,
                value=D.COMPTES_DEFAUT[i] if i < len(D.COMPTES_DEFAUT) else None)
        ws.cell(row=r, column=4)
        ws.cell(row=r, column=5, value="=IF($C{0}=\"\",\"\",Devise_Base)".format(r))
        so = ws.cell(row=r, column=6, value=0 if i < len(D.COMPTES_DEFAUT) else None)
        so.number_format = FMT_MONEY
        mv = ws.cell(
            row=r, column=7,
            value=(f'=IF($C{r}="","",SUMIFS(J_Montant,J_Compte,$C{r},J_Type,"Revenu",{REAL})'
                   f'-SUMIFS(J_Montant,J_Compte,$C{r},J_Type,"Dépense",{REAL})'
                   f'-SUMIFS(J_Montant,J_Compte,$C{r},J_Type,"Épargne",{REAL})'
                   f'-SUMIFS(J_Montant,J_Compte,$C{r},J_Type,"Investissement",{REAL})'
                   f'-SUMIFS(J_Montant,J_Compte,$C{r},J_Type,"Remboursement crédit",{REAL}))'))
        mv.number_format = FMT_MONEY
        sc = ws.cell(row=r, column=8, value=f'=IF($C{r}="","",N($F{r})+N($G{r}))')
        sc.number_format = FMT_MONEY
        sc.font = f(10, True, NAVY)
        dv_nature_cpt.add(ws.cell(row=r, column=4))
    A0, A1, AT = PARAM_ACC_FIRST, PARAM_ACC_LAST, PARAM_TOTAL_ROW
    paint(ws, f"C{A0}:H{A1}", border=box_border(LINE))
    input_style(ws, f"C{A0}:D{A1}")
    input_style(ws, f"F{A0}:F{A1}")
    paint(ws, f"E{A0}:E{A1}", fill=solid(BG))
    zebra(ws, f"C{A0}:H{A1}", A0)
    color_scale_ok(ws, f"H{A0}:H{A1}")
    currency_format(ws, f"F{A0}:H{A1}")
    ws[f"C{AT}"] = "TOTAL DISPONIBLE"
    ws[f"C{AT}"].font = f(11, True, WHITE)
    ws.merge_cells(f"C{AT}:G{AT}")
    paint(ws, f"C{AT}:G{AT}", fill=solid(NAVY))
    t = ws[f"H{AT}"]
    t.value = f"=SUM(H{A0}:H{A1})"
    t.number_format = FMT_MONEY
    t.font = f(12, True, WHITE)
    t.fill = solid(NAVY)
    t.alignment = al("right")

    note_box(ws, "C", "J", PARAM_TOTAL_ROW + 2,
             "Astuce : le solde de chaque compte se met à jour tout seul à partir du Journal. "
             "Les transferts entre vos propres comptes sont neutres et ne modifient pas le total.",
             height=30)

    # ⑤ Dîme & périmètre
    D0 = PARAM_TOTAL_ROW + 5
    section(ws, D0, "C", "J", "⑤  DÎME & PÉRIMÈTRE D'ANALYSE")
    dv_oui = DataValidation(type="list", formula1="Liste_OuiNon", allow_blank=False)
    ws.add_data_validation(dv_oui)
    dv_base = DataValidation(type="list", formula1="Liste_Bases_Dime", allow_blank=True)
    ws.add_data_validation(dv_base)
    dv_per = DataValidation(type="list", formula1="Liste_Perimetres", allow_blank=False)
    ws.add_data_validation(dv_per)
    field(D0 + 2, "Dîme activée", value="Oui", dv=dv_oui,
          help_text="Si Oui, le classeur calcule seul la dîme due chaque mois et "
                    "suit ce qui a été versé.")
    field(D0 + 3, "Taux de la dîme", value=0.10, numfmt=FMT_PCT,
          help_text="10 % par défaut.")
    field(D0 + 4, "Assiette de calcul", value="Tous les revenus", dv=dv_base,
          help_text="Sur quels revenus la dîme est calculée.")
    field(D0 + 5, "Périmètre d'analyse", value="Général seul", dv=dv_per,
          help_text="Ce que comptent le tableau de bord et la synthèse annuelle. "
                    "« Général seul » laisse le mariage, l'immobilier et le business "
                    "à leurs propres feuilles.")
    # cellule technique : traduit le périmètre en critère SUMIFS
    ws[f"K{D0 + 5}"] = (f'=IF($E${D0 + 5}="Tout confondu","*",'
                        f'IF($E${D0 + 5}="Général seul","Général",$E${D0 + 5}))')
    ws[f"K{D0 + 5}"].font = f(8, False, MUTED)
    note_box(ws, "C", "J", D0 + 7,
             "La dîme ne peut pas s'écrire toute seule dans le Journal — Excel sans macro "
             "ne crée pas de ligne. En revanche le montant dû est calculé automatiquement, "
             "affiché sur le tableau de bord et comparé à ce que vous avez réellement "
             f"versé (catégorie « {D.CAT_DIME} »).", height=32)

    section(ws, D0 + 10, "C", "J", "⑥  RACCOURCIS")
    shortcuts = [("Journal des opérations", "Journal"), ("Plan annuel", "Plan Annuel"),
                 ("Tableau de bord", "Tableau de Bord"), ("Cours des devises", "Listes")]
    for i, (label, sheet) in enumerate(shortcuts):
        r = PARAM_TOTAL_ROW + 12 + i
        link(ws, f"C{r}", "➜  " + label, sheet, "A1", fill_color=BLUE_SOFT)
        ws.merge_cells(f"C{r}:E{r}")
    return ws


def build_journal(wb):
    ws = wb.create_sheet("Journal")
    setup_sheet(ws, tab_color=BLUE, freeze=f"C{JR_FIRST}", zoom=85)
    nav_bar(ws, "W", row=3, current="Journal")
    banner(ws, "W", "BUDGET SMART — JOURNAL QUOTIDIEN",
           "Une ligne par opération. Vous ne remplissez que le premier bloc ; "
           "le sens, l'entrée, la sortie et le solde se calculent seuls.")

    widths(ws, {"A": 2, "B": 12, "C": 20, "D": 13, "E": 28, "F": 30, "G": 14, "H": 22,
                "I": 14, "J": 14, "K": 14, "L": 16,
                "M": 8, "N": 17, "O": 22, "P": 13, "Q": 11, "R": 10, "S": 26,
                "T": 10, "U": 16, "V": 7, "W": 8})

    note_box(ws, "B", "W", 5,
             "Saisissez toujours un montant POSITIF dans « Montant ». C'est la colonne "
             "« Type » qui dit si l'argent entre ou sort : le classeur remplit alors tout "
             "seul les colonnes ENTRÉE, SORTIE et SOLDE CUMULÉ. Statut « Prévu » = "
             "engagement futur, compté à part. Les cases jaunes sont les seules à remplir.",
             height=32)

    # ---- indicateurs de tête ------------------------------------------
    kpi_card(ws, 7, 2, "ÉCRITURES SAISIES", "=COUNT(J_Date)", FMT_INT, NAVY, span=3,
             note_formula='="Dernière saisie : "&IF(COUNT(J_Date)=0,"—",'
                          'RIGHT("0"&DAY(MAX(J_Date)),2)&"/"&RIGHT("0"&MONTH(MAX(J_Date)),2)'
                          '&"/"&YEAR(MAX(J_Date)))')
    kpi_card(ws, 7, 5, "TOTAL DES ENTRÉES",
             f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Type,"Revenu",{REAL})',
             FMT_MONEY, GREEN, span=3, note_formula='="Année "&Annee_Travail')
    kpi_card(ws, 7, 8, "TOTAL DES SORTIES",
             f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Type,"<>Revenu",'
             f'J_Type,"<>Transfert",{REAL})',
             FMT_MONEY, RED, span=3, note="Dépenses + épargne + crédits")
    kpi_card(ws, 7, 11, "SOLDE DE L'ANNÉE",
             f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Type,"Revenu",{REAL})'
             f'-SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Type,"<>Revenu",'
             f'J_Type,"<>Transfert",{REAL})',
             FMT_MONEY, BLUE, span=3)
    kpi_card(ws, 7, 14, "MIS DE CÔTÉ",
             f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Type,"Épargne",{REAL})'
             f'+SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Type,"Investissement",{REAL})',
             FMT_MONEY, TEAL, span=3, note="Épargne + investissement")
    kpi_card(ws, 7, 17, "EN ATTENTE / PRÉVU",
             '=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Statut,"Prévu")'
             '+SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Statut,"En attente")',
             FMT_MONEY, ORANGE, span=3, note="Engagements non décaissés")
    for rng in ("E8:G8", "H8:J8", "K8:M8", "N8:P8", "Q8:S8"):
        currency_format(ws, rng)

    # ---- bandeaux de blocs --------------------------------------------
    blocs = [("B", "H", "①  À REMPLIR", GOLD, INK),
             ("I", "L", "②  CALCULÉ POUR VOUS", NAVY),
             ("M", "S", "③  COMPLÉMENTS — facultatifs", NAVY_SOFT),
             ("T", "W", "④  TECHNIQUE", MUTED)]
    for c1, c2, txt, color, *fg in blocs:
        ws.merge_cells(f"{c1}{JR_BLOCS}:{c2}{JR_BLOCS}")
        c = ws[f"{c1}{JR_BLOCS}"]
        c.value = txt
        c.font = f(10, True, fg[0] if fg else WHITE)
        c.alignment = al("center", "center")
        paint(ws, f"{c1}{JR_BLOCS}:{c2}{JR_BLOCS}", fill=solid(color))
    ws.row_dimensions[JR_BLOCS].height = 20

    headers = ["Date", "Type d'opération", "Module", "Catégorie", "Libellé / Description",
               "Montant", "Compte",
               "Sens", "ENTRÉE", "SORTIE", "Solde cumulé",
               "Devise", "Moyen de paiement", "Projet / Bien", "Nature", "Statut",
               "Récurrent", "Notes",
               "Taux", "Montant (devise de base)", "Mois", "Année"]
    table_header(ws, JR_HEADER, "B", "W", headers, height=36)
    ws.auto_filter.ref = f"B{JR_HEADER}:W{JR_LAST}"

    for r in range(JR_FIRST, JR_LAST + 1):
        ws.cell(row=r, column=2).number_format = FMT_DATE
        ws.cell(row=r, column=7).number_format = FMT_MONEY
        ws.cell(row=r, column=9,
                value=(f'=IF($C{r}="","",IFERROR(INDEX(Sens_Types,'
                       f'MATCH($C{r},Liste_Types,0)),""))'))
        ws.cell(row=r, column=10,
                value=f'=IF($I{r}="▲ ENTRÉE",$U{r},"")').number_format = FMT_MONEY
        ws.cell(row=r, column=11,
                value=f'=IF($I{r}="▼ SORTIE",$U{r},"")').number_format = FMT_MONEY
        prev = "Solde_Initial" if r == JR_FIRST else f"$L{r - 1}"
        ws.cell(row=r, column=12,
                value=f"={prev}+N($J{r})-N($K{r})").number_format = FMT_MONEY
        ws.cell(
            row=r, column=20,
            value=(f'=IF($G{r}="","",IFERROR(INDEX(Taux_Devises,'
                   f'MATCH(IF($M{r}="",Devise_Base,$M{r}),Codes_Devises,0)),1))')
        ).number_format = FMT_RATE
        u = ws.cell(row=r, column=21, value=f'=IF($G{r}="","",ROUND($G{r}*$T{r},2))')
        u.number_format = FMT_MONEY
        ws.cell(row=r, column=22,
                value=f'=IF($B{r}="","",MONTH($B{r}))').number_format = "0"
        ws.cell(row=r, column=23,
                value=f'=IF($B{r}="","",YEAR($B{r}))').number_format = "0"

    F, L = JR_FIRST, JR_LAST
    paint(ws, f"B{F}:W{L}", border=box_border(LINE), font=f(10))
    input_style(ws, f"B{F}:H{L}")
    input_style(ws, f"M{F}:S{L}")
    paint(ws, f"I{F}:L{L}", fill=solid(BG))
    paint(ws, f"T{F}:W{L}", fill=solid(BG), font=f(9, False, MUTED))
    paint(ws, f"I{F}:I{L}", align=al("center"), font=f(9, True))
    paint(ws, f"J{F}:J{L}", font=f(10, True, GREEN))
    paint(ws, f"K{F}:K{L}", font=f(10, True, RED))
    paint(ws, f"L{F}:L{L}", font=f(10, True, NAVY))
    for col in ("M", "Q", "R", "V", "W"):
        paint(ws, f"{col}{F}:{col}{L}", align=al("center"))
    zebra(ws, f"B{F}:W{L}", F)
    currency_format(ws, f"G{F}:G{L}", hide_zero=True)
    currency_format(ws, f"J{F}:L{L}", hide_zero=True)
    currency_format(ws, f"U{F}:U{L}", hide_zero=True)
    ws.column_dimensions.group("T", "W", outline_level=1, hidden=False)

    # ---- listes déroulantes -------------------------------------------
    dvs = [("C", "Liste_Types"), ("D", "Liste_Modules"),
           ("E", 'INDIRECT("Cat_"&IFERROR(MATCH($D{row},Liste_Modules,0),1))'),
           ("H", "Liste_Comptes"), ("M", "Codes_Devises"), ("N", "Liste_Moyens"),
           ("O", "Liste_Projets"), ("P", "Liste_Natures"), ("Q", "Liste_Statuts"),
           ("R", "Liste_OuiNon")]
    for col, formula in dvs:
        f1 = formula.format(row=F) if "{row}" in formula else formula
        dv = DataValidation(type="list", formula1=f1, allow_blank=True,
                            showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}{F}:{col}{L}")

    dv_date = DataValidation(type="date", operator="between",
                             formula1="DATE(2000,1,1)", formula2="DATE(2100,12,31)",
                             allow_blank=True, showErrorMessage=True,
                             errorTitle="Date invalide",
                             error="Saisissez une date au format jj/mm/aaaa.")
    ws.add_data_validation(dv_date)
    dv_date.add(f"B{F}:B{L}")

    dv_amount = DataValidation(type="decimal", operator="greaterThanOrEqual",
                               formula1="0", allow_blank=True, showErrorMessage=True,
                               errorTitle="Montant invalide",
                               error="Saisissez un montant POSITIF. C'est la colonne "
                                     "« Type d'opération » qui indique si l'argent "
                                     "entre ou sort.")
    ws.add_data_validation(dv_amount)
    dv_amount.add(f"G{F}:G{L}")

    # ---- signaux visuels ----------------------------------------------
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.styles.numbers import NumberFormat

    def rule(formula, font=None, fill=None, numfmt=None, stop=False):
        nf = NumberFormat(numFmtId=900, formatCode=numfmt) if numfmt else None
        r = Rule(type="expression",
                 dxf=DifferentialStyle(font=font, fill=fill, numFmt=nf),
                 stopIfTrue=stop)
        r.formula = [formula]
        return r

    rng = f"B{F}:W{L}"
    ws.conditional_formatting.add(rng, rule(f'$I{F}="▲ ENTRÉE"',
                                            fill=PatternFill(bgColor=GREEN_SOFT)))
    ws.conditional_formatting.add(rng, rule(f'$Q{F}="Prévu"',
                                            fill=PatternFill(bgColor=ORANGE_SOFT)))
    ws.conditional_formatting.add(rng, rule(f'$Q{F}="En attente"',
                                            fill=PatternFill(bgColor="FFF7D6")))
    ws.conditional_formatting.add(rng, rule(f'$Q{F}="Annulé"',
                                            font=Font(color=MUTED, strike=True)))
    # le solde cumulé disparaît sur les lignes vides
    ws.conditional_formatting.add(f"L{F}:L{L}",
                                  rule(f'$B{F}=""', numfmt=";;;", stop=True))
    ws.conditional_formatting.add(f"L{F}:L{L}",
                                  rule(f'AND($B{F}<>"",$L{F}<0)',
                                       font=Font(color=RED, bold=True)))
    # lignes incomplètes
    for col in ("C", "G"):
        ws.conditional_formatting.add(
            f"{col}{F}:{col}{L}",
            rule(f'AND($B{F}<>"",${col}{F}="")', fill=PatternFill(bgColor=RED_SOFT)))
    return ws


SHEET_PLAN = "Estimation Annuelle"
PLAN_HEADER = 11
PLAN_FIRST = 12
PLAN_LAST = 211


def build_plan(wb):
    ws = wb.create_sheet(SHEET_PLAN)
    setup_sheet(ws, tab_color=PURPLE, freeze=f"F{PLAN_FIRST}", zoom=85)
    banner(ws, "V", "ESTIMATION ANNUELLE",
           "Étape ② — ce que vous prévoyez pour l'année, mois par mois. "
           "Le réalisé remonte tout seul du Journal.")
    nav_bar(ws, "V", row=3, current="Estimation Annuelle")
    widths(ws, {"A": 2, "B": 8, "C": 13, "D": 13, "E": 34})
    for i in range(12):
        widths(ws, {get_column_letter(6 + i): 12})
    widths(ws, {"R": 15, "S": 15, "T": 14, "U": 12, "V": 26})

    kpi_card(ws, 7, 2, "REVENUS PRÉVUS",
             f'=SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Revenu")',
             FMT_MONEY, GREEN, span=3)
    kpi_card(ws, 7, 5, "DÉPENSES PRÉVUES",
             f'=SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Dépense")',
             FMT_MONEY, RED, span=3)
    kpi_card(ws, 7, 8, "ÉPARGNE PRÉVUE",
             f'=SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Épargne")'
             f'+SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Investissement")',
             FMT_MONEY, TEAL, span=3)
    kpi_card(ws, 7, 11, "SOLDE PRÉVISIONNEL",
             f'=SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Revenu")'
             f'-SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Dépense")'
             f'-SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Remboursement crédit")',
             FMT_MONEY, BLUE, span=3)
    kpi_card(ws, 7, 14, "RÉALISÉ / PRÉVU (dépenses)",
             f'=IFERROR(SUMIFS($S${PLAN_FIRST}:$S${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Dépense")'
             f'/SUMIFS($R${PLAN_FIRST}:$R${PLAN_LAST},$B${PLAN_FIRST}:$B${PLAN_LAST},Annee_Travail,$D${PLAN_FIRST}:$D${PLAN_LAST},"Dépense"),0)',
             FMT_PCT, ORANGE, span=3)
    for rng in ("B8:D8", "E8:G8", "H8:J8", "K8:M8"):
        currency_format(ws, rng)

    headers = (["Année", "Module", "Type", "Catégorie"] + D.MONTHS_SHORT +
               ["Total prévu", "Réalisé", "Écart", "% consommé", "Commentaire"])
    table_header(ws, PLAN_HEADER, "B", "V", headers, height=32)

    seed = (
        [("Général", "Revenu", c) for c in
         ["Salaire", "Prime / 13e mois", "Freelance / Consulting", "Commerce / Ventes",
          "Loyers perçus", "Autre revenu"]] +
        [("Général", "Dépense", c) for c in
         ["Loyer", "Électricité", "Eau", "Internet / Box", "Téléphone",
          "Alimentation / Courses", "Restaurants / Déjeuners", "Carburant",
          "Transport en commun", "Taxi / VTC", "Entretien véhicule", "Assurance véhicule",
          "Santé / Pharmacie", "Assurance santé / Mutuelle", "Habillement",
          "Coiffure & soins", "Scolarité / Éducation", "Garde d'enfants",
          "Sport / Fitness", "Loisirs & sorties", "Abonnements (streaming, apps)",
          "Dîme / Offrandes", "Dons & solidarité", "Cadeaux", "Voyages / Vacances",
          "Impôts & taxes", "Frais bancaires", "Aide à la famille élargie",
          "Imprévus / Divers"]] +
        [("Général", "Épargne", "Épargne (versement)"),
         ("Général", "Investissement", "Investissement"),
         ("Général", "Remboursement crédit", "Remboursement de crédit")] +
        [("Mariage", "Dépense", c) for c in
         ["Dot & coutume", "Salle & mobilier", "Traiteur & repas", "Décoration",
          "Tenue de la mariée", "Tenue du marié", "Photographe", "DJ & animation",
          "Lune de miel", "Imprévus mariage"]] +
        [("Immobilier", "Dépense", c) for c in
         ["Prix d'achat terrain", "Frais de notaire", "Matériaux de construction",
          "Main d'œuvre", "Mensualité crédit immobilier", "Taxe foncière",
          "Entretien & réparations"]] +
        [("Immobilier", "Revenu", "Loyers perçus")] +
        [("Business", "Revenu", c) for c in
         ["Chiffre d'affaires", "Prestation de service", "Vente de produits"]] +
        [("Business", "Dépense", c) for c in
         ["Achats de marchandises", "Salaires & charges", "Loyer professionnel",
          "Marketing & publicité", "Logiciels & abonnements", "Transport & logistique",
          "Impôts & taxes pro", "Divers business"]]
    )

    for i in range(PLAN_LAST - PLAN_FIRST + 1):
        r = PLAN_FIRST + i
        ws.cell(row=r, column=2, value="=Annee_Travail").number_format = "0"
        if i < len(seed):
            mod, typ, cat = seed[i]
            ws.cell(row=r, column=3, value=mod)
            ws.cell(row=r, column=4, value=typ)
            ws.cell(row=r, column=5, value=cat)
        for m in range(12):
            ws.cell(row=r, column=6 + m).number_format = FMT_MONEY
        ws.cell(row=r, column=18, value=f"=IF($E{r}=\"\",\"\",SUM($F{r}:$Q{r}))"
                ).number_format = FMT_MONEY
        ws.cell(
            row=r, column=19,
            value=(f'=IF($E{r}="","",SUMIFS(J_Montant,J_Annee,$B{r},J_Module,$C{r},'
                   f'J_Type,$D{r},J_Cat,$E{r},{REAL}))')
        ).number_format = FMT_MONEY
        ws.cell(row=r, column=20,
                value=f'=IF($E{r}="","",IF($D{r}="Revenu",$S{r}-$R{r},$R{r}-$S{r}))'
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=21,
                value=f'=IF(OR($E{r}="",$R{r}=0),"",$S{r}/$R{r})').number_format = FMT_PCT

    paint(ws, f"B{PLAN_FIRST}:V{PLAN_LAST}", border=box_border(LINE), font=f(10))
    input_style(ws, f"C{PLAN_FIRST}:Q{PLAN_LAST}")
    paint(ws, f"B{PLAN_FIRST}:B{PLAN_LAST}", fill=solid(BG), align=al("center"))
    paint(ws, f"R{PLAN_FIRST}:U{PLAN_LAST}", fill=solid(BG))
    paint(ws, f"R{PLAN_FIRST}:R{PLAN_LAST}", font=f(10, True, NAVY))
    zebra(ws, f"B{PLAN_FIRST}:V{PLAN_LAST}", PLAN_FIRST)
    color_scale_ok(ws, f"T{PLAN_FIRST}:T{PLAN_LAST}")
    databar(ws, f"U{PLAN_FIRST}:U{PLAN_LAST}", PURPLE)
    hide_zeros(ws, f"F{PLAN_FIRST}:T{PLAN_LAST}")
    currency_format(ws, f"F{PLAN_FIRST}:T{PLAN_LAST}", hide_zero=True)

    for col, formula in [("C", "Liste_Modules"), ("D", "Liste_Types"),
                         ("E", 'INDIRECT("Cat_"&IFERROR(MATCH($C{row},Liste_Modules,0),1))')]:
        f1 = formula.format(row=PLAN_FIRST) if "{row}" in formula else formula
        dv = DataValidation(type="list", formula1=f1, allow_blank=True,
                            showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}{PLAN_FIRST}:{col}{PLAN_LAST}")

    ws.auto_filter.ref = f"B{PLAN_HEADER}:V{PLAN_LAST}"
    note_box(ws, "B", "V", 5,
             "Une ligne = une catégorie budgétée pour une année. Dupliquez les lignes en "
             "changeant l'année pour préparer l'exercice suivant : l'historique reste intact. "
             "« Écart » est toujours favorable quand il est positif.", height=28)
    return ws


REC_HEADER = 11
REC_FIRST = 12
REC_LAST = 61


def build_recurrents(wb):
    ws = wb.create_sheet("Récurrents")
    setup_sheet(ws, tab_color=TEAL, freeze=f"C{REC_FIRST}")
    nav_bar(ws, "O", row=3, current=None)
    banner(ws, "O", "CHARGES & REVENUS RÉCURRENTS",
           "Abonnements, loyers, mensualités : ce qui tombe chaque mois, avec les échéances à venir")
    widths(ws, {"A": 2, "B": 30, "C": 13, "D": 26, "E": 14, "F": 8, "G": 15, "H": 22,
                "I": 14, "J": 9, "K": 16, "L": 16, "M": 14, "N": 24, "O": 24})

    kpi_card(ws, 7, 2, "CHARGE MENSUELLE RÉCURRENTE",
             f'=SUMIFS($K${REC_FIRST}:$K${REC_LAST},$J${REC_FIRST}:$J${REC_LAST},"Oui")',
             FMT_MONEY, RED, span=3)
    kpi_card(ws, 7, 5, "COÛT ANNUALISÉ",
             f'=SUMIFS($L${REC_FIRST}:$L${REC_LAST},$J${REC_FIRST}:$J${REC_LAST},"Oui")',
             FMT_MONEY, ORANGE, span=3)
    kpi_card(ws, 7, 8, "ÉCHÉANCES DANS 7 JOURS",
             f'=SUMPRODUCT(($J${REC_FIRST}:$J${REC_LAST}="Oui")*'
             f'($G${REC_FIRST}:$G${REC_LAST}<>"")*'
             f'($G${REC_FIRST}:$G${REC_LAST}>=TODAY())*'
             f'($G${REC_FIRST}:$G${REC_LAST}<=TODAY()+7))',
             FMT_INT, BLUE, span=3)
    kpi_card(ws, 7, 11, "PART DES REVENUS MENSUELS",
             f'=IFERROR(SUMIFS($K${REC_FIRST}:$K${REC_LAST},$J${REC_FIRST}:$J${REC_LAST},"Oui")'
             f'/(SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Type,"Revenu",{REAL})/12),0)',
             FMT_PCT, PURPLE, span=3)
    currency_format(ws, "B8:G8")

    headers = ["Libellé", "Module", "Catégorie", "Montant", "Devise", "Fréquence",
               "Compte", "Moyen", "Actif", "Coût mensuel", "Coût annuel",
               "Prochaine échéance", "Alerte", "Notes"]
    table_header(ws, REC_HEADER, "B", "O", headers, height=32)

    for r in range(REC_FIRST, REC_LAST + 1):
        ws.cell(row=r, column=5).number_format = FMT_MONEY
        ws.cell(row=r, column=7).number_format = FMT_DATE
        ws.cell(row=r, column=10, value="Oui" if r == REC_FIRST else None)
        ws.cell(
            row=r, column=11,
            value=(f'=IF($E{r}="","",ROUND($E{r}*IFERROR(INDEX(Facteurs_Frequence,'
                   f'MATCH($G{r},Liste_Frequences,0)),1)*IFERROR(INDEX(Taux_Devises,'
                   f'MATCH(IF($F{r}="",Devise_Base,$F{r}),Codes_Devises,0)),1),0))')
        ).number_format = FMT_MONEY
        ws.cell(row=r, column=12, value=f'=IF($K{r}="","",$K{r}*12)').number_format = FMT_MONEY
        ws.cell(row=r, column=13,
                value=f'=IF($H{r}="","",$H{r})').number_format = FMT_DATE
        ws.cell(
            row=r, column=14,
            value=(f'=IF($H{r}="","",IF($H{r}<TODAY(),"⚠ En retard",'
                   f'IF($H{r}<=TODAY()+7,"⏳ Sous 7 jours",'
                   f'IF($H{r}<=TODAY()+30,"À venir (30 j)","OK"))))'))
    ws.column_dimensions["M"].hidden = True

    paint(ws, f"B{REC_FIRST}:O{REC_LAST}", border=box_border(LINE), font=f(10))
    input_style(ws, f"B{REC_FIRST}:J{REC_LAST}")
    paint(ws, f"K{REC_FIRST}:N{REC_LAST}", fill=solid(BG))
    paint(ws, f"F{REC_FIRST}:F{REC_LAST}", align=al("center"))
    paint(ws, f"J{REC_FIRST}:J{REC_LAST}", align=al("center"))
    paint(ws, f"N{REC_FIRST}:N{REC_LAST}", align=al("center"), font=f(9, True))
    ws.cell(row=REC_FIRST, column=8).number_format = FMT_DATE
    for r in range(REC_FIRST, REC_LAST + 1):
        ws.cell(row=r, column=8).number_format = FMT_DATE
    zebra(ws, f"B{REC_FIRST}:O{REC_LAST}", REC_FIRST)
    currency_format(ws, f"E{REC_FIRST}:E{REC_LAST}")
    currency_format(ws, f"K{REC_FIRST}:L{REC_LAST}")

    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    for text, color, bg in [("⚠ En retard", RED, "FBE8E6"),
                            ("⏳ Sous 7 jours", ORANGE, "FCEFD9"),
                            ("OK", GREEN, "E3F4EB")]:
        dxf = DifferentialStyle(font=Font(color=color, bold=True),
                                fill=PatternFill(bgColor=bg))
        rule = Rule(type="expression", dxf=dxf, stopIfTrue=False)
        rule.formula = [f'$N{REC_FIRST}="{text}"']
        ws.conditional_formatting.add(f"B{REC_FIRST}:O{REC_LAST}", rule)

    for col, formula in [("C", "Liste_Modules"),
                         ("D", 'INDIRECT("Cat_"&IFERROR(MATCH($C{row},Liste_Modules,0),1))'),
                         ("F", "Codes_Devises"), ("G", "Liste_Frequences"),
                         ("H", "Liste_Comptes"), ("I", "Liste_Moyens"),
                         ("J", "Liste_OuiNon")]:
        f1 = formula.format(row=REC_FIRST) if "{row}" in formula else formula
        dv = DataValidation(type="list", formula1=f1, allow_blank=True,
                            showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(f"{col}{REC_FIRST}:{col}{REC_LAST}")

    note_box(ws, "B", "O", 5,
             "Renseignez la date de la prochaine échéance : la colonne « Alerte » vous prévient "
             "à J-30, J-7 et en cas de retard. Le coût mensuel est ramené au mois quelle que "
             "soit la fréquence, et converti dans votre devise de base.", height=28)
    return ws

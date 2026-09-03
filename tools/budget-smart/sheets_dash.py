"""Moteur de calcul et tableaux de bord d'APEX Budget."""
from __future__ import annotations

from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import apex_data as D
from apex_style import (
    ChartZone, nav_bar, percent_labels, tune_chart,
    BG, BLUE, BLUE_SOFT, CARD, FMT_DATE, FMT_INT, FMT_MONEY, FMT_PCT, GOLD,
    GOLD_SOFT, GREEN, GREEN_SOFT, INK, LINE, MUTED, NAVY, NAVY_MID, NAVY_SOFT,
    ORANGE, ORANGE_SOFT, PURPLE, RED, RED_SOFT, SERIES, TEAL, WHITE, al, banner,
    box_border, cell, color_scale_ok, currency_format, databar, f, input_style,
    kpi_card, link, note_box, paint, section, setup_sheet, solid, table_header,
    widths, zebra,
)
from sheets_core import PARAM_ACC_FIRST, PARAM_TOTAL_ROW, REAL

OUT = 'J_Type,"<>Revenu",J_Type,"<>Transfert"'
# Filtre de périmètre : « * » quand tous les modules sont retenus
PER = "J_Module,Perimetre_Filtre"

# Bloc dîme sur la feuille Calculs
C_DIME_HEAD = 307
C_DIME_FIRST = 308

# Géométrie de la feuille Calculs
C_MONTH_FIRST = 6
C_MOD_FIRST = 22
C_CAT_HEAD = 36
C_CAT_FIRST = 37
C_TOPM_FIRST = 183
C_TOPY_FIRST = 196
C_ACC_FIRST = 209
C_CMP_FIRST = 232
C_PRJ_FIRST = 246
C_PRJ_LAST = 285
C_RMOD_FIRST = 289
C_RNAT_FIRST = 295
C_PVR_FIRST = 301


def build_calculs(wb):
    ws = wb.create_sheet("Calculs")
    setup_sheet(ws, tab_color=MUTED, freeze="B6", zoom=85)
    nav_bar(ws, "J", row=3)
    banner(ws, "J", "MOTEUR DE CALCUL",
           "Feuille technique alimentée automatiquement — ne rien saisir ici")
    widths(ws, {"A": 2, "B": 16, "C": 34, "D": 16, "E": 16, "F": 16, "G": 16,
                "H": 16, "I": 16, "J": 16})

    # ---- 1. Agrégats mensuels ----------------------------------------
    section(ws, 4, "B", "J", "1 ▸ AGRÉGATS MENSUELS DE L'ANNÉE DE TRAVAIL")
    table_header(ws, 5, "B", "J",
                 ["Mois n°", "Mois", "Revenus", "Dépenses", "Épargne",
                  "Investissement", "Remb. crédit", "Total sorties", "Solde"])
    ws["K5"], ws["L5"], ws["M5"], ws["N5"], ws["O5"] = (
        "Trésorerie cumulée", "Taux d'épargne", "Dépenses prévues",
        "Écart budget", "Revenus prévus")
    for col in "KLMNO":
        c = ws[f"{col}5"]
        c.fill = solid(NAVY_SOFT)
        c.font = f(9, True, WHITE)
        c.alignment = al("center", "center", wrap=True)
        widths(ws, {col: 16})

    def s(typ, extra=""):
        return (f'SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,$B{{r}},'
                f'J_Type,"{typ}",{PER},{REAL}{extra})')

    for i in range(12):
        r = C_MONTH_FIRST + i
        ws.cell(row=r, column=2, value=i + 1).alignment = al("center")
        ws.cell(row=r, column=3, value=f"=INDEX(Liste_Mois,$B{r})")
        ws.cell(row=r, column=4, value="=" + s("Revenu").format(r=r))
        ws.cell(row=r, column=5, value="=" + s("Dépense").format(r=r))
        ws.cell(row=r, column=6, value="=" + s("Épargne").format(r=r))
        ws.cell(row=r, column=7, value="=" + s("Investissement").format(r=r))
        ws.cell(row=r, column=8, value="=" + s("Remboursement crédit").format(r=r))
        ws.cell(row=r, column=9, value=f"=SUM($E{r}:$H{r})")
        ws.cell(row=r, column=10, value=f"=$D{r}-$I{r}")
        prev = "Solde_Initial" if i == 0 else f"$K{r - 1}"
        ws.cell(row=r, column=11, value=f"={prev}+$J{r}")
        ws.cell(row=r, column=12, value=f'=IFERROR(($F{r}+$G{r})/$D{r},0)')
        mcol = get_column_letter(6 + i)
        ws.cell(row=r, column=13,
                value=(f'=SUMIFS(P_M{i + 1},P_Annee,Annee_Travail,P_Type,"Dépense",'
                       f'P_Module,Perimetre_Filtre)'))
        ws.cell(row=r, column=14, value=f"=$M{r}-$E{r}")
        ws.cell(row=r, column=15,
                value=(f'=SUMIFS(P_M{i + 1},P_Annee,Annee_Travail,P_Type,"Revenu",'
                       f'P_Module,Perimetre_Filtre)'))
    r_tot = C_MONTH_FIRST + 12
    ws.cell(row=r_tot, column=3, value="TOTAL ANNÉE").font = f(10, True, WHITE)
    for col in range(4, 16):
        L = get_column_letter(col)
        if col == 11:
            ws.cell(row=r_tot, column=col, value=f"=$K{r_tot - 1}")
        elif col == 12:
            ws.cell(row=r_tot, column=col,
                    value=f"=IFERROR(($F{r_tot}+$G{r_tot})/$D{r_tot},0)")
        else:
            ws.cell(row=r_tot, column=col,
                    value=f"=SUM({L}{C_MONTH_FIRST}:{L}{C_MONTH_FIRST + 11})")
    paint(ws, f"B{r_tot}:O{r_tot}", fill=solid(NAVY), font=f(10, True, WHITE))
    paint(ws, f"B{C_MONTH_FIRST}:O{r_tot}", border=box_border(LINE))
    paint(ws, f"D{C_MONTH_FIRST}:K{r_tot}", numfmt=FMT_MONEY)
    paint(ws, f"M{C_MONTH_FIRST}:O{r_tot}", numfmt=FMT_MONEY)
    paint(ws, f"L{C_MONTH_FIRST}:L{r_tot}", numfmt=FMT_PCT)

    # ---- 2. Sorties par module ---------------------------------------
    section(ws, 20, "B", "J", "2 ▸ SORTIES PAR MODULE ET PAR MOIS")
    table_header(ws, 21, "B", "F", ["Mois"] + D.MODULES)
    for i in range(12):
        r = C_MOD_FIRST + i
        ws.cell(row=r, column=2, value=f"=INDEX(Liste_Mois,{i + 1})")
        for j, mod in enumerate(D.MODULES):
            ws.cell(row=r, column=3 + j,
                    value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{i + 1},'
                           f'J_Module,"{mod}",{OUT},{REAL})')).number_format = FMT_MONEY
    paint(ws, f"B{C_MOD_FIRST}:F{C_MOD_FIRST + 11}", border=box_border(LINE))

    # ---- 3. Analyse par catégorie ------------------------------------
    section(ws, 35, "B", "I", "3 ▸ ANALYSE PAR CATÉGORIE")
    table_header(ws, C_CAT_HEAD, "B", "I",
                 ["Module", "Catégorie", "Réel du mois", "Réel de l'année",
                  "Prévu de l'année", "Écart", "Clé mois", "Clé année"])
    r = C_CAT_FIRST
    for mod in D.MODULES:
        for cat in D.CATEGORIES[mod]:
            ws.cell(row=r, column=2, value=mod).font = f(9)
            ws.cell(row=r, column=3, value=cat).font = f(9)
            ws.cell(row=r, column=4,
                    value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,Mois_Num,'
                           f'J_Module,$B{r},J_Cat,$C{r},{OUT},{REAL})')
                    ).number_format = FMT_MONEY
            ws.cell(row=r, column=5,
                    value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,'
                           f'J_Module,$B{r},J_Cat,$C{r},{OUT},{REAL})')
                    ).number_format = FMT_MONEY
            ws.cell(row=r, column=6,
                    value=(f'=SUMIFS(P_Total,P_Annee,Annee_Travail,P_Module,$B{r},'
                           f'P_Cat,$C{r})')).number_format = FMT_MONEY
            ws.cell(row=r, column=7, value=f"=$F{r}-$E{r}").number_format = FMT_MONEY
            keep = (f'IF(OR(Perimetre_Filtre="*",$B{r}=Perimetre_Filtre),1,0)')
            ws.cell(row=r, column=8, value=f"=$D{r}*{keep}+ROW()/1000000")
            ws.cell(row=r, column=9, value=f"=$E{r}*{keep}+ROW()/1000000")
            r += 1
    cat_last = r - 1
    paint(ws, f"B{C_CAT_FIRST}:I{cat_last}", border=box_border(LINE))
    paint(ws, f"H{C_CAT_FIRST}:I{cat_last}", font=f(8, False, MUTED))

    # ---- 4. Tops -----------------------------------------------------
    def top_block(title, head_row, first_row, key_col, amount_col):
        section(ws, head_row - 1, "B", "F", title)
        table_header(ws, head_row, "B", "F",
                     ["Rang", "Catégorie", "Montant", "Part", "Clé"])
        for k in range(10):
            rr = first_row + k
            ws.cell(row=rr, column=2, value=k + 1).alignment = al("center")
            ws.cell(row=rr, column=6,
                    value=f"=IFERROR(LARGE(${key_col}${C_CAT_FIRST}:${key_col}${cat_last},{k + 1}),0)")
            ws.cell(row=rr, column=4,
                    value=(f'=IFERROR(INDEX(${amount_col}${C_CAT_FIRST}:'
                           f'${amount_col}${cat_last},MATCH($F{rr},'
                           f'${key_col}${C_CAT_FIRST}:${key_col}${cat_last},0)),0)')
                    ).number_format = FMT_MONEY
            ws.cell(row=rr, column=3,
                    value=(f'=IF($D{rr}<=0,"",IFERROR(INDEX($C${C_CAT_FIRST}:$C${cat_last},'
                           f'MATCH($F{rr},${key_col}${C_CAT_FIRST}:${key_col}${cat_last},0)),""))'))
            ws.cell(row=rr, column=5,
                    value=(f'=IFERROR($D{rr}/SUM(${amount_col}${C_CAT_FIRST}:'
                           f'${amount_col}${cat_last}),0)')).number_format = FMT_PCT
        paint(ws, f"B{first_row}:F{first_row + 9}", border=box_border(LINE))
        paint(ws, f"F{first_row}:F{first_row + 9}", font=f(8, False, MUTED))

    top_block("4 ▸ TOP 10 DES SORTIES DU MOIS", 182, C_TOPM_FIRST, "H", "D")
    top_block("5 ▸ TOP 10 DES SORTIES DE L'ANNÉE", 195, C_TOPY_FIRST, "I", "E")

    # ---- 6. Soldes par compte ----------------------------------------
    section(ws, 207, "B", "D", "6 ▸ SOLDES PAR COMPTE")
    table_header(ws, 208, "B", "D", ["Compte", "Solde actuel", "Part"])
    for i in range(20):
        r = C_ACC_FIRST + i
        pr = PARAM_ACC_FIRST + i
        ws.cell(row=r, column=2, value=f"=IF(Paramètres!$C${pr}=\"\",\"\",Paramètres!$C${pr})")
        ws.cell(row=r, column=3,
                value=f"=IF($B{r}=\"\",0,N(Paramètres!$H${pr}))").number_format = FMT_MONEY
        ws.cell(row=r, column=4,
                value=(f'=IFERROR($C{r}/SUM($C${C_ACC_FIRST}:$C${C_ACC_FIRST + 19}),0)')
                ).number_format = FMT_PCT
    paint(ws, f"B{C_ACC_FIRST}:D{C_ACC_FIRST + 19}", border=box_border(LINE))

    # ---- 7. Comparaison N / N-1 --------------------------------------
    section(ws, 230, "B", "F", "7 ▸ COMPARAISON AVEC L'ANNÉE PRÉCÉDENTE")
    table_header(ws, 231, "B", "F",
                 ["Mois", "Dépenses N", "Dépenses N-1", "Revenus N", "Revenus N-1"])
    for i in range(12):
        r = C_CMP_FIRST + i
        ws.cell(row=r, column=2, value=f"=INDEX(Liste_Mois,{i + 1})")
        ws.cell(row=r, column=3,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{i + 1},'
                       f'{OUT},{PER},{REAL})')).number_format = FMT_MONEY
        ws.cell(row=r, column=4,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail-1,J_Mois,{i + 1},'
                       f'{OUT},{PER},{REAL})')).number_format = FMT_MONEY
        ws.cell(row=r, column=5,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{i + 1},'
                       f'J_Type,"Revenu",{PER},{REAL})')).number_format = FMT_MONEY
        ws.cell(row=r, column=6,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail-1,J_Mois,{i + 1},'
                       f'J_Type,"Revenu",{PER},{REAL})')).number_format = FMT_MONEY
    paint(ws, f"B{C_CMP_FIRST}:F{C_CMP_FIRST + 11}", border=box_border(LINE))

    # ---- 8. Liste consolidée biens / projets / objectifs -------------
    section(ws, 245, "B", "C", "8 ▸ LISTE CONSOLIDÉE — BIENS, PROJETS, OBJECTIFS")
    sources = ([("Immobilier", 10 + i) for i in range(10)] +
               [("Business", 10 + i) for i in range(12)] +
               [("Objectifs", 10 + i) for i in range(12)] +
               [("Mariage", None) for _ in range(6)])
    for i, (sheet, srow) in enumerate(sources):
        r = C_PRJ_FIRST + i
        if srow is None:
            ws.cell(row=r, column=2, value=None)
        else:
            ws.cell(row=r, column=2,
                    value=f"=IF('{sheet}'!$C${srow}=\"\",\"\",'{sheet}'!$C${srow})")
        ws.cell(row=r, column=3, value=sheet if srow else "")
    paint(ws, f"B{C_PRJ_FIRST}:C{C_PRJ_LAST}", border=box_border(LINE), font=f(9))
    free_first = C_PRJ_FIRST + 34
    input_style(ws, f"B{free_first}:B{C_PRJ_LAST}")
    ws.cell(row=free_first - 1, column=4,
            value="↓ lignes libres : ajoutez ici un projet ou un bien "
                  "qui n'existe dans aucun module").font = f(9, False, MUTED, italic=True)

    # ---- 9. Répartitions du mois -------------------------------------
    section(ws, 287, "B", "D", "9 ▸ RÉPARTITION DU MOIS SÉLECTIONNÉ")
    table_header(ws, 288, "B", "C", ["Module", "Sorties du mois"])
    for i, mod in enumerate(D.MODULES):
        r = C_RMOD_FIRST + i
        ws.cell(row=r, column=2, value=mod)
        ws.cell(row=r, column=3,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,Mois_Num,'
                       f'J_Module,"{mod}",{OUT},{REAL})')).number_format = FMT_MONEY
    table_header(ws, 294, "B", "C", ["Nature", "Sorties du mois"])
    for i, nat in enumerate(D.NATURES):
        r = C_RNAT_FIRST + i
        ws.cell(row=r, column=2, value=nat)
        ws.cell(row=r, column=3,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,Mois_Num,'
                       f'J_Nature,"{nat}",{OUT},{PER},{REAL})')).number_format = FMT_MONEY

    section(ws, 299, "B", "E", "10 ▸ PRÉVU vs RÉALISÉ PAR MODULE (mois sélectionné)")
    table_header(ws, 300, "B", "E", ["Module", "Prévu", "Réalisé", "Écart"])
    for i, mod in enumerate(D.MODULES):
        r = C_PVR_FIRST + i
        ws.cell(row=r, column=2, value=mod)
        ws.cell(row=r, column=3,
                value=(f'=SUMPRODUCT((P_Annee=Annee_Travail)*(P_Module="{mod}")*'
                       f'(P_Type<>"Revenu")*INDEX(P_Mois,0,Mois_Num))')
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=4, value=f"=$C{C_RMOD_FIRST + i}").number_format = FMT_MONEY
        ws.cell(row=r, column=5, value=f"=$C{r}-$D{r}").number_format = FMT_MONEY
    section(ws, C_DIME_HEAD - 1, "B", "F", "11 ▸ DÎME — DUE ET VERSÉE, MOIS PAR MOIS")
    table_header(ws, C_DIME_HEAD, "B", "F",
                 ["Mois", "Revenus soumis", "Dîme due", "Déjà versée", "Reste à verser"])
    for i in range(12):
        r = C_DIME_FIRST + i
        m = i + 1
        base_all = (f'SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
                    f'J_Type,"Revenu",{REAL})')
        base_sal = (f'SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
                    f'J_Type,"Revenu",J_Cat,"Salaire",{REAL})')
        base_prm = (f'SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
                    f'J_Type,"Revenu",J_Cat,"Prime / 13e mois",{REAL})')
        ws.cell(row=r, column=2, value=f"=INDEX(Liste_Mois,{m})")
        ws.cell(row=r, column=3,
                value=(f'=IF(Dime_Active<>"Oui",0,'
                       f'IF(Dime_Base="Salaire uniquement",{base_sal},'
                       f'IF(Dime_Base="Salaire + primes",{base_sal}+{base_prm},'
                       f'{base_all})))')).number_format = FMT_MONEY
        ws.cell(row=r, column=4,
                value=f'=ROUND($C{r}*N(Dime_Taux),0)').number_format = FMT_MONEY
        ws.cell(row=r, column=5,
                value=(f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,{m},'
                       f'J_Cat,"{D.CAT_DIME}",J_Type,"Dépense",{REAL})')
                ).number_format = FMT_MONEY
        ws.cell(row=r, column=6,
                value=f'=MAX(0,$D{r}-$E{r})').number_format = FMT_MONEY
    paint(ws, f"B{C_DIME_FIRST}:F{C_DIME_FIRST + 11}", border=box_border(LINE))

    paint(ws, f"B{C_RMOD_FIRST}:C{C_RMOD_FIRST + 3}", border=box_border(LINE))
    paint(ws, f"B{C_RNAT_FIRST}:C{C_RNAT_FIRST + 2}", border=box_border(LINE))
    paint(ws, f"B{C_PVR_FIRST}:E{C_PVR_FIRST + 3}", border=box_border(LINE))
    return ws, cat_last


def _color_series(serie, hexcolor):
    serie.graphicalProperties = GraphicalProperties(solidFill=hexcolor)
    serie.graphicalProperties.line.solidFill = hexcolor


def build_dashboard(wb, wsc):
    ws = wb.create_sheet("Tableau de Bord")
    setup_sheet(ws, tab_color=NAVY, freeze="A6", zoom=85)
    nav_bar(ws, "P", row=3, current="Tableau de Bord")
    banner(ws, "P", "TABLEAU DE BORD MENSUEL",
           "=\"Suivi de \"&Nom_Foyer&\"  •  \"&Mois_Suivi&\" \"&Annee_Travail"
           "&\"  •  périmètre : \"&Perimetre&\"  •  montants en \"&Devise_Base")
    widths(ws, {"A": 2, "B": 5, "C": 22, "D": 14, "E": 14, "F": 14, "G": 14, "H": 8,
                "I": 14, "J": 5, "K": 24, "L": 14, "M": 14, "N": 14, "O": 14, "P": 14})

    # --- sélecteur de période ----------------------------------------
    ws.merge_cells("B4:C4")
    c = ws["B4"]
    c.value = "PÉRIODE ANALYSÉE"
    c.font = f(10, True, WHITE)
    c.alignment = al("left", "center", indent=1)
    paint(ws, "B4:C4", fill=solid(NAVY_MID))
    ws["D4"] = 2026
    ws["D4"].number_format = "0"
    ws["G4"] = "Janvier"
    for coord in ("D4", "G4"):
        ws[coord].font = f(12, True, INK)
        ws[coord].alignment = al("center")
    input_style(ws, "D4:D4")
    input_style(ws, "G4:G4")
    ws["E4"] = "◄ année"
    ws["H4"] = "=IFERROR(MATCH($G$4,Liste_Mois,0),1)"
    ws["H4"].font = f(8, False, MUTED)
    ws["H4"].alignment = al("center")
    ws["F4"] = "mois ►"
    for coord in ("E4", "F4"):
        ws[coord].font = f(9, False, MUTED)
        ws[coord].alignment = al("center")
    dv_y = DataValidation(type="whole", operator="between", formula1="2000",
                          formula2="2100", allow_blank=False)
    ws.add_data_validation(dv_y)
    dv_y.add(ws["D4"])
    dv_m = DataValidation(type="list", formula1="Liste_Mois", allow_blank=False)
    ws.add_data_validation(dv_m)
    dv_m.add(ws["G4"])
    ws.merge_cells("I4:P4")
    ws["I4"] = ('="Mise à jour au \"&TEXT(TODAY(),"dd/mm/yyyy")&"   •   "'
                '&COUNT(J_Date)&" écritures enregistrées   •   Devise : "&Devise_Base'
                '&" ("&Symbole_Base&")"')
    ws["I4"].font = f(9, False, MUTED, italic=True)
    ws["I4"].alignment = al("right", "center", indent=1)
    ws.row_dimensions[4].height = 24

    M = f"INDEX(Calculs!$D${C_MONTH_FIRST}:$O${C_MONTH_FIRST + 11},Mois_Num,{{c}})"

    def mval(col_index):
        return "=" + M.format(c=col_index)

    # --- KPI ligne 1 --------------------------------------------------
    kpi_card(ws, 6, 2, "REVENUS DU MOIS", mval(1), FMT_MONEY, GREEN, span=3,
             note_formula='="Prévu : "&TEXT(' + M.format(c=12) + ',"#,##0")')
    kpi_card(ws, 6, 5, "DÉPENSES DU MOIS", mval(2), FMT_MONEY, RED, span=3,
             note_formula='="Budget : "&TEXT(' + M.format(c=10) + ',"#,##0")')
    kpi_card(ws, 6, 8, "ÉPARGNE + INVESTI.",
             "=" + M.format(c=3) + "+" + M.format(c=4), FMT_MONEY, TEAL, span=3,
             note_formula='="Objectif : "&TEXT(Taux_Cible*' + M.format(c=1) + ',"#,##0")')
    kpi_card(ws, 6, 11, "SOLDE DU MOIS", mval(7), FMT_MONEY, BLUE, span=3,
             note_formula='=IF(' + M.format(c=7) + '>=0,"Excédent","Déficit à couvrir")')
    kpi_card(ws, 6, 14, "TAUX D'ÉPARGNE", mval(9), FMT_PCT, PURPLE, span=3,
             note_formula='="Cible : "&TEXT(Taux_Cible,"0%")')

    # --- KPI ligne 2 --------------------------------------------------
    kpi_card(ws, 10, 2, "BUDGET CONSOMMÉ",
             "=IFERROR(" + M.format(c=2) + "/" + M.format(c=10) + ",0)",
             FMT_PCT, ORANGE, span=3,
             note_formula='="Écart : "&TEXT(' + M.format(c=11) + ',"#,##0")')
    kpi_card(ws, 10, 5, "TRÉSORERIE CUMULÉE", mval(8), FMT_MONEY, NAVY, span=3,
             note_formula=f'="Dont comptes : "&TEXT(Paramètres!$H${PARAM_TOTAL_ROW},"#,##0")')
    kpi_card(ws, 10, 8, "CHARGES FIXES DU MOIS",
             f'=SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Mois,Mois_Num,'
             f'J_Nature,"Fixe",{OUT},{PER},{REAL})', FMT_MONEY, NAVY_SOFT, span=3,
             note="Loyer, abonnements, mensualités")
    kpi_card(ws, 10, 11, "RESTE À VIVRE",
             "=" + M.format(c=1) + f'-SUMIFS(J_Montant,J_Annee,Annee_Travail,'
             f'J_Mois,Mois_Num,J_Nature,"Fixe",{OUT},{PER},{REAL})',
             FMT_MONEY, GOLD, span=3, note="Revenus – charges fixes")
    kpi_card(ws, 10, 14, "DÉPENSE MOYENNE / JOUR",
             "=IFERROR(" + M.format(c=2) + "/DAY(EOMONTH(DATE(Annee_Travail,Mois_Num,1),0)),0)",
             FMT_MONEY, TEAL, span=3, note="Sur le mois sélectionné")

    # le format monétaire dynamique ne s'applique qu'aux cartes en valeur
    for rng in ("B7:D7", "E7:G7", "H7:J7", "K7:M7",
                "E11:G11", "H11:J11", "K11:M11", "N11:P11"):
        currency_format(ws, rng)
    ws["N7"].number_format = FMT_PCT
    ws["B11"].number_format = FMT_PCT

    # --- Graphiques ---------------------------------------------------
    section(ws, 78, "B", "P", "GRAPHIQUES")

    ch1 = BarChart()
    ch1.type = "col"
    ch1.grouping = "clustered"
    ch1.title = "Revenus, dépenses et épargne par mois"
    data = Reference(wsc, min_col=4, max_col=6, min_row=5, max_row=C_MONTH_FIRST + 11)
    cats = Reference(wsc, min_col=3, min_row=C_MONTH_FIRST, max_row=C_MONTH_FIRST + 11)
    ch1.add_data(data, titles_from_data=True)
    ch1.set_categories(cats)
    for i, serie in enumerate(ch1.series):
        _color_series(serie, [GREEN, RED, TEAL][i])
    ch1.height, ch1.width = 9.0, 17.0
    ch1.y_axis.majorGridlines = None
    ch1.gapWidth = 60
    tune_chart(ch1, axis_size=750, legend="t")
    zg = ChartZone(ws, "B", 80)
    zd = ChartZone(ws, "J", 80)
    zg.add(ch1)

    ch2 = LineChart()
    ch2.title = "Trésorerie cumulée"
    d2 = Reference(wsc, min_col=11, min_row=5, max_row=C_MONTH_FIRST + 11)
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(cats)
    _color_series(ch2.series[0], BLUE)
    ch2.series[0].smooth = False
    ch2.height, ch2.width = 9.0, 15.0
    tune_chart(ch2, axis_size=750, no_legend=True)
    zd.add(ch2)

    section(ws, 14, "B", "P", "RÉPARTITION DU MOIS SÉLECTIONNÉ")

    ch3 = DoughnutChart()
    ch3.title = "Sorties par module"
    d3 = Reference(wsc, min_col=3, min_row=C_RMOD_FIRST, max_row=C_RMOD_FIRST + 3)
    c3 = Reference(wsc, min_col=2, min_row=C_RMOD_FIRST, max_row=C_RMOD_FIRST + 3)
    ch3.add_data(d3, titles_from_data=False)
    ch3.set_categories(c3)
    ch3.dataLabels = percent_labels()
    ch3.holeSize = 55
    ch3.height, ch3.width = 9.0, 13.0
    tune_chart(ch3, axis_size=800)
    zg.add(ch3)

    ch4 = PieChart()
    ch4.title = "Sorties par nature"
    d4 = Reference(wsc, min_col=3, min_row=C_RNAT_FIRST, max_row=C_RNAT_FIRST + 2)
    c4 = Reference(wsc, min_col=2, min_row=C_RNAT_FIRST, max_row=C_RNAT_FIRST + 2)
    ch4.add_data(d4, titles_from_data=False)
    ch4.set_categories(c4)
    ch4.dataLabels = percent_labels()
    ch4.height, ch4.width = 9.0, 12.0
    tune_chart(ch4, axis_size=800)
    zd.add(ch4)

    # Top 10 du mois (tableau)
    table_header(ws, 15, "K", "P",
                 ["Top 10 des dépenses du mois", "", "Montant", "", "Part", ""])
    ws.merge_cells("K15:L15")
    ws.merge_cells("M15:N15")
    ws.merge_cells("O15:P15")
    for k in range(10):
        r = 16 + k
        ws.merge_cells(f"K{r}:L{r}")
        ws[f"K{r}"] = f"=IF(Calculs!$C${C_TOPM_FIRST + k}=\"\",\"—\",Calculs!$C${C_TOPM_FIRST + k})"
        ws[f"K{r}"].font = f(10)
        ws[f"K{r}"].alignment = al("left", "center", indent=1)
        ws.merge_cells(f"M{r}:N{r}")
        ws[f"M{r}"] = f"=Calculs!$D${C_TOPM_FIRST + k}"
        ws[f"M{r}"].number_format = FMT_MONEY
        ws[f"M{r}"].font = f(10, True, NAVY)
        ws.merge_cells(f"O{r}:P{r}")
        ws[f"O{r}"] = f"=Calculs!$E${C_TOPM_FIRST + k}"
        ws[f"O{r}"].number_format = FMT_PCT
    paint(ws, "K16:P25", border=box_border(LINE))
    zebra(ws, "K16:P25", 16)
    databar(ws, "O16:P25", ORANGE)
    currency_format(ws, "M16:N25")

    # --- Prévu vs réalisé ---------------------------------------------
    section(ws, 28, "B", "P", "PRÉVU vs RÉALISÉ — MOIS SÉLECTIONNÉ")
    table_header(ws, 29, "B", "H",
                 ["", "Module", "Prévu", "Réalisé", "Écart", "% consommé", "Statut"])
    for i, mod in enumerate(D.MODULES):
        r = 30 + i
        src = C_PVR_FIRST + i
        ws[f"C{r}"] = mod
        ws[f"D{r}"] = f"=Calculs!$C${src}"
        ws[f"E{r}"] = f"=Calculs!$D${src}"
        ws[f"F{r}"] = f"=Calculs!$E${src}"
        ws[f"G{r}"] = f'=IFERROR($E{r}/$D{r},0)'
        ws[f"H{r}"] = (f'=IF($D{r}=0,"Non budgété",IF($G{r}>1,"⚠ Dépassement",'
                       f'IF($G{r}>Seuil_Alerte,"Vigilance","✓ Sous contrôle")))')
        for col in "DEF":
            ws[f"{col}{r}"].number_format = FMT_MONEY
        ws[f"G{r}"].number_format = FMT_PCT
        ws[f"H{r}"].font = f(9, True)
        ws[f"H{r}"].alignment = al("center")
    r = 34
    ws[f"C{r}"] = "TOTAL"
    ws[f"D{r}"] = "=SUM(D30:D33)"
    ws[f"E{r}"] = "=SUM(E30:E33)"
    ws[f"F{r}"] = "=SUM(F30:F33)"
    ws[f"G{r}"] = "=IFERROR(E34/D34,0)"
    for col in "DEF":
        ws[f"{col}{r}"].number_format = FMT_MONEY
    ws[f"G{r}"].number_format = FMT_PCT
    paint(ws, "B30:H34", border=box_border(LINE))
    paint(ws, "B34:H34", fill=solid(NAVY), font=f(10, True, WHITE))
    color_scale_ok(ws, "F30:F33")
    currency_format(ws, "D30:F34")

    ch5 = BarChart()
    ch5.type = "bar"
    ch5.title = "Prévu vs réalisé par module"
    d5 = Reference(wsc, min_col=3, max_col=4, min_row=300, max_row=C_PVR_FIRST + 3)
    c5 = Reference(wsc, min_col=2, min_row=C_PVR_FIRST, max_row=C_PVR_FIRST + 3)
    ch5.add_data(d5, titles_from_data=True)
    ch5.set_categories(c5)
    _color_series(ch5.series[0], NAVY_SOFT)
    _color_series(ch5.series[1], GOLD)
    ch5.height, ch5.width = 8.0, 14.0
    ch5.gapWidth = 70
    tune_chart(ch5, axis_size=750, legend="t")
    tune_chart(ch5, axis_size=750, legend="t")
    zg.add(ch5)

    # --- Dîme ---------------------------------------------------------
    section(ws, 38, "B", "H", "DÎME DU MOIS")
    dime = [
        ("Revenus soumis à la dîme", f"=INDEX(Calculs!$C${C_DIME_FIRST}:"
                                     f"$C${C_DIME_FIRST + 11},Mois_Num)", FMT_MONEY),
        ("Dîme due", f"=INDEX(Calculs!$D${C_DIME_FIRST}:$D${C_DIME_FIRST + 11},Mois_Num)",
         FMT_MONEY),
        ("Déjà versée", f"=INDEX(Calculs!$E${C_DIME_FIRST}:$E${C_DIME_FIRST + 11},Mois_Num)",
         FMT_MONEY),
        ("Reste à verser", f"=INDEX(Calculs!$F${C_DIME_FIRST}:$F${C_DIME_FIRST + 11},"
                           f"Mois_Num)", FMT_MONEY),
    ]
    for i, (lab, formula, fmt) in enumerate(dime):
        r = 39 + i
        ws.merge_cells(f"C{r}:E{r}")
        ws[f"C{r}"] = lab
        ws[f"C{r}"].font = f(10, i == 3, INK)
        ws[f"C{r}"].alignment = al("left", "center", indent=1)
        ws.merge_cells(f"F{r}:G{r}")
        v = ws[f"F{r}"]
        v.value = formula
        v.number_format = fmt
        v.font = f(11, True, GOLD if i == 1 else (RED if i == 3 else NAVY))
        v.alignment = al("right", "center", indent=1)
    ws.merge_cells("C43:G43")
    st = ws["C43"]
    st.value = ('=IF(Dime_Active<>"Oui","Dîme désactivée dans Paramètres.",'
                'IF($F$40=0,"Aucun revenu soumis ce mois-ci.",'
                'IF($F$42<=0,"✓ Dîme du mois entièrement versée.",'
                '"Il reste "&FIXED($F$42,0)&" "&Symbole_Base&" à verser.")))')
    st.font = f(10, True, NAVY)
    st.alignment = al("left", "center", indent=1)
    paint(ws, "C39:G43", border=box_border(LINE))
    paint(ws, "C39:G43", fill=solid(GOLD_SOFT))
    currency_format(ws, "F39:G42")

    # --- Alertes ------------------------------------------------------
    section(ws, 46, "B", "P", "ALERTES & SIGNAUX AUTOMATIQUES")
    alerts = [
        ('=IF(' + M.format(c=7) + '<0,"🔴 Le mois est déficitaire de "&'
         'TEXT(ABS(' + M.format(c=7) + '),"#,##0")&" "&Symbole_Base&". Réduisez les postes '
         'variables ou puisez dans vos réserves.","🟢 Le mois est excédentaire.")'),
        ('=IF(' + M.format(c=9) + '<Taux_Cible,"🟠 Taux d\'épargne de "&'
         'TEXT(' + M.format(c=9) + ',"0.0%")&" contre "&TEXT(Taux_Cible,"0%")&" visé.",'
         '"🟢 Objectif d\'épargne atteint.")'),
        ('=IF(AND(' + M.format(c=10) + '>0,' + M.format(c=2) + '>' + M.format(c=10) +
         '),"🔴 Budget de dépenses dépassé de "&TEXT(' + M.format(c=2) + '-' +
         M.format(c=10) + ',"#,##0")&" "&Symbole_Base&".",'
         'IF(' + M.format(c=10) + '=0,"⚪ Aucun budget saisi pour ce mois dans le Estimation annuelle.",'
         '"🟢 Dépenses contenues dans le budget."))'),
        ('=IF(COUNTIFS(Récurrents!$J$8:$J$57,"Oui",Récurrents!$H$8:$H$57,"<"&TODAY())>0,'
         '"🔴 "&COUNTIFS(Récurrents!$J$8:$J$57,"Oui",Récurrents!$H$8:$H$57,"<"&TODAY())&'
         '" échéance(s) récurrente(s) en retard.","🟢 Aucune échéance récurrente en retard.")'),
        ('=IF(SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Statut,"Prévu")>0,'
         '"🟠 "&TEXT(SUMIFS(J_Montant,J_Annee,Annee_Travail,J_Statut,"Prévu"),"#,##0")&" "&'
         'Symbole_Base&" d\'engagements prévus non encore décaissés.",'
         '"🟢 Aucun engagement en attente.")'),
        ('=IF(Paramètres!$H$' + str(PARAM_TOTAL_ROW) + '<0,"🔴 Trésorerie de vos comptes négative.",'
         'IF(Paramètres!$H$' + str(PARAM_TOTAL_ROW) + '<' + M.format(c=2) + ',"🟠 Votre trésorerie couvre moins d\'un '
         'mois de dépenses.","🟢 Trésorerie confortable."))'),
        ('=IF(Dime_Active<>"Oui","⚪ Dîme désactivée.",'
         'IF(INDEX(Calculs!$F$' + str(C_DIME_FIRST) + ':$F$' + str(C_DIME_FIRST + 11) +
         ',Mois_Num)>0,"🟠 Dîme : il reste "&FIXED(INDEX(Calculs!$F$' + str(C_DIME_FIRST) +
         ':$F$' + str(C_DIME_FIRST + 11) + ',Mois_Num),0)&" "&Symbole_Base&" à verser ce mois-ci.",'
         '"🟢 Dîme du mois à jour."))'),
        ('=IF(COUNT(J_Date)=0,"⚪ Commencez par saisir vos premières opérations dans le Journal.",'
         '"🟢 "&COUNT(J_Date)&" opérations enregistrées depuis le début.")'),
    ]
    for i, formula in enumerate(alerts):
        r = 47 + i
        ws.merge_cells(f"B{r}:P{r}")
        c = ws[f"B{r}"]
        c.value = formula
        c.font = f(10)
        c.alignment = al("left", "center", indent=1)
        ws.row_dimensions[r].height = 20
    paint(ws, "B47:P54", fill=solid(CARD), border=box_border(LINE))

    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.styles.differential import DifferentialStyle
    for token, color, bg in [("🔴", RED, RED_SOFT), ("🟠", ORANGE, ORANGE_SOFT),
                             ("🟢", GREEN, GREEN_SOFT)]:
        dxf = DifferentialStyle(font=Font(color=color, bold=True),
                                fill=PatternFill(bgColor=bg))
        rule = Rule(type="expression", dxf=dxf, stopIfTrue=False)
        rule.formula = [f'ISNUMBER(SEARCH("{token}",$B47))']
        ws.conditional_formatting.add("B47:P54", rule)

    note_box(ws, "B", "P", 56,
             "Changez l'année ou le mois en haut de cette feuille : tous les indicateurs, "
             "graphiques et alertes se recalculent instantanément à partir du Journal.",
             height=26)
    return ws


def build_annuel(wb, wsc):
    ws = wb.create_sheet("Synthèse Annuelle")
    setup_sheet(ws, tab_color=GOLD, freeze="A6", zoom=85)
    nav_bar(ws, "P", row=3)
    banner(ws, "P", "SYNTHÈSE ANNUELLE",
           '="Année "&Annee_Travail&"  •  "&Nom_Foyer&"  •  montants en "&Devise_Base')
    widths(ws, {"A": 2, "B": 5, "C": 16, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15,
                "I": 15, "J": 15, "K": 15, "L": 15, "M": 15, "N": 15, "O": 15, "P": 15})

    T = f"Calculs!${{c}}${C_MONTH_FIRST + 12}"
    kpi_card(ws, 5, 2, "REVENUS DE L'ANNÉE", "=" + T.format(c="D"), FMT_MONEY, GREEN, span=3)
    kpi_card(ws, 5, 5, "DÉPENSES DE L'ANNÉE", "=" + T.format(c="E"), FMT_MONEY, RED, span=3)
    kpi_card(ws, 5, 8, "ÉPARGNE + INVESTISSEMENT",
             "=" + T.format(c="F") + "+" + T.format(c="G"), FMT_MONEY, TEAL, span=3)
    kpi_card(ws, 5, 11, "SOLDE ANNUEL", "=" + T.format(c="J"), FMT_MONEY, BLUE, span=3)
    kpi_card(ws, 5, 14, "TAUX D'ÉPARGNE MOYEN", "=" + T.format(c="L"), FMT_PCT, PURPLE, span=3)

    kpi_card(ws, 9, 2, "TRÉSORERIE FIN D'ANNÉE", "=" + T.format(c="K"), FMT_MONEY, NAVY, span=3)
    kpi_card(ws, 9, 5, "DÉPENSE MOYENNE / MOIS",
             "=IFERROR(" + T.format(c="E") + "/12,0)", FMT_MONEY, ORANGE, span=3)
    kpi_card(ws, 9, 8, "MEILLEUR MOIS",
             f'=IFERROR(INDEX(Calculs!$C${C_MONTH_FIRST}:$C${C_MONTH_FIRST + 11},'
             f'MATCH(MAX(Calculs!$J${C_MONTH_FIRST}:$J${C_MONTH_FIRST + 11}),'
             f'Calculs!$J${C_MONTH_FIRST}:$J${C_MONTH_FIRST + 11},0)),"—")',
             "General", GREEN, span=3,
             note_formula=f'="Solde : "&TEXT(MAX(Calculs!$J${C_MONTH_FIRST}:'
                          f'$J${C_MONTH_FIRST + 11}),"#,##0")')
    kpi_card(ws, 9, 11, "MOIS LE PLUS TENDU",
             f'=IFERROR(INDEX(Calculs!$C${C_MONTH_FIRST}:$C${C_MONTH_FIRST + 11},'
             f'MATCH(MIN(Calculs!$J${C_MONTH_FIRST}:$J${C_MONTH_FIRST + 11}),'
             f'Calculs!$J${C_MONTH_FIRST}:$J${C_MONTH_FIRST + 11},0)),"—")',
             "General", RED, span=3,
             note_formula=f'="Solde : "&TEXT(MIN(Calculs!$J${C_MONTH_FIRST}:'
                          f'$J${C_MONTH_FIRST + 11}),"#,##0")')
    kpi_card(ws, 9, 14, "ÉCART AU BUDGET",
             "=" + T.format(c="N"), FMT_MONEY, GOLD, span=3,
             note="Positif = dépenses inférieures au plan")
    for rng in ("B6:D6", "E6:G6", "H6:J6", "K6:M6",
                "B10:D10", "E10:G10", "N10:P10"):
        currency_format(ws, rng)
    ws["N6"].number_format = FMT_PCT
    ws["H10"].number_format = "General"
    ws["K10"].number_format = "General"

    section(ws, 13, "B", "P", "TABLEAU ANNUEL DÉTAILLÉ")
    table_header(ws, 14, "B", "P",
                 ["", "Mois", "Revenus", "Dépenses", "Épargne", "Investissement",
                  "Remb. crédit", "Total sorties", "Solde", "Trésorerie cumulée",
                  "Taux d'épargne", "Dépenses prévues", "Écart budget",
                  "Revenus prévus", "Écart revenus"])
    for i in range(12):
        r = 15 + i
        src = C_MONTH_FIRST + i
        ws[f"C{r}"] = f"=Calculs!$C${src}"
        for j, col in enumerate("DEFGHIJKLMNO"):
            ws[f"{col}{r}"] = f"=Calculs!${chr(ord('D') + j)}${src}"
            ws[f"{col}{r}"].number_format = FMT_MONEY
        ws[f"L{r}"].number_format = FMT_PCT
        ws[f"P{r}"] = f"=Calculs!$D${src}-Calculs!$O${src}"
        ws[f"P{r}"].number_format = FMT_MONEY
    r = 27
    ws[f"C{r}"] = "TOTAL"
    for col in "DEFGHIJMNOP":
        ws[f"{col}{r}"] = f"=SUM({col}15:{col}26)"
        ws[f"{col}{r}"].number_format = FMT_MONEY
    ws[f"K{r}"] = "=K26"
    ws[f"K{r}"].number_format = FMT_MONEY
    ws[f"L{r}"] = f"=Calculs!$L${C_MONTH_FIRST + 12}"
    ws[f"L{r}"].number_format = FMT_PCT
    paint(ws, "B15:P27", border=box_border(LINE))
    paint(ws, f"B{r}:P{r}", fill=solid(NAVY), font=f(10, True, WHITE))
    zebra(ws, "B15:P26", 15)
    color_scale_ok(ws, "J15:J26")
    color_scale_ok(ws, "N15:N26")
    currency_format(ws, "D15:K27")
    currency_format(ws, "M15:P27")

    section(ws, 30, "B", "P", "GRAPHIQUES DE L'ANNÉE")
    cats = Reference(wsc, min_col=3, min_row=C_MONTH_FIRST, max_row=C_MONTH_FIRST + 11)

    ch1 = BarChart()
    ch1.type = "col"
    ch1.grouping = "stacked"
    ch1.overlap = 100
    ch1.title = "Sorties par module"
    d1 = Reference(wsc, min_col=3, max_col=6, min_row=21, max_row=C_MOD_FIRST + 11)
    ch1.add_data(d1, titles_from_data=True)
    ch1.set_categories(Reference(wsc, min_col=2, min_row=C_MOD_FIRST,
                                 max_row=C_MOD_FIRST + 11))
    for i, serie in enumerate(ch1.series):
        _color_series(serie, SERIES[i])
    ch1.height, ch1.width = 10.0, 16.0
    tune_chart(ch1, axis_size=750, legend="t")
    za = ChartZone(ws, "B", 100)
    zb = ChartZone(ws, "J", 100)
    za.add(ch1)

    ch2 = LineChart()
    ch2.title = "Taux d'épargne mensuel"
    d2 = Reference(wsc, min_col=12, min_row=5, max_row=C_MONTH_FIRST + 11)
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(cats)
    _color_series(ch2.series[0], PURPLE)
    ch2.y_axis.numFmt = "0%"
    ch2.height, ch2.width = 9.0, 15.0
    tune_chart(ch2, axis_size=750, no_legend=True)
    zb.add(ch2)

    section(ws, 50, "B", "P", "COMPARAISON AVEC L'ANNÉE PRÉCÉDENTE")
    table_header(ws, 51, "B", "I",
                 ["", "Mois", "Dépenses N", "Dépenses N-1", "Variation",
                  "Revenus N", "Revenus N-1", "Variation"])
    for i in range(12):
        r = 52 + i
        src = C_CMP_FIRST + i
        ws[f"C{r}"] = f"=Calculs!$B${src}"
        ws[f"D{r}"] = f"=Calculs!$C${src}"
        ws[f"E{r}"] = f"=Calculs!$D${src}"
        ws[f"F{r}"] = f"=IFERROR($D{r}/$E{r}-1,0)"
        ws[f"G{r}"] = f"=Calculs!$E${src}"
        ws[f"H{r}"] = f"=Calculs!$F${src}"
        ws[f"I{r}"] = f"=IFERROR($G{r}/$H{r}-1,0)"
        for col in "DEGH":
            ws[f"{col}{r}"].number_format = FMT_MONEY
        for col in "FI":
            ws[f"{col}{r}"].number_format = FMT_PCT
    paint(ws, "B52:I63", border=box_border(LINE))
    zebra(ws, "B52:I63", 52)
    currency_format(ws, "D52:E63")
    currency_format(ws, "G52:H63")

    ch3 = BarChart()
    ch3.type = "col"
    ch3.title = "Dépenses : année en cours vs précédente"
    d3 = Reference(wsc, min_col=3, max_col=4, min_row=231, max_row=C_CMP_FIRST + 11)
    ch3.add_data(d3, titles_from_data=True)
    ch3.set_categories(Reference(wsc, min_col=2, min_row=C_CMP_FIRST,
                                 max_row=C_CMP_FIRST + 11))
    _color_series(ch3.series[0], NAVY_SOFT)
    _color_series(ch3.series[1], "9AA9BD")
    ch3.height, ch3.width = 9.0, 15.0
    ch3.gapWidth = 60
    tune_chart(ch3, axis_size=750, legend="t")
    za.add(ch3)

    section(ws, 68, "B", "P", "TOP 10 DES POSTES DE DÉPENSE DE L'ANNÉE")
    table_header(ws, 69, "B", "F", ["Rang", "Catégorie", "Montant annuel", "Part", ""])
    for k in range(10):
        r = 70 + k
        src = C_TOPY_FIRST + k
        ws[f"B{r}"] = k + 1
        ws[f"B{r}"].alignment = al("center")
        ws[f"C{r}"] = f'=IF(Calculs!$C${src}="","—",Calculs!$C${src})'
        ws[f"D{r}"] = f"=Calculs!$D${src}"
        ws[f"D{r}"].number_format = FMT_MONEY
        ws[f"E{r}"] = f"=Calculs!$E${src}"
        ws[f"E{r}"].number_format = FMT_PCT
        ws[f"F{r}"] = f'=IF($D{r}=0,"",REPT("█",ROUND($E{r}*40,0)))'
        ws[f"F{r}"].font = f(9, False, GOLD)
    paint(ws, "B70:F79", border=box_border(LINE))
    zebra(ws, "B70:F79", 70)
    currency_format(ws, "D70:D79")

    ch4 = DoughnutChart()
    ch4.title = "Répartition annuelle des dépenses (top 10)"
    d4 = Reference(wsc, min_col=4, min_row=C_TOPY_FIRST, max_row=C_TOPY_FIRST + 9)
    c4 = Reference(wsc, min_col=3, min_row=C_TOPY_FIRST, max_row=C_TOPY_FIRST + 9)
    ch4.add_data(d4, titles_from_data=False)
    ch4.set_categories(c4)
    ch4.holeSize = 55
    ch4.height, ch4.width = 11.0, 16.0
    tune_chart(ch4, axis_size=750)
    zb.add(ch4)

    section(ws, 83, "B", "P", "SOLDES PAR COMPTE")
    table_header(ws, 84, "B", "E", ["", "Compte", "Solde actuel", "Part"])
    for i in range(10):
        r = 85 + i
        src = C_ACC_FIRST + i
        ws[f"C{r}"] = f"=Calculs!$B${src}"
        ws[f"D{r}"] = f"=Calculs!$C${src}"
        ws[f"D{r}"].number_format = FMT_MONEY
        ws[f"E{r}"] = f"=Calculs!$D${src}"
        ws[f"E{r}"].number_format = FMT_PCT
    paint(ws, "B85:E94", border=box_border(LINE))
    zebra(ws, "B85:E94", 85)
    databar(ws, "E85:E94", TEAL)
    currency_format(ws, "D85:D94")
    return ws

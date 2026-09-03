"""Palette, formats et briques de mise en page du classeur Budget Smart."""
from __future__ import annotations

import math

from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.numbers import NumberFormat
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# --------------------------------------------------------------------------
# Palette
# --------------------------------------------------------------------------
NAVY = "1A3557"          # bleu nuit APEX AFRICA
NAVY_MID = "234468"
NAVY_SOFT = "2E5480"
GOLD = "B8860B"          # or de la charte, sur fond clair
GOLD_SOFT = "FDF6E3"     # crème
GOLD_LINE = "D9A93C"     # or clair, déclinaison sur fond sombre
INK = "404040"           # gris ardoise, corps de texte
MUTED = "7A8794"
LINE = "D6DEE8"
BG = "F2F2F2"
CARD = "FFFFFF"
WHITE = "FFFFFF"

GREEN = "1E6B3C"
GREEN_SOFT = "D4EDDA"
RED = "8B1A1A"
RED_SOFT = "F8D7DA"
BLUE = "2E5480"
BLUE_SOFT = "D6E4F0"
TEAL = "10706B"
TEAL_SOFT = "DCEDEC"
PURPLE = "4A5C8C"
PURPLE_SOFT = "E2E7F2"
ORANGE = "E07B22"
ORANGE_SOFT = "FBEBD8"

# Couleurs de séries pour les graphiques
SERIES = [BLUE, GOLD, GREEN, ORANGE, TEAL, PURPLE, RED, "9C3A5F",
          "9FB0C4", GOLD_LINE]

# --------------------------------------------------------------------------
# Formats de nombre
# --------------------------------------------------------------------------
FMT_MONEY = '#,##0;[Red]-#,##0'
FMT_MONEY_BLANK = '#,##0;[Red]-#,##0;'   # les zéros restent invisibles
FMT_MONEY_D = '#,##0.00;[Red]-#,##0.00'
FMT_INT = '#,##0'
FMT_PCT = '0.0%'
FMT_PCT0 = '0%'
FMT_DATE = 'dd/mm/yyyy'
FMT_RATE = '#,##0.0000'
FMT_TEXT = '@'

# Format monétaire appliqué dynamiquement selon la devise choisie dans Paramètres.
CURRENCY_FORMATS = [
    ("XOF", '#,##0 "FCFA"'),
    ("XAF", '#,##0 "FCFA"'),
    ("EUR", '#,##0 "€"'),
    ("USD", '"$" #,##0'),
    ("GBP", '"£" #,##0'),
    ("CAD", '"C$" #,##0'),
    ("CHF", '#,##0 "CHF"'),
    ("MAD", '#,##0 "DH"'),
    ("NGN", '"₦" #,##0'),
    ("GHS", '"₵" #,##0'),
    ("CDF", '#,##0 "FC"'),
    ("ZAR", '"R" #,##0'),
    ("AED", '#,##0 "AED"'),
    ("MRU", '#,##0 "MRU"'),
    ("GNF", '#,##0 "GNF"'),
]

_NUMFMT_SEED = 200

# --------------------------------------------------------------------------
# Bordures
# --------------------------------------------------------------------------
def _side(color=LINE, style="thin"):
    return Side(style=style, color=color)


BORDER_CELL = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
BORDER_NONE = Border()
BORDER_TOP_GOLD = Border(top=_side(GOLD, "medium"))
BORDER_BOTTOM = Border(bottom=_side())


def box_border(color=LINE, style="thin"):
    s = _side(color, style)
    return Border(left=s, right=s, top=s, bottom=s)


# --------------------------------------------------------------------------
# Helpers de mise en forme
# --------------------------------------------------------------------------
def solid(color):
    return PatternFill("solid", fgColor=color)


def f(size=10, bold=False, color=INK, italic=False, name="Arial"):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def al(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)


def paint(ws, ref, fill=None, font=None, border=None, align=None, numfmt=None):
    """Applique un style homogène sur une plage 'A1:D10'."""
    for row in ws[ref]:
        for c in row:
            if fill is not None:
                c.fill = fill
            if font is not None:
                c.font = font
            if border is not None:
                c.border = border
            if align is not None:
                c.alignment = align
            if numfmt is not None:
                c.number_format = numfmt


def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


def heights(ws, mapping):
    for row, h in mapping.items():
        ws.row_dimensions[row].height = h


def setup_sheet(ws, tab_color=NAVY, zoom=90, freeze=None, hide_grid=True):
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = not hide_grid
    ws.sheet_view.zoomScale = zoom
    if freeze:
        ws.freeze_panes = freeze


def cell(ws, coord, value=None, fill=None, font=None, align=None,
         numfmt=None, border=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if fill is not None:
        c.fill = fill
    if font is not None:
        c.font = font
    if align is not None:
        c.alignment = align
    if numfmt is not None:
        c.number_format = numfmt
    if border is not None:
        c.border = border
    return c


def banner(ws, last_col, title, subtitle="", row=1, height=52, sub_height=20):
    """Bandeau de titre en tête de feuille."""
    ws.merge_cells(f"B{row}:{last_col}{row}")
    c = ws[f"B{row}"]
    c.value = title
    c.fill = solid(NAVY)
    c.font = f(19, True, WHITE)
    c.alignment = al("left", "center", indent=1)
    ws.row_dimensions[row].height = height
    paint(ws, f"B{row}:{last_col}{row}", fill=solid(NAVY))
    if subtitle:
        ws.merge_cells(f"B{row + 1}:{last_col}{row + 1}")
        s = ws[f"B{row + 1}"]
        s.value = subtitle
        s.font = f(10, False, WHITE)
        s.alignment = al("left", "center", indent=1)
        paint(ws, f"B{row + 1}:{last_col}{row + 1}", fill=solid(NAVY_MID))
        ws.row_dimensions[row + 1].height = sub_height


def section(ws, row, first_col, last_col, text, color=NAVY_MID, fg=WHITE, size=11):
    ws.merge_cells(f"{first_col}{row}:{last_col}{row}")
    c = ws[f"{first_col}{row}"]
    c.value = text
    c.font = f(size, True, fg)
    c.alignment = al("left", "center", indent=1)
    paint(ws, f"{first_col}{row}:{last_col}{row}", fill=solid(color))
    ws.row_dimensions[row].height = 24


def table_header(ws, row, first_col, last_col, labels, color=NAVY_SOFT, height=30):
    """Écrit une ligne d'en-têtes de tableau."""
    ci = ws[f"{first_col}1"].column
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=ci + i, value=label)
        c.fill = solid(color)
        c.font = f(9, True, WHITE)
        c.alignment = al("center", "center", wrap=True)
        c.border = box_border(NAVY_SOFT)
    ws.row_dimensions[row].height = height


def zebra(ws, ref, first_data_row, color=BG):
    """Alternance de lignes (formule) sur une plage de données."""
    dxf = DifferentialStyle(fill=PatternFill(bgColor=color))
    rule = Rule(type="expression", dxf=dxf, stopIfTrue=False)
    rule.formula = [f"MOD(ROW()-{first_data_row},2)=1"]
    ws.conditional_formatting.add(ref, rule)


def currency_format(ws, ref, hide_zero=False):
    """Affiche les montants avec le symbole de la devise choisie (Paramètres).

    `hide_zero` laisse les cellules à zéro visuellement vides : indispensable sur les
    grandes grilles, où une colonne de « 0 FCFA » noie les vrais montants.
    """
    global _NUMFMT_SEED
    for code, fmt in CURRENCY_FORMATS:
        _NUMFMT_SEED += 1
        if hide_zero:
            fmt = f"{fmt};-{fmt};"
        dxf = DifferentialStyle(numFmt=NumberFormat(numFmtId=_NUMFMT_SEED, formatCode=fmt))
        rule = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        rule.formula = [f'Devise_Base="{code}"']
        ws.conditional_formatting.add(ref, rule)


def hide_zeros(ws, ref):
    """Applique le format monétaire à zéro masqué sur une plage."""
    paint(ws, ref, numfmt=FMT_MONEY_BLANK)


def color_scale_ok(ws, ref):
    dxf_pos = DifferentialStyle(font=Font(color=GREEN, bold=True))
    r = Rule(type="cellIs", operator="greaterThan", formula=["0"], dxf=dxf_pos)
    ws.conditional_formatting.add(ref, r)
    dxf_neg = DifferentialStyle(font=Font(color=RED, bold=True))
    r2 = Rule(type="cellIs", operator="lessThan", formula=["0"], dxf=dxf_neg)
    ws.conditional_formatting.add(ref, r2)


def databar(ws, ref, color=BLUE):
    from openpyxl.formatting.rule import DataBarRule
    ws.conditional_formatting.add(
        ref,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color=color, showValue=True, minLength=None, maxLength=None),
    )


def link(ws, coord, label, target_sheet, target_cell="A1",
         fill_color=WHITE, font_color=NAVY, bold=True, size=11):
    c = ws[coord]
    c.value = label
    c.hyperlink = Hyperlink(ref=coord, location=f"'{target_sheet}'!{target_cell}")
    c.font = f(size, bold, font_color)
    c.fill = solid(fill_color)
    c.alignment = al("left", "center", indent=1)
    c.border = box_border(LINE)
    return c


def kpi_card(ws, row, col, label, formula, numfmt=FMT_MONEY, accent=BLUE,
             span=3, note=None, note_formula=None):
    """Carte indicateur : libellé, valeur, note. Occupe `span` colonnes, 3 lignes."""
    c1 = get_column_letter(col)
    c2 = get_column_letter(col + span - 1)
    # libellé
    ws.merge_cells(f"{c1}{row}:{c2}{row}")
    lab = ws[f"{c1}{row}"]
    lab.value = label
    lab.font = f(9, True, MUTED)
    lab.alignment = al("left", "center", indent=1)
    # valeur
    ws.merge_cells(f"{c1}{row + 1}:{c2}{row + 1}")
    val = ws[f"{c1}{row + 1}"]
    val.value = formula
    val.font = f(18, True, accent)
    val.alignment = al("left", "center", indent=1)
    val.number_format = numfmt
    # note
    ws.merge_cells(f"{c1}{row + 2}:{c2}{row + 2}")
    nt = ws[f"{c1}{row + 2}"]
    nt.value = note_formula if note_formula else (note or "")
    nt.font = f(8, False, MUTED)
    nt.alignment = al("left", "center", indent=1)

    paint(ws, f"{c1}{row}:{c2}{row + 2}", fill=solid(CARD))
    for r in range(row, row + 3):
        for cc in range(col, col + span):
            cur = ws.cell(row=r, column=cc)
            top = _side(LINE) if r == row else None
            bottom = _side(LINE) if r == row + 2 else None
            left = _side(LINE) if cc == col else None
            right = _side(LINE) if cc == col + span - 1 else None
            cur.border = Border(top=top, bottom=bottom, left=left, right=right)
    # liseré d'accent
    top_cell = ws.cell(row=row, column=col)
    top_cell.border = Border(top=_side(accent, "medium"), left=_side(LINE))
    for cc in range(col + 1, col + span):
        cur = ws.cell(row=row, column=cc)
        cur.border = Border(top=_side(accent, "medium"),
                            right=_side(LINE) if cc == col + span - 1 else None)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 30
    ws.row_dimensions[row + 2].height = 16
    return val


def note_box(ws, first_col, last_col, row, text, color=GOLD_SOFT, fg=INK, height=None):
    ws.merge_cells(f"{first_col}{row}:{last_col}{row}")
    c = ws[f"{first_col}{row}"]
    c.value = text
    c.font = f(9, False, fg, italic=True)
    c.alignment = al("left", "center", wrap=True, indent=1)
    paint(ws, f"{first_col}{row}:{last_col}{row}", fill=solid(color),
          border=box_border(GOLD_LINE))
    if height:
        ws.row_dimensions[row].height = height


def input_style(ws, ref):
    """Champs saisissables : fond jaune pâle + bordure."""
    paint(ws, ref, fill=solid("FFFBEA"), border=box_border("E3CE84"),
          font=f(10, False, INK))


def style_chart(chart, title, height=8.0, width=16.0):
    chart.title = title
    chart.height = height
    chart.width = width
    chart.style = 2
    if getattr(chart, "x_axis", None) is not None:
        chart.x_axis.delete = False
    if getattr(chart, "y_axis", None) is not None:
        chart.y_axis.delete = False
    return chart


# --------------------------------------------------------------------------
# Navigation — la même barre sur toutes les feuilles
# --------------------------------------------------------------------------
NAV_LINKS = [
    ("⌂  ACCUEIL", "Accueil"),
    ("①  MES INFOS", "Paramètres"),
    ("②  ESTIMATION", "Estimation Annuelle"),
    ("③  BUDGET SMART", "Journal"),
    ("TABLEAU DE BORD", "Tableau de Bord"),
    ("BUDGET 12 MOIS", "Budget 12 Mois"),
]


def nav_bar(ws, last_col, row=3, current=None):
    """Barre de navigation : un bouton par étape du parcours."""
    from openpyxl.utils import column_index_from_string
    last = column_index_from_string(last_col)
    span = max(2, (last - 1) // len(NAV_LINKS))
    col = 2
    for label, sheet in NAV_LINKS:
        end = min(col + span - 1, last)
        c1, c2 = get_column_letter(col), get_column_letter(end)
        ws.merge_cells(f"{c1}{row}:{c2}{row}")
        actif = (sheet == current)
        cell = ws[f"{c1}{row}"]
        cell.value = label
        if not actif:
            cell.hyperlink = Hyperlink(ref=f"{c1}{row}", location=f"'{sheet}'!A1")
        cell.font = f(9, True, WHITE if actif else NAVY)
        cell.alignment = al("center", "center")
        paint(ws, f"{c1}{row}:{c2}{row}",
              fill=solid(NAVY if actif else GOLD_SOFT),
              border=box_border(GOLD_LINE))
        col = end + 1
        if col > last:
            break
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 6


def page(ws, last_col, title, subtitle="", current=None, tab_color=NAVY,
         freeze=None, zoom=85):
    """Gabarit commun : bandeau, sous-titre, barre de navigation, respiration."""
    setup_sheet(ws, tab_color=tab_color, zoom=zoom, freeze=freeze)
    banner(ws, last_col, title, subtitle)
    nav_bar(ws, last_col, row=3, current=current)
    return 5   # première ligne de contenu


# --------------------------------------------------------------------------
# Gros boutons de la page d'accueil
# --------------------------------------------------------------------------
def big_button(ws, row, col, span, height_rows, label, sublabel, sheet,
               color, fg=WHITE):
    """Un pavé cliquable, lisible de loin."""
    c1, c2 = get_column_letter(col), get_column_letter(col + span - 1)
    r2 = row + height_rows - 1
    ws.merge_cells(f"{c1}{row}:{c2}{row + height_rows - 2}")
    main = ws[f"{c1}{row}"]
    main.value = label
    main.hyperlink = Hyperlink(ref=f"{c1}{row}", location=f"'{sheet}'!A1")
    main.font = f(16, True, fg)
    main.alignment = al("center", "center", wrap=True)
    ws.merge_cells(f"{c1}{r2}:{c2}{r2}")
    sub = ws[f"{c1}{r2}"]
    sub.value = sublabel
    sub.font = f(9, False, fg)
    sub.alignment = al("center", "center", wrap=True)
    paint(ws, f"{c1}{row}:{c2}{r2}", fill=solid(color), border=box_border(color))
    for r in range(row, r2 + 1):
        ws.row_dimensions[r].height = 22
    return main


# --------------------------------------------------------------------------
# Zone de graphiques — empilement automatique, jamais de superposition
# --------------------------------------------------------------------------
ROW_CM = 0.529   # hauteur d'une ligne par défaut, en centimètres


class ChartZone:
    """Empile les graphiques dans une colonne dédiée.

    La hauteur de chaque graphique détermine où commence le suivant : plus
    aucun graphique ne peut en recouvrir un autre, ni recouvrir un tableau.
    """

    def __init__(self, ws, col, row=5, gap=2):
        self.ws, self.col, self.row, self.gap = ws, col, row, gap

    def add(self, chart, title_row=None):
        if title_row:
            self.ws[f"{self.col}{self.row}"] = title_row
            self.ws[f"{self.col}{self.row}"].font = f(11, True, NAVY)
            self.row += 1
        self.ws.add_chart(chart, f"{self.col}{self.row}")
        self.row += math.ceil(chart.height / ROW_CM) + self.gap
        return chart


# --------------------------------------------------------------------------
# Lisibilité des textes DANS les graphiques
# --------------------------------------------------------------------------
def _chart_text(size=800, bold=False, color=INK):
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.text import (CharacterProperties, Font as DrawFont,
                                       Paragraph, ParagraphProperties,
                                       RichTextProperties)
    cp = CharacterProperties(sz=size, b=bold, solidFill=color,
                             latin=DrawFont(typeface="Arial"))
    return RichText(bodyPr=RichTextProperties(),
                    p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])


def tune_chart(chart, axis_size=750, legend="r", skip=None, no_legend=False,
               rotate_x=None):
    """Empêche les textes d'un graphique de se chevaucher.

    Petites polices sur les axes et la légende, légende sortie du tracé, et
    espacement des étiquettes d'axe quand les catégories sont nombreuses.
    """
    for name in ("x_axis", "y_axis"):
        axis = getattr(chart, name, None)
        if axis is None:
            continue
        axis.txPr = _chart_text(axis_size)
        axis.delete = False
    if no_legend:
        chart.legend = None
    elif chart.legend is not None:
        chart.legend.position = legend
        chart.legend.overlay = False
        chart.legend.txPr = _chart_text(axis_size)
    if skip and getattr(chart, "x_axis", None) is not None:
        chart.x_axis.tickLblSkip = skip
        chart.x_axis.tickMarkSkip = skip
    if rotate_x is not None and getattr(chart, "x_axis", None) is not None:
        chart.x_axis.txPr = _chart_text(axis_size)
        chart.x_axis.txPr.bodyPr.rot = rotate_x
    return chart


def percent_labels():
    """Étiquettes de secteur : le pourcentage SEUL.

    Activer showPercent sans éteindre le reste laisse Excel afficher aussi le
    nom de catégorie, le nom de série et la valeur : quatre textes empilés sur
    chaque part, illisibles. Tout est explicitement désactivé ici.
    """
    from openpyxl.chart.label import DataLabelList
    dl = DataLabelList()
    dl.showLegendKey = False
    dl.showVal = False
    dl.showCatName = False
    dl.showSerName = False
    dl.showPercent = True
    dl.showBubbleSize = False
    dl.separator = None
    dl.txPr = _chart_text(800, bold=True, color=WHITE)
    return dl
